#!/usr/bin/env python3
"""Rebuilds Summer_of_Success_Recap.xlsx -- the Summer of Success 2026
qualifier/amplify payout recap -- from this folder's sales.csv and goals.csv.

    python3 make_sos_recap.py

Four tabs, matching the workbook this replaces:

  Recap           one row per rep x supplier: qualifier goal vs actual CE,
                  status, tier earned, tier payout, amplify rollup, total.
  Program Setup   every constant the model runs on (tier payouts, rate
                  multipliers, the deck's tier CE thresholds) plus the rules
                  and the open questions. Blue cells are typed in; every
                  other tab is formulas pointing at them.
  Amplify Detail  one row per rep x supplier x amplify brand, since amplify
                  pays PER BRAND on cases above that brand's goal.
  Rep Totals      per-rep rollup of the Recap tab.

Only six kinds of cell are values: rep/manager/supplier/brand labels, the
qualifier goal, the actual CE figures, and the Program Setup constants.
Everything else is a live formula, so the workbook recalculates if a
constant is edited -- change a tier threshold on Program Setup and the whole
model follows.

HOW THE SOURCE MAPS IN
  sales.csv    RDE "2026 Summer of Success" export. Its Qualifier Brands
               column tags each row "Qualifier: ..." or "Amplify: ...":
                 Qualifier rows -> summed per rep+supplier into Recap!E
                                   (Actual CE).
                 Amplify rows   -> one Amplify Detail row each, carrying the
                                   2026 CE as actual and the 2025 CE as that
                                   brand's goal (amplify = growth over last
                                   year; see Program Setup note (a)).
  goals.csv    Per-rep qualifier goals, keyed on the rep plus the SAME brand
               text that appears after "Qualifier: " in sales.csv -- that is
               what ties a goal to a supplier, since goals.csv names brands
               and sales.csv names suppliers.

Two source quirks, both carried over from the workbook this replaces (see
Program Setup rules 6 and 7): Default / Office Tell Sell / Chris Politano
are excluded to match the dashboard, and supplier "Bell's" is rolled into
"New Belgium Brewing Company".

Managers come from incentive-tracking/index.html's DM_GROUPS, the repo's one
rep->DM roster. A rep missing from it gets a blank manager rather than being
dropped.

Run summer26/generate.py first if sales.csv has just been replaced, so the
live dashboard and this workbook are built from the same export.
"""
import csv
import collections
import datetime
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as gl

HERE = Path(__file__).parent
SALES = HERE / "sales.csv"
GOALS = HERE / "goals.csv"
DM_SOURCE = HERE.parent / "incentive-tracking" / "index.html"
OUT = HERE / "Summer_of_Success_Recap.xlsx"

SEASON = "6/1/2026–8/31/2026"
EXCLUDED_REPS = {"Default", "Office Tell Sell", "Chris Politano"}
SUPPLIER_ROLLUP = {"Bell's": "New Belgium Brewing Company"}

# Program constants -- the blue cells on Program Setup. Everything else in
# the workbook is a formula reading these.
TIER_PAYOUTS = [(1, 750), (2, 500), (3, 250)]
RATE_MULTIPLIERS = [("Qualified", 1), ("90% Partial Amplify", 0.5), ("Not Qualified", 0)]
TIER_THRESHOLDS = [
    # deck name,          name as it appears in the data,   T1,    T2,   T3
    ("Constellation Brands", "Constellation Brands",       25000, 6000, 1000),
    ("Molson Coors",         "MolsonCoors Beverage Company", 16000, 8000, 1000),
    ("Heineken (HUSA)",      "Heineken USA",                 8200, 2300,  100),
    ("Yuengling",            "DG Yuengling Inc",             2600, 1000,  100),
    ("Mike's (MABI)",        "Mark Anthony",                 4000, 1000,  500),
    ("Boston Beer co.",      "Boston Beer Company",          4000, 2000,  500),
]

NAVY, GREY, BLUE = "1F3864", "595959", "0000FF"
FONT = "Arial"
CE = "#,##0;(#,##0);-"
USD = "$#,##0;($#,##0);-"
USD2 = "$#,##0.00;($#,##0.00);-"


def style(cell, *, bold=False, size=10, color=None, bg=None, numfmt=None, wrap=False):
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if numfmt:
        cell.number_format = numfmt
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def sheet_head(ws, title, subtitle, size=15):
    style(ws.cell(1, 1, title), bold=True, size=size, color=NAVY)
    style(ws.cell(2, 1, subtitle), size=9, color=GREY)


def table_head(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        style(ws.cell(4, c, h), bold=True, color="FFFFFF", bg=NAVY)
        ws.column_dimensions[gl(c)].width = w
    ws.freeze_panes = "A5"


def load_managers():
    """rep -> DM, from incentive-tracking/index.html's DM_GROUPS roster."""
    text = DM_SOURCE.read_text()
    block = re.search(r"const DM_GROUPS = \[(.*?)\n\];", text, re.S)
    if not block:
        raise SystemExit(f"DM_GROUPS not found in {DM_SOURCE} -- has that roster moved? "
                         f"The recap reads it for the Manager column.")
    managers = {}
    for dm, reps in re.findall(r"\{dm:'([^']+)', reps:\[(.*?)\]\}", block.group(1), re.S):
        # Names are single- OR double-quoted in that file, and one of them
        # ("John O'Donoghue") carries an apostrophe -- so match each quote
        # style against its own delimiter rather than one "any quote" class.
        for single, double in re.findall(r"'([^']*)'|\"([^\"]*)\"", reps):
            managers[single or double] = dm
    return managers


def num(raw):
    raw = (raw or "").strip()
    return float(raw) if raw else 0.0


def load_source():
    with open(SALES, newline="") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        raise SystemExit(f"{SALES} is empty")
    cols = rows[0].keys()
    try:
        c25 = next(c for c in cols if c.startswith("Case Equiv") and "2025" in c)
        c26 = next(c for c in cols if c.startswith("Case Equiv") and "2026" in c)
    except StopIteration:
        raise SystemExit(f"{SALES}: could not find the 2025/2026 'Case Equiv' columns "
                         f"in {list(cols)}")

    qualifier, amplify = [], []
    for r in rows:
        rep = r["Sales Rep Assigned"].strip()
        if not rep or rep in EXCLUDED_REPS:
            continue
        tag = r["Qualifier Brands"].strip()
        supplier = r["Supplier"].strip()
        supplier = SUPPLIER_ROLLUP.get(supplier, supplier)
        rec = {"rep": rep, "supplier": supplier, "brand": r["Brand Family"].strip(),
               "ce26": num(r[c26]), "ce25": num(r[c25]),
               "tag": tag.split(": ", 1)[1] if ": " in tag else ""}
        if tag.startswith("Qualifier"):
            qualifier.append(rec)
        elif tag.startswith("Amplify"):
            amplify.append(rec)
    return qualifier, amplify


def load_goals(qualifier):
    """(rep, supplier) -> qualifier goal.

    goals.csv names BRANDS ("Modelo/Corona"); sales.csv names suppliers. The
    qualifier rows carry both -- their tag is the same brand text -- so they
    are what ties one to the other.
    """
    brand_to_supplier = {(q["rep"], q["tag"]): q["supplier"] for q in qualifier}
    goals, unmapped = {}, []
    with open(GOALS, newline="") as f:
        for g in csv.DictReader(f):
            if g["Type"].strip() != "Qualifier":
                continue
            rep = g["Sales Rep"].strip()
            if rep in EXCLUDED_REPS:
                continue
            supplier = brand_to_supplier.get((rep, g["Brand(s)"].strip()))
            if supplier is None:
                unmapped.append((rep, g["Brand(s)"].strip()))
                continue
            goals[(rep, supplier)] = num(g["2026 Goal"])
    return goals, unmapped


def total_row(ws, row, label_col, cols, fmts):
    """Bold TOTAL line under a table, summing each named column."""
    style(ws.cell(row, label_col, "TOTAL"), bold=True)
    for c in cols:
        style(ws.cell(row, c, f"=SUM({gl(c)}5:{gl(c)}{row - 1})"), bold=True,
              numfmt=fmts.get(c))


def build_program_setup(wb, notes):
    ws = wb.create_sheet("Program Setup")
    sheet_head(ws, "Summer of Success 2026 — program rules used in this recap",
               f"Season {SEASON}. Blue = typed in from the deck / source data. "
               f"Everything on the other tabs is a formula off these cells.")
    for col, w in zip("ABCDE", (30, 40, 14, 14, 14)):
        ws.column_dimensions[col].width = w

    style(ws.cell(4, 1, "Tier payout amounts"), bold=True)
    for c, h in enumerate(("Tier", "Payout"), 1):
        style(ws.cell(5, c, h), bold=True, color="FFFFFF", bg=NAVY)
    for i, (tier, payout) in enumerate(TIER_PAYOUTS):
        style(ws.cell(6 + i, 1, tier), color=BLUE)
        style(ws.cell(6 + i, 2, payout), color=BLUE, numfmt=USD)

    style(ws.cell(10, 1, "Amplify rate multiplier by qualifier result"), bold=True)
    for c, h in enumerate(("Qualifier result", "Amplify rate multiplier"), 1):
        style(ws.cell(11, c, h), bold=True, color="FFFFFF", bg=NAVY)
    for i, (label, mult) in enumerate(RATE_MULTIPLIERS):
        style(ws.cell(12 + i, 1, label))
        style(ws.cell(12 + i, 2, mult), color=BLUE, numfmt="0%")

    style(ws.cell(16, 1, "Tier CE thresholds by supplier — the deck's “Rewards Payouts — "
                         "by Route Volume” grid"), bold=True)
    for c, h in enumerate(("Supplier (deck)", "Supplier (as named in the data)",
                           "Tier 1 CE", "Tier 2 CE", "Tier 3 CE"), 1):
        style(ws.cell(17, c, h), bold=True, color="FFFFFF", bg=NAVY)
    for i, (deck, data_name, t1, t2, t3) in enumerate(TIER_THRESHOLDS):
        r = 18 + i
        style(ws.cell(r, 1, deck))
        style(ws.cell(r, 2, data_name))
        for c, v in zip((3, 4, 5), (t1, t2, t3)):
            style(ws.cell(r, c, v), color=BLUE, numfmt=CE)

    style(ws.cell(26, 1, "Rules applied"), bold=True)
    for i, line in enumerate(notes["rules"]):
        style(ws.cell(27 + i, 1, line), size=9, color=GREY)
    warn_row = 27 + len(notes["rules"]) + 1
    style(ws.cell(warn_row, 1, "⚠ FOUR THINGS TO CONFIRM BEFORE THIS PAYS ANYONE"),
          bold=True, color="C00000")
    for i, line in enumerate(notes["confirm"]):
        style(ws.cell(warn_row + 1 + i, 1, line), size=9, color=GREY)
    ws.freeze_panes = "A18"
    return ws


def main():
    qualifier, amplify = load_source()
    goals, unmapped = load_goals(qualifier)
    managers = load_managers()

    # Recap rows: every rep+supplier appearing on either side of the program.
    actual = collections.Counter()
    for q in qualifier:
        actual[(q["rep"], q["supplier"])] += q["ce26"]
    keys = sorted({(r["rep"], r["supplier"]) for r in qualifier + amplify})

    amp_rows = sorted(
        ((a["rep"], a["supplier"], a["brand"], a["ce26"], a["ce25"]) for a in amplify),
        key=lambda t: (t[0], t[1], t[2]))
    last_recap = 4 + len(keys)
    last_amp = 4 + len(amp_rows)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- Recap --------------------------------------------------------
    ws = wb.create_sheet("Recap")
    sheet_head(ws, "Summer of Success 2026 — Recap by rep and supplier",
               "Filter any column from row 4. Status drives everything: Qualified = tier "
               "payout + full Amplify rate · 90% Partial = half Amplify rate, no tier · "
               "Not Qualified = nothing.", size=16)
    table_head(ws, ["Rep", "Manager", "Supplier", "Qualifier goal", "Actual CE", "% to goal",
                    "Status", "Tier earned", "Tier payout", "Amplify actual CE",
                    "Amplify goal CE", "Eligible amplify CE", "Amplify payout",
                    "Total payout", "Rate multiplier", "Key"],
               [20, 16, 30, 14, 13, 11, 21, 11, 13, 15, 15, 16, 14, 14, 13, 34])
    A, P = f"$A$5:$A${last_amp}", f"$C$5:$C${last_amp}"
    for i, (rep, supplier) in enumerate(keys):
        r = 5 + i
        style(ws.cell(r, 1, rep))
        style(ws.cell(r, 2, managers.get(rep, "")))
        style(ws.cell(r, 3, supplier))
        goal = goals.get((rep, supplier))
        style(ws.cell(r, 4, goal if goal else None), color=BLUE, numfmt=CE)
        # A rep+supplier with qualifier rows that net to 0 still shows 0 --
        # blank means "no qualifier rows at all", which is a different thing.
        has_qual = (rep, supplier) in actual
        style(ws.cell(r, 5, actual[(rep, supplier)] if has_qual else None),
              color=BLUE, numfmt=CE)
        f = {
            6: f'=IF($D{r}>0,$E{r}/$D{r},"")',
            7: (f'=IF($D{r}>0,IF($F{r}>=1,"Qualified",IF($F{r}>=0.9,'
                f'"90% Partial Amplify","Not Qualified")),"No qualifier - Amplify open")'),
            8: (f'=IF($G{r}<>"Qualified","-",IFERROR(IF($E{r}>=INDEX('
                f"'Program Setup'!$C$18:$C$23,MATCH($C{r},'Program Setup'!$B$18:$B$23,0)),1,"
                f"IF($E{r}>=INDEX('Program Setup'!$D$18:$D$23,MATCH($C{r},"
                f"'Program Setup'!$B$18:$B$23,0)),2,IF($E{r}>=INDEX('Program Setup'!"
                f"$E$18:$E$23,MATCH($C{r},'Program Setup'!$B$18:$B$23,0)),3,\"-\"))),\"-\"))"),
            9: (f"=IF(ISNUMBER($H{r}),INDEX('Program Setup'!$B$6:$B$8,MATCH($H{r},"
                f"'Program Setup'!$A$6:$A$8,0)),0)"),
            10: f"=SUMIFS('Amplify Detail'!$E$5:$E${last_amp},'Amplify Detail'!{A},$A{r},"
                f"'Amplify Detail'!{P},$C{r})",
            11: f"=SUMIFS('Amplify Detail'!$F$5:$F${last_amp},'Amplify Detail'!{A},$A{r},"
                f"'Amplify Detail'!{P},$C{r})",
            12: f"=SUMIFS('Amplify Detail'!$G$5:$G${last_amp},'Amplify Detail'!{A},$A{r},"
                f"'Amplify Detail'!{P},$C{r})",
            13: f"=SUMIFS('Amplify Detail'!$J$5:$J${last_amp},'Amplify Detail'!{A},$A{r},"
                f"'Amplify Detail'!{P},$C{r})",
            14: f"=$I{r}+$M{r}",
            15: (f'=IF($G{r}="No qualifier - Amplify open",1,IFERROR(INDEX('
                 f"'Program Setup'!$B$12:$B$14,MATCH($G{r},'Program Setup'!$A$12:$A$14,0)),0))"),
            16: f'=$A{r}&"|"&$C{r}',
        }
        fmt = {6: "0.0%", 9: USD, 10: CE, 11: CE, 12: CE, 13: USD, 14: USD, 15: "0%"}
        for c, formula in f.items():
            style(ws.cell(r, c, formula), numfmt=fmt.get(c))
    ws.auto_filter.ref = f"A4:O{last_recap}"
    total_row(ws, last_recap + 1, 1, (9, 10, 11, 12, 13, 14),
              {9: USD, 10: CE, 11: CE, 12: CE, 13: USD, 14: USD})

    # ---- Program Setup ------------------------------------------------
    build_program_setup(wb, program_notes(keys, goals, actual, amplify))

    # ---- Amplify Detail -----------------------------------------------
    ws = wb.create_sheet("Amplify Detail")
    sheet_head(ws, "Amplify detail — brand level",
               "Amplify pays ONLY on cases above the brand's goal: Eligible = "
               "MAX(Actual − Goal, 0), per brand. A brand below goal pays $0 and does not "
               "offset others.")
    table_head(ws, ["Rep", "Manager", "Supplier", "Amplify brand", "Actual CE (2026)",
                    "Amplify goal CE (2025)", "Eligible CE (above goal)", "Normal rate",
                    "Applied rate", "Payout", "Key"],
               [20, 16, 30, 26, 15, 16, 16, 12, 12, 14, 34])
    for i, (rep, supplier, brand, ce26, ce25) in enumerate(amp_rows):
        r = 5 + i
        style(ws.cell(r, 1, rep))
        style(ws.cell(r, 2, managers.get(rep, "")))
        style(ws.cell(r, 3, supplier))
        style(ws.cell(r, 4, brand))
        style(ws.cell(r, 5, ce26), color=BLUE, numfmt=CE)
        style(ws.cell(r, 6, ce25), color=BLUE, numfmt=CE)
        style(ws.cell(r, 7, f"=MAX(0,$E{r}-$F{r})"), numfmt=CE)
        # $2/case brands, per the deck: Fever Tree, Mike's Harder, MXD.
        style(ws.cell(r, 8, f'=IF(OR(ISNUMBER(SEARCH("fever",$D{r})),'
                            f'ISNUMBER(SEARCH("mike",$D{r})),'
                            f'ISNUMBER(SEARCH("mxd",$D{r}))),2,1)'), numfmt=USD2)
        style(ws.cell(r, 9, f"=$H{r}*IFERROR(INDEX('Recap'!$O$5:$O${last_recap},"
                            f"MATCH($K{r},'Recap'!$P$5:$P${last_recap},0)),0)"), numfmt=USD2)
        style(ws.cell(r, 10, f"=$G{r}*$I{r}"), numfmt=USD)
        style(ws.cell(r, 11, f'=$A{r}&"|"&$C{r}'))
    ws.auto_filter.ref = f"A4:J{last_amp}"
    total_row(ws, last_amp + 1, 1, (7, 10), {7: CE, 10: USD})

    # ---- Rep Totals ---------------------------------------------------
    ws = wb.create_sheet("Rep Totals")
    sheet_head(ws, "Total payout by rep", "Rolls up the Recap tab.")
    table_head(ws, ["Rep", "Manager", "Suppliers with a goal", "Qualified", "90% partial",
                    "Eligible amplify CE", "Tier payout", "Amplify payout", "Total payout"],
               [20, 16, 20, 12, 13, 17, 14, 15, 15])
    reps = sorted({rep for rep, _ in keys})
    RA = f"'Recap'!$A$5:$A${last_recap}"
    for i, rep in enumerate(reps):
        r = 5 + i
        style(ws.cell(r, 1, rep))
        style(ws.cell(r, 2, managers.get(rep, "")))
        style(ws.cell(r, 3, f'=COUNTIFS({RA},$A{r},'
                            f"'Recap'!$D$5:$D${last_recap},\">0\")"), numfmt=CE)
        style(ws.cell(r, 4, f'=COUNTIFS({RA},$A{r},'
                            f"'Recap'!$G$5:$G${last_recap},\"Qualified\")"), numfmt=CE)
        style(ws.cell(r, 5, f'=COUNTIFS({RA},$A{r},'
                            f"'Recap'!$G$5:$G${last_recap},\"90% Partial Amplify\")"),
              numfmt=CE)
        style(ws.cell(r, 6, f"=SUMIFS('Recap'!$L$5:$L${last_recap},{RA},$A{r})"), numfmt=CE)
        style(ws.cell(r, 7, f"=SUMIFS('Recap'!$I$5:$I${last_recap},{RA},$A{r})"), numfmt=USD)
        style(ws.cell(r, 8, f"=SUMIFS('Recap'!$M$5:$M${last_recap},{RA},$A{r})"), numfmt=USD)
        style(ws.cell(r, 9, f"=$G{r}+$H{r}"), numfmt=USD)
    last_rep = 4 + len(reps)
    ws.auto_filter.ref = f"A4:I{last_rep}"
    total_row(ws, last_rep + 1, 1, (4, 5, 6, 7, 8, 9),
              {4: CE, 5: CE, 6: CE, 7: USD, 8: USD, 9: USD})

    wb.save(OUT)

    no_goal = sum(1 for k in keys if not goals.get(k) and actual.get(k))
    print(f"Wrote {OUT.name}  (source: {SALES.name}, rebuilt {datetime.date.today():%m/%d/%Y})")
    print(f"  Recap {len(keys)} rep+supplier rows · Amplify Detail {len(amp_rows)} brand rows "
          f"· Rep Totals {len(reps)} reps")
    print(f"  {sum(1 for k in keys if goals.get(k))} rows carry a qualifier goal; "
          f"{no_goal} have qualifier CE but no goal (Status 'No qualifier - Amplify open')")
    if unmapped:
        print(f"  NOTE: {len(unmapped)} goals.csv qualifier rows had no matching supplier in "
              f"sales.csv and were skipped, e.g. {unmapped[:3]}")


def program_notes(keys, goals, actual, amplify):
    """The Rules / Confirm blocks, with every figure computed from THIS export.

    The workbook this replaces hard-coded its worked examples and counts
    ("11 rep/supplier rows", "196 of 220"). On a payout document those go
    stale silently the moment the data is refreshed, so they are derived.
    """
    tiers = {name: (t1, t2, t3) for _, name, t1, t2, t3 in TIER_THRESHOLDS}

    def worked_example():
        """A qualified rep+supplier with amplify brands, to illustrate rules 2 and 3."""
        best = None
        for rep, supplier in keys:
            goal = goals.get((rep, supplier))
            act = actual.get((rep, supplier), 0)
            if not goal or act < goal or supplier not in tiers:
                continue
            brands = [a for a in amplify
                      if a["rep"] == rep and a["supplier"] == supplier
                      and a["ce26"] > a["ce25"]]
            if len(brands) >= 2 and (best is None or act > best[2]):
                best = (rep, supplier, act, goal, brands)
        return best

    ex = worked_example()
    if ex:
        rep, supplier, act, goal, brands = ex
        t1, t2, t3 = tiers[supplier]
        tier, line, pay = ((1, t1, 750) if act >= t1 else
                           (2, t2, 500) if act >= t2 else
                           (3, t3, 250) if act >= t3 else (None, None, 0))
        rule2 = (f"     Example: {rep}, {supplier}, goal {goal:,.0f}, actual {act:,.0f} = "
                 f"{act / goal:.0%} → Qualified"
                 + (f"; {act:,.0f} clears the {line:,.0f} Tier {tier} line → ${pay}."
                    if tier else "; no tier line cleared → $0."))
        parts, tot = [], 0.0
        for b in sorted(brands, key=lambda b: -(b["ce26"] - b["ce25"]))[:3]:
            rate = 2 if re.search(r"fever|mike|mxd", b["brand"], re.I) else 1
            elig = b["ce26"] - b["ce25"]
            tot += elig * rate
            parts.append(f"{b['brand']} {b['ce26']:,.0f} − {b['ce25']:,.0f} = "
                         f"{elig:,.0f} × ${rate} = ${elig * rate:,.0f}")
        rule3 = f"     Example: {rep}, {supplier} — " + "; ".join(parts) + \
                f". Total ${tot:,.0f}."
    else:
        rule2 = "     (No qualified rep+supplier with a tier grid in this export.)"
        rule3 = "     (No qualified rep+supplier with amplify growth in this export.)"

    # (a) how often goals.csv's supplier-level Amplify goal equals the sum of
    #     the brand 2025 figures this recap uses instead.
    brand_2025 = collections.Counter()
    for a in amplify:
        brand_2025[(a["rep"], a["supplier"])] += a["ce25"]
    # goals.csv names amplify goals by the same brand-group text that follows
    # "Amplify: " in sales.csv, so the export's own rows are the join.
    amp_tag_supplier = {(a["rep"], a["tag"]): a["supplier"] for a in amplify}
    supplier_goal = {}
    with open(GOALS, newline="") as f:
        for g in csv.DictReader(f):
            if g["Type"].strip() != "Amplify":
                continue
            rep = g["Sales Rep"].strip()
            if rep in EXCLUDED_REPS:
                continue
            supplier = amp_tag_supplier.get((rep, g["Brand(s)"].strip()))
            if supplier:
                supplier_goal[(rep, supplier)] = (supplier_goal.get((rep, supplier), 0)
                                                  + num(g["2026 Goal"]))
    compared = [(k, supplier_goal[k], brand_2025[k]) for k in supplier_goal]
    agree = sum(1 for _, a, b in compared if abs(a - b) < 0.5)
    worst = sorted(compared, key=lambda t: -abs(t[1] - t[2]))[:2]
    worst_txt = "; ".join(f"{k[0]} / {k[1]} (goal {a:,.0f} vs {b:,.0f} of 2025 CE)"
                          for k, a, b in worst) or "none"

    # (d) qualifier volume with no goal
    no_goal = [(rep, sup, actual[(rep, sup)]) for rep, sup in keys
               if not goals.get((rep, sup)) and actual.get((rep, sup))]
    no_goal.sort(key=lambda t: -t[2])
    biggest = (f"     The largest is {no_goal[0][0]}, {no_goal[0][1]}, "
               f"{no_goal[0][2]:,.0f} CE. Filter Status for that label to see them. If "
               f"those reps were meant to have goals, their tier payout is being missed."
               if no_goal else "     None in this export.")

    return {"rules": [
        "1. QUALIFIER — a rep clears a supplier when their 2026 qualifier CE reaches their "
        "own personal goal (goals are per rep, not flat).",
        "2. TIER PAYOUT — only at 100%+ of goal. The tier is then read off the grid above "
        "using that same actual 2026 qualifier CE.",
        rule2,
        "3. AMPLIFY — pays ONLY on cases ABOVE the brand's Amplify goal:  Eligible CE = "
        "MAX(Actual CE − Amplify goal CE, 0), calculated PER BRAND.",
        "     A brand below its goal pays $0 and does NOT offset a brand that is above goal.",
        rule3,
        "4. AMPLIFY RATE — $1/case, except Fever Tree, Mike's Harder and MXD at $2/case.",
        "5. 90% PARTIAL — 90.00%–99.99% of the qualifier goal earns NO tier payout but HALF "
        "the Amplify rate ($0.50 / $1.00). Below 90% earns no Amplify at all.",
        "6. Suppliers with no qualifier in the data (Sapporo, Cape May, New Belgium) have no "
        "tier payout; their Amplify is open at full rate.",
        "7. Excluded, matching the dashboard: Default, Office Tell Sell, Chris Politano. "
        "Bell's is rolled into New Belgium Brewing Company.",
    ], "confirm": [
        "a) AMPLIFY GOAL SOURCE — payout is per brand, and the only per-brand figure in the "
        "source is that brand's 2025 CE, so 2025 CE is used as each brand's Amplify goal.",
        "     That makes Amplify = growth over last year, which matches the live summer26 "
        "dashboard.",
        f"     goals.csv carries a SUPPLIER-level Amplify goal too; it equals the sum of the "
        f"brand 2025 figures for {agree} of {len(compared)} rep/supplier rows. The rest "
        f"differ, mostly by ±1 (rounding).",
        f"     The largest disagreements: {worst_txt}.",
        "b) TIER BASIS — the deck writes the thresholds as “+25,000 CE's” / “+16,000 CE's”. "
        "The “+” may mean CE GROWTH over 2025 rather than total 2026 CE.",
        "     This recap uses total 2026 CE for the tier, per the worked example above. On a "
        "growth reading almost no one clears Tier 1.",
        "c) PER REP vs HOUSE — summer26/README.txt describes this same grid as a company-wide "
        "aggregate across all reps combined. This recap applies it per rep, as instructed.",
        f"d) QUALIFIER VOLUME WITH NO GOAL — {len(no_goal)} rep/supplier rows have qualifier "
        f"CE but no goal in goals.csv, so they cannot qualify and fall to “No qualifier - "
        f"Amplify open”.",
        biggest,
    ]}


if __name__ == "__main__":
    main()
