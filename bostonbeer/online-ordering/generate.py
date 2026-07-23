#!/usr/bin/env python3
"""
Rebuilds data/forecast.json for the Boston Beer Online Ordering pulse-check
page. This is deliberately a simple join, not a projection engine: for each
SKU it lines up Boston Beer's own forecast (F) and the confirmed order (A)
per week against Encompass's actual sell-through/inventory numbers and
Boston Beer's own forward-looking system forecast, with no invented math on
top (no trend line, no seasonal fallback, no recommended quantity). The page
itself decides whether to draw the eye to a week where F looks out of line
with recent sales -- this script just gets the raw numbers lined up.

Three independent inputs, each refreshable on its own:
  076KOH_Forecasts.csv        Boston Beer portal forecast export -- Product,
                               Distributor, BBC SKU, Customer SKU, then a
                               pair of columns per week ("<date> F" / "<date> A").
  Boston_Beer_Inventory_Report.csv   Encompass "Boston Beer Inventory Report"
                               export -- Available, the last 8 weeks of actual
                               case sales (Cases -6 Weeks through Cases This
                               Week), Last Receive Date/Quantity, and a
                               "Product Link" column ("<Customer SKU> <clean
                               product name>") that's the only source of the
                               cleaned-up product name now (Encompass's
                               separate "Products" export is no longer needed
                               for that).
  ForecastReport.xlsx         Encompass "ForecastReport" export -- Boston
                               Beer's own forward-looking system forecast,
                               in case-equivalents, one column per week
                               several months out. This is what lets you
                               gauge whether a product's sales look set to
                               pick up (seasonal ramp) or keep climbing
                               (growing brand) beyond what recent weeks show.

Encompass's inventory SKU sometimes carries a suffix the Boston Beer portal
SKU doesn't (e.g. portal "AJ0153" vs Encompass "AJ0153A1") -- both the
Inventory Report and ForecastReport use the same suffixed codes, so both are
matched to the portal SKU by exact match first, then by prefix.

L4 Avg is computed here (not read from Encompass -- this export doesn't
carry a pre-computed average column) as the mean of the last 4 *complete*
weeks (Cases -3 Weeks through Cases Last Week). Cases This Week is excluded
from the average since it's a partial, still-in-progress week that would
understate it -- it's still shown in the Recent Sales detail, just labeled
as in progress.

Also builds data/trend_forecast.json for the "Trend Forecast" tab, from two
more (optional) inputs -- see the docstring on build_trend_forecast() below
for the full method.

Run: python3 generate.py [forecast_csv] [inventory_csv] [forecast_report_xlsx] [l13_csv] [ly_by_week_csv]
(all five default to the filenames committed in this folder; the trend
forecast is skipped with a message if l13_csv/ly_by_week_csv aren't found)
"""
import csv
import json
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
DEFAULT_FORECAST_CSV = HERE / "076KOH_Forecasts.csv"
DEFAULT_INVENTORY_CSV = HERE / "Boston_Beer_Inventory_Report.csv"
DEFAULT_FORECAST_REPORT_XLSX = HERE / "ForecastReport.xlsx"
DEFAULT_L13_CSV = HERE / "L13_Trend.csv"
DEFAULT_LY_BY_WEEK_CSV = HERE / "LY_By_Week.csv"
OUT_JSON = HERE / "data" / "forecast.json"
TREND_OUT_JSON = HERE / "data" / "trend_forecast.json"

WEEK_COL_RE = re.compile(r"^(\d{2}/\d{2}/\d{4}) F$")
PRODUCT_LINK_RE = re.compile(r"^\s*\S+\s+(.*\S)\s*$")
DATE_HEADER_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
L13_WINDOW_COL_RE = re.compile(r"^Cases (\d{1,2}/\d{1,2}/\d{4}) - (\d{1,2}/\d{1,2}/\d{4})$")
LY_WEEK_COL_RE = re.compile(r"^Cases (\d{4}) (\d{2})$")

# Chronological order, oldest to newest. Index 6 (Last Week) is the newest
# *complete* week; index 7 (This Week) is still in progress.
RECENT_WEEK_COLS = [
    "Cases -6 Weeks", "Cases -5 Weeks", "Cases -4 Weeks", "Cases -3 Weeks",
    "Cases -2 Weeks", "Cases -1 Weeks", "Cases Last Week", "Cases This Week",
]
RECENT_WEEK_LABELS = ["-6 wk", "-5 wk", "-4 wk", "-3 wk", "-2 wk", "-1 wk", "Last wk", "This wk"]


def to_num(raw):
    raw = (raw or "").strip()
    if raw == "":
        return None
    try:
        return int(raw)
    except ValueError:
        return float(raw)


def resolve(sku, lookup):
    """Exact match first, else the one key that starts with sku (Encompass's
    suffixed inventory-SKU variants)."""
    if sku in lookup:
        return lookup[sku]
    for k, v in lookup.items():
        if k.startswith(sku):
            return v
    return None


def load_inventory(path):
    with open(path, newline="") as f:
        rows = list(csv.DictReader(f))
    inv = {}
    for r in rows:
        sku = r["Supplier Product ID"].strip()
        if not sku:
            continue
        weeks = [to_num(r[c]) for c in RECENT_WEEK_COLS]
        complete_weeks = [w for w in weeks[:7] if w is not None]  # excludes This Week
        last4_complete = [w for w in weeks[3:7] if w is not None]
        m = PRODUCT_LINK_RE.match(r.get("Product Link") or "")
        inv[sku] = {
            "brandFamily": (r.get("Brand Family") or "").strip() or None,
            "package": (r.get("Package") or "").strip() or None,
            "cleanName": m.group(1).strip() if m else None,
            "available": to_num(r.get("Available")),
            "recentWeeks": weeks,
            "l4Avg": round(sum(last4_complete) / len(last4_complete), 1) if last4_complete else None,
            "l8Avg": round(sum(complete_weeks) / len(complete_weeks), 1) if complete_weeks else None,
            "lastReceiveDate": (r.get("Last Receive Date") or "").strip() or None,
            "lastReceiveQty": to_num(r.get("Last Receive Quantity")),
        }
    return inv


def load_forecast_report(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    date_cols = [(i, str(h)) for i, h in enumerate(header, start=1) if h and DATE_HEADER_RE.match(str(h))]
    if not date_cols:
        raise SystemExit(f"{path}: could not find any date columns in row 1")

    out = {}
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(row=r, column=1).value
        product = ws.cell(row=r, column=3).value
        if not sku or not product:
            continue
        sku = str(sku).strip()
        weeks = []
        for col, wd in date_cols:
            v = ws.cell(row=r, column=col).value
            weeks.append({"weekEnding": wd, "cases": round(v, 1) if isinstance(v, (int, float)) else None})
        out[sku] = weeks
    return out


def to_num_rde(raw):
    """L13_Trend.csv / LY_By_Week.csv are RDE exports: comma thousands
    separators and parenthesized negatives (e.g. " (2,448.00)"), unlike the
    Boston Beer portal/Encompass CSVs to_num() above handles."""
    raw = (raw or "").strip()
    if raw == "":
        return None
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()").replace(",", "")
    if raw == "":
        return None
    val = float(raw)
    return -val if neg else val


def parse_mdy(s):
    m, d, y = s.split("/")
    return date(int(y), int(m), int(d))


def week1_start(year):
    """The Sunday on or before Jan 1 of `year` -- the anchor of the retail
    week-numbering convention L13_Trend.csv / LY_By_Week.csv use (verified
    below, not assumed)."""
    jan1 = date(year, 1, 1)
    offset = (jan1.weekday() + 1) % 7  # Mon=0..Sun=6 -> days back to that Sunday
    return jan1 - timedelta(days=offset)


def week_number(d):
    """(year, week_num) for date d under the week1_start() convention -- week
    N runs Sunday..Saturday, N=1 being the week containing (or ending at)
    week1_start(d.year). Verified against L13_Trend.csv's own stated
    windows: this reproduces 4/20/2025 as week 17's start and 7/19/2025 as
    week 29's end exactly, for both the 2025 and 2026 windows the file
    ships with -- that agreement (not a hardcoded week number) is why this
    rule is trusted to project forward to arbitrary future dates too."""
    year = d.year
    start = week1_start(year)
    if d < start:
        year -= 1
        start = week1_start(year)
    return year, (d - start).days // 7 + 1


def load_l13_trend(path):
    with open(path, newline="") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        raise SystemExit(f"{path}: no data rows")
    windows = []
    for h in rows[0].keys():
        m = L13_WINDOW_COL_RE.match(h)
        if m:
            windows.append((parse_mdy(m.group(1)), parse_mdy(m.group(2)), h))
    if len(windows) != 2:
        raise SystemExit(f"{path}: expected exactly 2 'Cases <start> - <end>' window columns, found {len(windows)}")
    windows.sort(key=lambda w: w[1])  # earlier end date = last year's window
    ly_start, ly_end, ly_col = windows[0]
    ty_start, ty_end, ty_col = windows[1]

    out = {}
    for r in rows:
        sku = (r.get("Supplier Product ID") or "").strip()
        product = (r.get("Product Name") or "").strip()
        if not sku or not product:
            continue
        out[sku] = {"product": product, "lyCases": to_num_rde(r[ly_col]) or 0, "tyCases": to_num_rde(r[ty_col]) or 0}
    windows_out = {"lyStart": ly_start, "lyEnd": ly_end, "tyStart": ty_start, "tyEnd": ty_end}
    return out, windows_out


def load_ly_by_week(path):
    with open(path, newline="") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        raise SystemExit(f"{path}: no data rows")
    week_cols = [(int(m.group(1)), int(m.group(2)), h) for h in rows[0].keys() if (m := LY_WEEK_COL_RE.match(h))]
    if not week_cols:
        raise SystemExit(f"{path}: no 'Cases <year> <NN>' week columns found")

    out = {}
    for r in rows:
        sku = (r.get("Supplier Product ID") or "").strip()
        if not sku:
            continue
        weeks = {}
        for yr, wn, col in week_cols:
            v = to_num_rde(r[col])
            if v is not None:
                weeks[(yr, wn)] = v
        out[sku] = weeks
    return out


def build_trend_forecast(forecast_rows, week_dates, l13_lookup, l13_windows, ly_lookup, inv_lookup):
    """For each SKU on the Boston Beer forecast (same driver list as the
    Pulse Check tab, matched into L13/LY by the same resolve() used for
    Inventory/ForecastReport), projects the next up-to-8 weeks two ways and
    lines them up:

      trend forecast = last year's actual cases for that same week (by
        retail week number, see week_number() above) x (1 + that SKU's
        13-week YoY case trend from L13_Trend.csv), floored at 0.
      bbc forecast = Boston Beer's own "<date> F" value for that week,
        already in forecast_rows.

    A SKU with no positive last-year case volume in the L13 window (new
    item, no trend to project) or no matching LY_By_Week row (no weekly
    breakdown to scale) gets trendForecast=None for every week rather than
    a fabricated number. diffPct/flagged (>10% apart) are only set when
    both numbers exist and the trend forecast is positive.

    Each product also carries "available" and "lastReceiveDate" (current
    on-hand cases and last delivery date, from the same Inventory Report
    lookup the Pulse Check tab uses) so the Trend Forecast tab can show
    current inventory alongside the trend, without a second network fetch.
    """
    fc_weeks = week_dates[:8]
    ly_year = l13_windows["lyEnd"].year
    week_meta = []
    for wd in fc_weeks:
        _, wn = week_number(parse_mdy(wd))
        week_meta.append({"weekEnding": wd, "lyYear": ly_year, "lyWeekNum": wn})

    products = []
    unmatched_l13 = unmatched_ly = 0
    for r in forecast_rows:
        bbc = r["BBC SKU"].strip()
        l13 = resolve(bbc, l13_lookup)
        ly_weeks = resolve(bbc, ly_lookup)
        inv = resolve(bbc, inv_lookup)
        if l13 is None:
            unmatched_l13 += 1
        if ly_weeks is None:
            unmatched_ly += 1

        trend_pct = (l13["tyCases"] / l13["lyCases"] - 1) if (l13 and l13["lyCases"] > 0) else None

        weeks_out = []
        for wm in week_meta:
            f_raw = (r.get(f"{wm['weekEnding']} F") or "").strip()
            bbc_val = float(f_raw) if f_raw != "" else None
            ly_val = ly_weeks.get((wm["lyYear"], wm["lyWeekNum"])) if ly_weeks else None
            trend_val = max(0, round(ly_val * (1 + trend_pct))) if (trend_pct is not None and ly_val is not None) else None
            diff_pct = (bbc_val - trend_val) / trend_val if (trend_val and bbc_val is not None) else None
            weeks_out.append({
                "weekEnding": wm["weekEnding"],
                "lyWeekNum": wm["lyWeekNum"],
                "lyCases": ly_val,
                "trendForecast": trend_val,
                "bbcForecast": bbc_val,
                "diffPct": round(diff_pct, 4) if diff_pct is not None else None,
                "flagged": diff_pct is not None and abs(diff_pct) > 0.10,
            })

        products.append({
            "sku": bbc,
            "product": (l13["product"] if l13 else None) or r["Product"].strip(),
            "available": inv["available"] if inv else None,
            "lastReceiveDate": inv["lastReceiveDate"] if inv else None,
            "lastReceiveQty": inv["lastReceiveQty"] if inv else None,
            "trendPct": round(trend_pct, 4) if trend_pct is not None else None,
            "l13LyCases": l13["lyCases"] if l13 else None,
            "l13TyCases": l13["tyCases"] if l13 else None,
            "hasTrend": trend_pct is not None and any(w["lyCases"] is not None for w in weeks_out),
            "weeks": weeks_out,
        })

    return {
        "meta": {
            "l13Window": {
                "lastYear": f"{l13_windows['lyStart'].isoformat()} - {l13_windows['lyEnd'].isoformat()}",
                "thisYear": f"{l13_windows['tyStart'].isoformat()} - {l13_windows['tyEnd'].isoformat()}",
            },
            "weeks": week_meta,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
        },
        "products": products,
        "unmatchedL13": unmatched_l13,
        "unmatchedLy": unmatched_ly,
    }


def main():
    forecast_csv = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FORECAST_CSV
    inventory_csv = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_INVENTORY_CSV
    forecast_report_xlsx = Path(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_FORECAST_REPORT_XLSX
    l13_csv = Path(sys.argv[4]) if len(sys.argv) > 4 else DEFAULT_L13_CSV
    ly_by_week_csv = Path(sys.argv[5]) if len(sys.argv) > 5 else DEFAULT_LY_BY_WEEK_CSV
    for p in (forecast_csv, inventory_csv, forecast_report_xlsx):
        if not p.exists():
            raise SystemExit(f"Not found: {p}")

    with open(forecast_csv, newline="") as f:
        forecast_rows = list(csv.DictReader(f))
    if not forecast_rows:
        raise SystemExit("Forecast CSV has no data rows")

    week_dates = [m.group(1) for col in forecast_rows[0].keys() if (m := WEEK_COL_RE.match(col))]
    if not week_dates:
        raise SystemExit("Could not find any '<date> F' week columns in the forecast CSV")

    inv_lookup = load_inventory(inventory_csv)
    fr_lookup = load_forecast_report(forecast_report_xlsx)

    skus_out = []
    unmatched_inv = unmatched_fr = 0
    matched_inv_skus = set()
    for r in forecast_rows:
        bbc = r["BBC SKU"].strip()
        inv = resolve(bbc, inv_lookup)
        fr_weeks = resolve(bbc, fr_lookup)
        if inv is None:
            unmatched_inv += 1
        else:
            for k, v in inv_lookup.items():
                if v is inv:
                    matched_inv_skus.add(k)
                    break
        if fr_weeks is None:
            unmatched_fr += 1

        weeks = []
        for i, wd in enumerate(week_dates):
            f_raw = (r.get(f"{wd} F") or "").strip()
            a_raw = (r.get(f"{wd} A") or "").strip()
            weeks.append({
                "weekEnding": wd,
                "forecast": float(f_raw) if f_raw != "" else None,
                "onFile": float(a_raw) if a_raw != "" else None,
                "locked": i == 0,
            })

        skus_out.append({
            "sku": bbc,
            "product": (inv["cleanName"] if inv and inv["cleanName"] else None) or r["Product"],
            "rawProduct": r["Product"],
            "cleanName": bool(inv and inv["cleanName"]),
            "distributor": r["Distributor"],
            "customerSku": r["Customer SKU"],
            "weeks": weeks,
            "brandFamily": inv["brandFamily"] if inv else None,
            "package": inv["package"] if inv else None,
            "available": inv["available"] if inv else None,
            "l4Avg": inv["l4Avg"] if inv else None,
            "l8Avg": inv["l8Avg"] if inv else None,
            "recentWeeks": inv["recentWeeks"] if inv else None,
            "recentWeekLabels": RECENT_WEEK_LABELS,
            "lastReceiveDate": inv["lastReceiveDate"] if inv else None,
            "lastReceiveQty": inv["lastReceiveQty"] if inv else None,
            "forecastReportWeeks": fr_weeks,
        })

    orphan_inventory = []
    for sku, inv in inv_lookup.items():
        if sku not in matched_inv_skus:
            orphan_inventory.append({
                "sku": sku,
                "product": inv["cleanName"] or sku,
                "brandFamily": inv["brandFamily"],
                "available": inv["available"],
                "l4Avg": inv["l4Avg"],
                "lastReceiveDate": inv["lastReceiveDate"],
            })

    out = {
        "meta": {
            "weekDates": week_dates,
            "distributor": forecast_rows[0]["Distributor"],
            "lockedWeeks": 1,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
        },
        "skus": skus_out,
        "orphanInventory": orphan_inventory,
    }
    OUT_JSON.write_text(json.dumps(out))
    print(f"Wrote {len(skus_out)} SKUs, {len(week_dates)} weeks ({week_dates[0]} - {week_dates[-1]}).")
    if unmatched_inv:
        print(f"{unmatched_inv} SKU(s) in the forecast CSV have no matching Inventory Report row (new/unlisted item) -- shown with blanks for Available/recent sales.")
    if unmatched_fr:
        print(f"{unmatched_fr} SKU(s) in the forecast CSV have no matching ForecastReport row -- shown with no forward forecast.")
    print(f"{len(orphan_inventory)} Inventory Report row(s) have no matching forecast-CSV row (in inventory but not on the portal's forecast).")

    if l13_csv.exists() and ly_by_week_csv.exists():
        l13_lookup, l13_windows = load_l13_trend(l13_csv)
        ly_lookup = load_ly_by_week(ly_by_week_csv)
        trend = build_trend_forecast(forecast_rows, week_dates, l13_lookup, l13_windows, ly_lookup, inv_lookup)
        TREND_OUT_JSON.write_text(json.dumps(trend))
        wk = trend["meta"]["weeks"]
        print(f"\nWrote trend forecast for {len(trend['products'])} SKUs, {len(wk)} weeks "
              f"({wk[0]['weekEnding']} - {wk[-1]['weekEnding']}).")
        if trend["unmatchedL13"]:
            print(f"{trend['unmatchedL13']} SKU(s) have no matching L13_Trend.csv row -- no trend to project, shown BBC-forecast-only.")
        if trend["unmatchedLy"]:
            print(f"{trend['unmatchedLy']} SKU(s) have no matching LY_By_Week.csv row -- no weekly baseline to scale, shown BBC-forecast-only.")
    else:
        print(f"\nSkipping trend forecast: {l13_csv.name} and/or {ly_by_week_csv.name} not found.")


if __name__ == "__main__":
    main()
