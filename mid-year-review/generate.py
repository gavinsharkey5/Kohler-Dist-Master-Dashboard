#!/usr/bin/env python3
"""
Builds data/data.json for the 6-Month Review dashboard: for every brand
family, current YTD trend % vs. its 2026 Brewery Goal % and Kohler Goal %
(both from the 2026 planning workbook), so brand managers can see at a
glance whether they're ahead of, on, or behind pace -- and recalibrate the
back half of the year accordingly.

Inputs (keep these filenames when refreshing):
  2026_planning_source.xlsx  The 2026 Planning by Brand workbook. Gives us
                              the brand -> supplier / brand manager taxonomy
                              plus each brand's 2026 Brewery & Kohler Goal %
                              (columns K/N of the '2026 Planning by Brand'
                              tab). Same workbook used by ../2027-planning/.
  ytd_comparison.csv          RDE "Comparison" export, same-period-both-years
                              (e.g. "Case Equiv 1/1/2025-7/28/2025" vs
                              "Case Equiv 1/1/2026-7/28/2026" plus a
                              "Case Equiv % +/-" column) -- this IS the
                              current trend %, no projection math needed.
                              Column headers' date ranges shift every time
                              this gets re-pulled; matched by "Case Equiv"
                              prefix, not the exact header string.
  denise_food_bev_product_detail.csv (optional)
                              Product-level RDE export for Food & Bev
                              Enterprise LLC. RDE's own Brand Family
                              tagging conflates several of that
                              supplier's workbook brands (see
                              FOOD_BEV_BRAND_KEYWORDS below); this file's
                              Product Name text still distinguishes them,
                              recovering the real per-brand split.
  segment_package_trend.csv (optional)
                              Fusion "Segment / Package" export (Supplier,
                              Brand Family, Segment, Sub-Segments, Package,
                              Cases for the same two YTD windows as
                              ytd_comparison.csv, $Vol for both). Feeds the
                              Segment Trend header panel (a dropdown drills
                              from Segment down to that segment's Sub-
                              Segments) -- Cases-based, since there's no
                              Case Equiv version of this export yet (would
                              need a re-pull with the CE formula instead of
                              Cases; see brand_package_trend.csv below).
                              Also feeds a Cases-based Package Trend panel
                              as a FALLBACK only -- see
                              brand_package_trend.csv, which overrides it
                              with a Case-Equiv version whenever that file
                              is present (added 2026-08-10, per Gavin: CE
                              is this whole page's native unit, and Cases
                              was the odd one out). Both panels skipped
                              entirely if this file isn't present.
  brand_package_trend.csv (optional)
                              Fusion product-level export (Supplier, Brand
                              Family, Product Name, Package, Premise, Year
                              Month, Case Equiv for the same two YTD
                              windows as ytd_comparison.csv -- one row per
                              product/package/premise/month, each row's
                              Case Equiv landing in whichever year column
                              matches its own Year Month). Feeds:
                              (1) the "i" trend-driver popovers on the
                              Supplier + Brand tab (added 2026-08-10 per a
                              manager's suggestion): for each supplier (and
                              one company-wide "Overall" popover by the tab
                              heading), which brand families drove growth
                              vs. dragged it down, and which SPECIFIC
                              package (the raw Package label, e.g.
                              "1/15/19.2oz Can" -- deliberately not
                              bucketed into a coarse Cans/Bottles/Kegs
                              grouping, per Gavin 2026-08-10) grew vs.
                              shrank; and
                              (2) the header's Package Trend panel (also
                              2026-08-10), OVERRIDING segment_package_
                              trend.csv's Cases-based version with this
                              file's Case-Equiv one -- see main()'s
                              override right after both files are parsed.
                              See parse_brand_package_trend(). Skipped
                              entirely, with no popovers and Package Trend
                              falling back to Cases, if this file isn't
                              present.
  brand_geography_trend.csv (optional)
                              Encompass "Comparison" export (Brand Family,
                              Supplier, Package, City, County, Sales Rep
                              Assigned, Case Equiv for the same two YTD
                              windows -- reconciles exactly to
                              ytd_comparison.csv's own Total row). Adds a
                              "Counties growing/shrinking" section to the
                              SAME "i" popovers brand_package_trend.csv
                              feeds (added 2026-08-10, per Gavin) -- merged
                              into those insight objects in main()
                              (countyGainers/countyDecliners keys), not
                              kept as its own structure. Only County is
                              used (9 distinct, small enough for a top-3
                              mover list); City (219 distinct) is in the
                              file but unused for now. See
                              parse_brand_geography_trend(). Requires
                              brand_package_trend.csv to also be present
                              (nothing to merge county data into
                              otherwise); skipped entirely, with no county
                              sections, if this file isn't present.

A brand with no Brewery/Kohler goal set in the workbook (new items launched
after the plan was built, e.g. Carbliss, Monaco, Noca) is kept OUT of the
main vs-goal table and instead listed on its own "New in 2026" tab --
there's no goal to compare it against yet.

Also builds a supplier-level rollup (payload["supplierRollup"], the "By
Supplier" tab) using the SAME vs-goal math, but sourced from each
supplier's own grey header row in the workbook (its own Brewery/Kohler
Goal % and 2025 Finish, not a sum of its brands') and its own subtotal
row in ytd_comparison.csv (not a re-sum of the brand-level rows).

Run: python3 generate.py
"""
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
WORKBOOK = HERE / "2026_planning_source.xlsx"
CSV_YTD = HERE / "ytd_comparison.csv"
DENISE_PRODUCT_DETAIL = HERE / "denise_food_bev_product_detail.csv"
CSV_SEGMENT_PACKAGE = HERE / "segment_package_trend.csv"
CSV_BRAND_PACKAGE = HERE / "brand_package_trend.csv"
CSV_BRAND_GEOGRAPHY = HERE / "brand_geography_trend.csv"
CSV_DISTRICT_MANAGER = HERE / "district_manager_trend.csv"
HTML = HERE / "index.html"
OUT = HERE / "data" / "data.json"

# Same aliasing quirk documented in ../2027-planning/build_data.py: the
# workbook's hand-entered "Pabst Brand" is RDE's "Pabst Blue Ribbon".
NAME_ALIASES = {
    "pabst blue ribbon": "Pabst Brand",
}

# Per Gavin, 2026-08-04: the workbook itself mis-lists these brands' Supplier
# column -- e.g. "Fresca Mixed" is tagged "Constellation Brands" in the
# workbook, but ytd_comparison.csv's own row order puts it under its own
# "SAZERAC INC" header, separate from Constellation's brands entirely (a
# workbook data-entry error, not an RDE/parsing issue). Corrected here rather
# than editing 2026_planning_source.xlsx directly.
SUPPLIER_OVERRIDES = {
    "Fresca Mixed": "Sazerac Inc",
}


def to_num(raw):
    if raw is None:
        return 0.0
    raw = str(raw).strip()
    if raw == "" or raw == "-":
        return 0.0
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()").replace("$", "").replace(",", "").replace("%", "")
    if raw == "":
        return 0.0
    val = float(raw)
    return -val if neg else val


def load_workbook_taxonomy():
    wb_values = openpyxl.load_workbook(WORKBOOK, data_only=True)
    wb_formulas = openpyxl.load_workbook(WORKBOOK, data_only=False)
    wsv = wb_values["2026 Planning by Brand"]
    wsf = wb_formulas["2026 Planning by Brand"]

    brands = {}
    brands_lower = {}
    supplier_names = set()
    brand_manager_by_supplier = {}
    # Supplier-level rollup: the grey header row for each supplier carries its
    # OWN 2026 Brewery/Kohler Goal % and 2025 Finish (same columns J/K/N as a
    # brand row), not just its per-brand children -- this is what powers the
    # "By Supplier" tab.
    supplier_goals = {}

    for r in range(3, wsv.max_row + 1):
        brand = wsv.cell(r, 1).value
        if brand is None or str(brand).strip() == "":
            continue
        is_grey = wsf.cell(r, 1).fill.patternType == "solid"
        supplier = wsv.cell(r, 2).value

        if is_grey:
            grey_name = str(brand).strip()
            supplier_names.add(grey_name)
            grey_manager = wsv.cell(r, 3).value
            if grey_manager:
                brand_manager_by_supplier.setdefault(grey_name, str(grey_manager).strip())
            grey_finish_2025 = wsv.cell(r, 10).value
            grey_brewery_pct = wsv.cell(r, 11).value
            grey_kohler_pct = wsv.cell(r, 14).value
            supplier_goals[grey_name] = {
                "supplier": grey_name,
                "brand_manager": str(grey_manager).strip() if grey_manager else None,
                "finish_2025_ce": grey_finish_2025 if isinstance(grey_finish_2025, (int, float)) else None,
                "goal2026_brewery_pct": grey_brewery_pct if isinstance(grey_brewery_pct, (int, float)) else None,
                "goal2026_kohler_pct": grey_kohler_pct if isinstance(grey_kohler_pct, (int, float)) else None,
            }
            continue

        manager = wsv.cell(r, 3).value
        finish_2025 = wsv.cell(r, 10).value
        brewery_pct = wsv.cell(r, 11).value
        kohler_pct = wsv.cell(r, 14).value

        name = str(brand).strip()
        supplier_name = str(supplier).strip() if supplier else None
        supplier_name = SUPPLIER_OVERRIDES.get(name, supplier_name)
        brands[name] = {
            "brand": name,
            "supplier": supplier_name,
            "brand_manager": manager,
            "finish_2025_ce": finish_2025 if isinstance(finish_2025, (int, float)) else None,
            "goal2026_brewery_pct": brewery_pct if isinstance(brewery_pct, (int, float)) else None,
            "goal2026_kohler_pct": kohler_pct if isinstance(kohler_pct, (int, float)) else None,
        }
        if supplier:
            supplier_names.add(str(supplier).strip())
        if supplier and manager:
            brand_manager_by_supplier.setdefault(str(supplier).strip(), str(manager).strip())
        brands_lower[name.lower()] = name

    return brands, brands_lower, supplier_names, brand_manager_by_supplier, supplier_goals


def parse_ytd_csv(path, brands, brands_lower, supplier_names):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        prior_col = next(c for c in fieldnames if c.startswith("Case Equiv 1/1/2025") or c.startswith("Case Equiv 1/1/202") and "2025" in c)
        # Robust column matching: first "Case Equiv <date>" column is the prior
        # year, the second is the current year; the % column starts "Case Equiv %".
        ce_cols = [c for c in fieldnames if c.startswith("Case Equiv") and not c.startswith("Case Equiv %") and not c.startswith("Case Equiv ±")]
        pct_col = next(c for c in fieldnames if c.startswith("Case Equiv %"))
        prior_col, current_col = ce_cols[0], ce_cols[1]
        # e.g. "Case Equiv 1/1/2025 - 7/31/2025" -> "1/1/2025 - 7/31/2025" --
        # the exact comparison window pulled this refresh, so the dashboard's
        # column headers/notices can show it without ever being hand-edited.
        range_prior = prior_col.replace("Case Equiv", "").strip()
        range_current = current_col.replace("Case Equiv", "").strip()

        rows = []
        for r in reader:
            name = (r.get("Supplier / Brand Family") or "").strip()
            if not name or name == "Total":
                continue
            rows.append((name, {
                "ce_prior": to_num(r[prior_col]),
                "ce_current": to_num(r[current_col]),
                "pct_change": to_num(r[pct_col]) / 100.0,
            }))

    supplier_names_lower = {s.lower() for s in supplier_names}
    results = {}
    supplier_ytd = {}
    current_supplier = None
    unclassified = []
    skipped_phantom_headers = []

    for i, (raw_name, metrics) in enumerate(rows):
        name_l = raw_name.lower()
        next_name_l = rows[i + 1][0].lower() if i + 1 < len(rows) else None
        next_metrics = rows[i + 1][1] if i + 1 < len(rows) else None

        alias = NAME_ALIASES.get(name_l)
        canonical = alias or brands_lower.get(name_l)

        # A brand-family name that's immediately followed by an identical-name
        # row is functioning as this brand's own supplier-subtotal header here
        # (e.g. "Heineken USA" / "Heineken USA", "Sapporo" / "Sapporo") -- the
        # SECOND occurrence is the real leaf data, this first one is skipped.
        # When that header name is itself a known supplier (single-brand
        # suppliers like these), its own row IS that supplier's total.
        if canonical and next_name_l == name_l:
            current_supplier = canonical
            if canonical.lower() in supplier_names_lower:
                exact_supplier = next(s for s in supplier_names if s.lower() == canonical.lower())
                supplier_ytd.setdefault(exact_supplier, metrics)
            continue

        # A known brand name is treated as real leaf data even when that same
        # name is ALSO registered as a supplier label in the workbook (e.g.
        # single-brand entities like "Carbliss", "Monaco") -- some RDE pulls
        # emit just the one row for these instead of a header+leaf pair, and
        # without this check that lone row would get swallowed as a phantom
        # header (see the Kohler-side supplier check just below).
        if canonical:
            supplier = brands[canonical]["supplier"] or current_supplier
            results[(supplier, canonical)] = metrics
            # Single-brand suppliers (Sapporo, Kirin Ichiban, ...) sometimes
            # collapse to just this one row instead of a header+leaf pair --
            # when the brand's own registered supplier IS itself (same name),
            # this row's total doubles as that supplier's own YTD total.
            if supplier and supplier.lower() == canonical.lower() and supplier.lower() in supplier_names_lower:
                exact_supplier = next(s for s in supplier_names if s.lower() == supplier.lower())
                supplier_ytd.setdefault(exact_supplier, metrics)
            continue

        if name_l in supplier_names_lower:
            current_supplier = next(s for s in supplier_names if s.lower() == name_l)
            # This row's own Case Equiv figures ARE that supplier's total
            # (sum of every brand-family row that follows it) -- captured
            # here for the "By Supplier" rollup tab.
            supplier_ytd.setdefault(current_supplier, metrics)
            continue

        # Unrecognized name, not a known brand or supplier. If the very next
        # row has identical Case Equiv figures, this row is almost certainly
        # a phantom supplier-subtotal for that single next brand (RDE's own
        # distributor/supplier label for it, which isn't in the workbook at
        # all -- e.g. "SN Food & Beverage LLC" heading straight into
        # "Carbliss" with the same total) rather than real brand data of its
        # own; skip it so its volume isn't double-counted under two names.
        if next_metrics is not None and next_metrics == metrics:
            skipped_phantom_headers.append(raw_name)
            current_supplier = raw_name
            continue

        unclassified.append((current_supplier, raw_name, metrics))

    if skipped_phantom_headers:
        print(f"Skipped {len(skipped_phantom_headers)} phantom header row(s) (unrecognized name, "
              f"identical totals to the very next row): {skipped_phantom_headers}")

    return results, unclassified, supplier_ytd, range_prior, range_current


# ---------------------------------------------------------------------------
# Segment Trend (drill from Segment down to Sub-Segment) and Package Trend
# (top 20 individual package movers) -- two header panels, both Cases YoY
# from the Fusion segment/package export.
# ---------------------------------------------------------------------------

# Cases threshold (in whichever year is larger) for a package to be eligible
# for the top-movers ranking -- filters out packages so small that a swing
# from, say, 2 cases to 6 cases (+200%) would otherwise rank above a real
# story like 20,000 -> 30,000 cases (+50%).
MIN_PACKAGE_VOLUME = 500


# Shared by parse_segment_package_trend() (Cases, its original unit) and
# parse_brand_package_trend() (Case Equiv, added 2026-08-10 per Gavin --
# see that function's own docstring for why the Package Trend header panel
# now overrides its packageMovers with THIS unit instead). Field names are
# deliberately unit-agnostic (valPrior/valCurrent/valDiff, not
# casesPrior/casesCurrent) since which of the two units populates them is a
# call-site decision, not baked into the shape -- a field named
# "casesCurrent" holding Case Equivalents is exactly the kind of mislabeling
# that caused the Package-Trend-vs-Top-Headlines mismatch this was added to
# fix.
def build_package_movers(pkg_totals, min_volume):
    def to_row(label, prior, current):
        diff = current - prior
        pct = (current / prior - 1) if prior else (1.0 if current else 0.0)
        return {"label": label, "valPrior": round(prior, 1), "valCurrent": round(current, 1),
                "valDiff": round(diff, 1), "pctChange": round(pct, 4)}

    # Top movers: require real (nonzero) volume in BOTH years -- a package
    # that's brand-new this year or fully discontinued has an undefined /
    # infinite % swing that isn't a comparable "trend" and would otherwise
    # crowd out genuine percentage movers -- those counts are surfaced
    # separately instead (newCount/discontinuedCount) so nothing's hidden,
    # just not force-ranked on an undefined percentage.
    comparable = [(label, p, c) for label, (p, c) in pkg_totals.items()
                  if p > 0 and c > 0 and max(p, c) >= min_volume]
    new_count = sum(1 for p, c in pkg_totals.values() if p == 0 and c >= min_volume)
    discontinued_count = sum(1 for p, c in pkg_totals.values() if c == 0 and p >= min_volume)
    below_min_count = sum(1 for p, c in pkg_totals.values() if max(p, c) < min_volume)
    movers = [to_row(label, p, c) for label, p, c in comparable]
    return {
        "up": sorted(movers, key=lambda x: -x["pctChange"])[:10],
        "down": sorted(movers, key=lambda x: x["pctChange"])[:10],
        "minVolume": min_volume,
        "newCount": new_count, "discontinuedCount": discontinued_count, "belowMinCount": below_min_count,
    }


def parse_segment_package_trend(path):
    if not path.exists():
        return None
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        cases_cols = [c for c in fieldnames if c.startswith("Cases")]
        if len(cases_cols) < 2:
            return None
        prior_col, current_col = cases_cols[0], cases_cols[1]
        range_prior = prior_col.replace("Cases", "").strip()
        range_current = current_col.replace("Cases", "").strip()

        seg_totals = defaultdict(lambda: [0.0, 0.0])
        # segment -> sub-segment -> [prior, current]. Scoped per-segment
        # (not a global by-sub-segment rollup) because a handful of
        # Sub-Segments labels (e.g. "Beer - Non-Alc", "FMB") appear under
        # more than one Segment in this export -- drilling into "Beer"
        # should only show that sub-segment's Beer-segment volume.
        subseg_totals = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        pkg_totals = defaultdict(lambda: [0.0, 0.0])
        for r in reader:
            prior, current = to_num(r[prior_col]), to_num(r[current_col])
            segment = (r.get("Segment") or "").strip() or "Unclassified"
            seg_totals[segment][0] += prior
            seg_totals[segment][1] += current
            subsegment = (r.get("Sub-Segments") or "").strip() or "Unclassified"
            if subsegment.lower() == "is null":  # Fusion export quirk for a blank value
                subsegment = "Unclassified"
            ss = subseg_totals[segment][subsegment]
            ss[0] += prior
            ss[1] += current
            pkg = (r.get("Package") or "").strip()
            if pkg:
                pt = pkg_totals[pkg]
                pt[0] += prior
                pt[1] += current

    def to_row(label, prior, current):
        pct = (current / prior - 1) if prior else (1.0 if current else 0.0)
        return {"label": label, "casesPrior": round(prior, 1), "casesCurrent": round(current, 1),
                "pctChange": round(pct, 4)}

    segments = sorted((to_row(label, p, c) for label, (p, c) in seg_totals.items()), key=lambda x: -x["casesCurrent"])
    sub_segments = {
        segment: sorted((to_row(label, p, c) for label, (p, c) in subs.items()), key=lambda x: -x["casesCurrent"])
        for segment, subs in subseg_totals.items()
    }

    return {
        "rangePrior": range_prior, "rangeCurrent": range_current,
        "segments": segments,
        "subSegments": sub_segments,
        # Cases-based -- the fallback if brand_package_trend.csv isn't
        # present. main() overrides this with a Case-Equiv version from
        # that file when it is (see parse_brand_package_trend()), since CE
        # is this whole page's native unit everywhere else.
        "packageMovers": build_package_movers(pkg_totals, MIN_PACKAGE_VOLUME),
        "packageMoversUnit": "Cases",
    }


# ---------------------------------------------------------------------------
# Brand + package trend-driver popovers ("i" icons on the Supplier + Brand
# tab, added 2026-08-10 per a manager's suggestion): for each supplier, and
# once company-wide, which brand families are driving growth vs. dragging it
# down, and which SPECIFIC package (the raw Fusion Package label, e.g.
# "1/15/19.2oz Can" -- per Gavin, 2026-08-10, deliberately NOT bucketed into
# a coarse Cans/Bottles/Kegs grouping, since the point is to name the exact
# package a manager should go push) is growing vs. shrinking -- a quick read
# on WHY a supplier's trend % looks the way it does, without leaving the
# Supplier + Brand tab.
# ---------------------------------------------------------------------------
MIN_MOVER_CE = 0.5  # floor below which a Case Equiv swing is rounding noise, not a real mover
TOP_BRAND_MOVERS = 4
TOP_PACKAGE_MOVERS = 4


def _top_movers(totals, limit):
    """totals: {label: [prior, current]}. Returns (gainers, decliners), each
    a list of {label, cePrior, ceCurrent, ceDiff} sorted by magnitude,
    largest swing first, floored at MIN_MOVER_CE so rounding dust doesn't
    crowd out real movers."""
    rows = []
    for label, (prior, current) in totals.items():
        diff = current - prior
        if abs(diff) < MIN_MOVER_CE:
            continue
        rows.append({"label": label, "cePrior": round(prior, 1), "ceCurrent": round(current, 1), "ceDiff": round(diff, 1)})
    gainers = sorted((r for r in rows if r["ceDiff"] > 0), key=lambda r: -r["ceDiff"])[:limit]
    decliners = sorted((r for r in rows if r["ceDiff"] < 0), key=lambda r: r["ceDiff"])[:limit]
    return gainers, decliners


def parse_brand_package_trend(path):
    if not path.exists():
        return None
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        ce_cols = [c for c in fieldnames if c.startswith("Case Equiv") and "Difference" not in c]
        if len(ce_cols) < 2:
            return None
        prior_col, current_col = ce_cols[0], ce_cols[1]
        range_prior = prior_col.replace("Case Equiv", "").strip()
        range_current = current_col.replace("Case Equiv", "").strip()

        overall_brand = defaultdict(lambda: [0.0, 0.0])
        overall_pkg = defaultdict(lambda: [0.0, 0.0])
        overall_premise = defaultdict(lambda: [0.0, 0.0])
        by_supplier_brand = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        by_supplier_pkg = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        # supplier -> brand -> package -> [prior, current] -- package movers
        # scoped to ONE brand family, not the whole supplier, for the
        # brand-level "i" popover added 2026-08-10 per Gavin ("just want to
        # see the packages that have grown and declined" at the brand
        # level, no brand-vs-brand comparison needed since there's nothing
        # below a brand family to compare against here).
        by_brand_pkg = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0.0, 0.0])))
        supplier_totals = defaultdict(lambda: [0.0, 0.0])
        overall_totals = [0.0, 0.0]

        for r in reader:
            supplier = (r.get("Supplier") or "").strip()
            if not supplier:
                continue
            prior, current = to_num(r.get(prior_col)), to_num(r.get(current_col))
            brand = (r.get("Brand Family") or "").strip() or "Unclassified"
            pkg = (r.get("Package") or "").strip() or "Unspecified"
            premise = (r.get("Premise") or "").strip() or "Unclassified"

            overall_brand[brand][0] += prior
            overall_brand[brand][1] += current
            overall_pkg[pkg][0] += prior
            overall_pkg[pkg][1] += current
            overall_premise[premise][0] += prior
            overall_premise[premise][1] += current
            overall_totals[0] += prior
            overall_totals[1] += current

            sb = by_supplier_brand[supplier][brand]
            sb[0] += prior
            sb[1] += current
            sp = by_supplier_pkg[supplier][pkg]
            sp[0] += prior
            sp[1] += current
            bp = by_brand_pkg[supplier][brand][pkg]
            bp[0] += prior
            bp[1] += current
            st = supplier_totals[supplier]
            st[0] += prior
            st[1] += current

    def build_insight(brand_totals, pkg_totals, prior, current):
        brand_gainers, brand_decliners = _top_movers(brand_totals, TOP_BRAND_MOVERS)
        pkg_gainers, pkg_decliners = _top_movers(pkg_totals, TOP_PACKAGE_MOVERS)
        trend = (current / prior - 1) if prior else None
        return {
            "cePrior": round(prior, 1), "ceCurrent": round(current, 1), "trendPct": round(trend, 4) if trend is not None else None,
            "brandGainers": brand_gainers, "brandDecliners": brand_decliners,
            "pkgGainers": pkg_gainers, "pkgDecliners": pkg_decliners,
        }

    overall_insight = build_insight(overall_brand, overall_pkg, overall_totals[0], overall_totals[1])
    by_supplier = {
        supplier: build_insight(by_supplier_brand[supplier], by_supplier_pkg[supplier], *supplier_totals[supplier])
        for supplier in by_supplier_brand
    }

    # Brand-level "i" popover data (added 2026-08-10, per Gavin): package
    # movers only (no brandGainers/brandDecliners section -- there's
    # nothing below "brand family" to compare against, unlike the
    # supplier-level popover above). Nested supplier -> brand, matching how
    # the Supplier + Brand combo tree already looks a child brand up (see
    # main()'s attachment loop).
    def build_brand_insight(pkg_totals, prior, current):
        pkg_gainers, pkg_decliners = _top_movers(pkg_totals, TOP_PACKAGE_MOVERS)
        trend = (current / prior - 1) if prior else None
        return {
            "cePrior": round(prior, 1), "ceCurrent": round(current, 1), "trendPct": round(trend, 4) if trend is not None else None,
            "pkgGainers": pkg_gainers, "pkgDecliners": pkg_decliners,
        }

    by_brand = {
        supplier: {
            brand: build_brand_insight(by_brand_pkg[supplier][brand], *by_supplier_brand[supplier][brand])
            for brand in by_supplier_brand[supplier]
        }
        for supplier in by_supplier_brand
    }

    # Case-Equiv package movers (added 2026-08-10, per Gavin): the header's
    # Package Trend panel used to source its top-10-up/top-10-down list from
    # segment_package_trend.csv, in Cases -- a different unit than every
    # other number on this page, which reconciled a package's % change but
    # not its absolute unit count against the CE-based Top Headlines tile
    # (same package, same %, different raw number -- Cases and CE come from
    # different per-product fields, WholesaleUnitsPerCase vs CaseEquiv, so
    # they're not a fixed ratio at the Package-label level). main()
    # overrides segment_package_trend's packageMovers with THIS instead
    # whenever this file is present, so the whole page agrees on CE.
    package_movers_ce = build_package_movers(overall_pkg, MIN_PACKAGE_VOLUME)

    # On/Off Premise split (added 2026-08-10, per Gavin's "more headlines"
    # request) -- this file's own Premise column, not derived from anything
    # else, feeding one more Top Headlines sentence.
    def premise_row(prior, current):
        trend = (current / prior - 1) if prior else None
        return {"cePrior": round(prior, 1), "ceCurrent": round(current, 1), "trendPct": round(trend, 4) if trend is not None else None}
    premise_split = {label: premise_row(p, c) for label, (p, c) in overall_premise.items()}

    return {
        "rangePrior": range_prior, "rangeCurrent": range_current,
        "overall": overall_insight,
        "bySupplier": by_supplier,
        "byBrand": by_brand,
        "packageMoversCE": package_movers_ce,
        "premiseSplit": premise_split,
    }


TOP_COUNTY_MOVERS = 3  # only 9 counties in the territory total, so top 3 each direction covers most of it


# ---------------------------------------------------------------------------
# County-level "i" popover data (added 2026-08-10, per Gavin, from an
# Encompass "Comparison" export the user attached in chat -- same
# methodology/totals as ytd_comparison.csv and brand_package_trend.csv,
# reconciles to the same Total row, just with City/County added and broken
# out to Brand Family + Supplier + Package + Sales Rep grain). Only County
# is used here (City is available too -- 219 distinct vs. County's 9 -- but
# County alone is answers "where did this brand grow/shrink" at a glance;
# City would need its own top-N-movers treatment like Package Trend's if
# ever wanted). Merged into brand_package_trend's overall/bySupplier/
# byBrand insight objects in main() (adding countyGainers/countyDecliners
# keys to each) rather than kept as a separate top-level structure, so the
# existing "i" popovers just gain one more section with no new UI plumbing.
# ---------------------------------------------------------------------------
def parse_brand_geography_trend(path):
    if not path.exists():
        return None
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        ce_cols = [c for c in fieldnames if c.startswith("Case Equiv") and "±" not in c]
        if len(ce_cols) < 2:
            return None
        prior_col, current_col = ce_cols[0], ce_cols[1]

        overall_county = defaultdict(lambda: [0.0, 0.0])
        by_supplier_county = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        by_brand_county = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0.0, 0.0])))
        # Sales Rep Assigned -> County -> [prior, current] -- this file's
        # rep dimension, unused by the Supplier + Brand tab's own popovers
        # but reused by parse_district_manager_trend() (added 2026-08-10)
        # to add a Counties section to the District Manager Trends tab's
        # popovers too, joined by rep name (confirmed 100% name overlap
        # against district_manager_trend.csv before relying on this).
        by_rep_county = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))

        for r in reader:
            brand = (r.get("Brand Family") or "").strip()
            if not brand or brand == "Total":
                continue
            supplier = (r.get("Supplier") or "").strip()
            if not supplier:
                continue
            prior, current = to_num(r.get(prior_col)), to_num(r.get(current_col))
            county = (r.get("County") or "").strip() or "Unclassified"
            rep = (r.get("Sales Rep Assigned") or "").strip()

            oc = overall_county[county]
            oc[0] += prior
            oc[1] += current
            sc = by_supplier_county[supplier][county]
            sc[0] += prior
            sc[1] += current
            bc = by_brand_county[supplier][brand][county]
            bc[0] += prior
            bc[1] += current
            if rep:
                rc = by_rep_county[rep][county]
                rc[0] += prior
                rc[1] += current

    overall_gainers, overall_decliners = _top_movers(overall_county, TOP_COUNTY_MOVERS)
    by_supplier = {}
    for supplier, counties in by_supplier_county.items():
        gainers, decliners = _top_movers(counties, TOP_COUNTY_MOVERS)
        by_supplier[supplier] = {"countyGainers": gainers, "countyDecliners": decliners}
    by_brand = {}
    for supplier, brands in by_brand_county.items():
        by_brand[supplier] = {}
        for brand, counties in brands.items():
            gainers, decliners = _top_movers(counties, TOP_COUNTY_MOVERS)
            by_brand[supplier][brand] = {"countyGainers": gainers, "countyDecliners": decliners}

    return {
        "overallCountyGainers": overall_gainers, "overallCountyDecliners": overall_decliners,
        "bySupplier": by_supplier,
        "byBrand": by_brand,
        "byRepCounty": by_rep_county,
    }


TOP_PRODUCT_TYPE_MOVERS = 4
# "None"/blank District Manager and its lone "Default" rep are a tiny
# catch-all bucket (1,881 -> 3,255 CE as of the export this was built
# from, +73% swing off a near-zero base -- noise, not a real district),
# not an actual district manager -- excluded entirely rather than shown
# as a misleading "Unassigned" row.
DM_EXCLUDE_NAMES = {"none", ""}
REP_EXCLUDE_NAMES = {"default"}
# These 2 Product Types carry $0 CE in every row of the export this was
# built from (accounting adjustments, not real volume) -- excluded from
# product-type aggregation specifically (not from the rep/DM CE totals
# overall, though the effect is the same either way since they're zero).
PRODUCT_TYPE_EXCLUDE = {"finance charges", "hh finance charges"}


# ---------------------------------------------------------------------------
# District Manager Trends tab (added 2026-08-10, per Gavin, from an
# Encompass "Comparison" export the user attached in chat: District
# Manager, Sales Rep Assigned, Brand Family, Package, Product Type,
# On-Off Premise, Case Equiv for the same two YTD windows -- Total row
# again reconciles exactly to ytd_comparison.csv's own). Builds a
# District Manager -> Sales Rep tree (same collapsed-parent/expandable-
# children UI pattern as the Supplier + Brand combo tab, just without any
# goal-% machinery -- there's no rep-level goal data on hand) with an "i"
# popover at both levels covering: top brand families driving growth/
# decline, top Product Types (Case Beer/Keg Beer/Liquor/Wine/etc.) growing/
# shrinking, the On/Off-Premise CE split, and -- IF brand_geography_trend's
# byRepCounty came back non-None -- top Counties growing/shrinking too,
# joined by Sales Rep Assigned name (100% overlap confirmed against that
# file before relying on this join).
# ---------------------------------------------------------------------------
def build_dm_level_insight(brand_totals, ptype_totals, premise_totals, prior, current, county_totals=None):
    brand_gainers, brand_decliners = _top_movers(brand_totals, TOP_BRAND_MOVERS)
    ptype_gainers, ptype_decliners = _top_movers(ptype_totals, TOP_PRODUCT_TYPE_MOVERS)
    trend = (current / prior - 1) if prior else None
    premise_split = {}
    for label, (p, c) in premise_totals.items():
        t = (c / p - 1) if p else None
        premise_split[label] = {"cePrior": round(p, 1), "ceCurrent": round(c, 1), "trendPct": round(t, 4) if t is not None else None}
    result = {
        "cePrior": round(prior, 1), "ceCurrent": round(current, 1), "trendPct": round(trend, 4) if trend is not None else None,
        "brandGainers": brand_gainers, "brandDecliners": brand_decliners,
        "productTypeGainers": ptype_gainers, "productTypeDecliners": ptype_decliners,
        "premiseSplit": premise_split,
    }
    if county_totals is not None:
        county_gainers, county_decliners = _top_movers(county_totals, TOP_COUNTY_MOVERS)
        result["countyGainers"] = county_gainers
        result["countyDecliners"] = county_decliners
    return result


def parse_district_manager_trend(path, rep_county_totals=None):
    if not path.exists():
        return None
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        ce_cols = [c for c in fieldnames if c.startswith("Case Equiv") and "±" not in c]
        if len(ce_cols) < 2:
            return None
        prior_col, current_col = ce_cols[0], ce_cols[1]
        range_prior = prior_col.replace("Case Equiv", "").strip()
        range_current = current_col.replace("Case Equiv", "").strip()

        dm_totals = defaultdict(lambda: [0.0, 0.0])
        dm_brand = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        dm_ptype = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        dm_premise = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        dm_reps = defaultdict(set)
        rep_totals = defaultdict(lambda: [0.0, 0.0])
        rep_brand = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        rep_ptype = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        rep_premise = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        overall_totals = [0.0, 0.0]
        overall_brand = defaultdict(lambda: [0.0, 0.0])
        overall_ptype = defaultdict(lambda: [0.0, 0.0])
        overall_premise = defaultdict(lambda: [0.0, 0.0])

        for r in reader:
            dm = (r.get("District Manager") or "").strip()
            if dm.lower() in DM_EXCLUDE_NAMES:
                continue
            rep = (r.get("Sales Rep Assigned") or "").strip()
            if not rep or rep.lower() in REP_EXCLUDE_NAMES:
                continue
            prior, current = to_num(r.get(prior_col)), to_num(r.get(current_col))
            brand = (r.get("Brand Family") or "").strip() or "Unclassified"
            ptype = (r.get("Product Type") or "").strip()
            premise = (r.get("On-Off Premise") or "").strip() or "Unclassified"

            dm_totals[dm][0] += prior
            dm_totals[dm][1] += current
            dm_brand[dm][brand][0] += prior
            dm_brand[dm][brand][1] += current
            dm_premise[dm][premise][0] += prior
            dm_premise[dm][premise][1] += current
            dm_reps[dm].add(rep)

            rep_totals[rep][0] += prior
            rep_totals[rep][1] += current
            rep_brand[rep][brand][0] += prior
            rep_brand[rep][brand][1] += current
            rep_premise[rep][premise][0] += prior
            rep_premise[rep][premise][1] += current

            overall_totals[0] += prior
            overall_totals[1] += current
            overall_brand[brand][0] += prior
            overall_brand[brand][1] += current
            overall_premise[premise][0] += prior
            overall_premise[premise][1] += current

            if ptype.lower() not in PRODUCT_TYPE_EXCLUDE:
                dm_ptype[dm][ptype][0] += prior
                dm_ptype[dm][ptype][1] += current
                rep_ptype[rep][ptype][0] += prior
                rep_ptype[rep][ptype][1] += current
                overall_ptype[ptype][0] += prior
                overall_ptype[ptype][1] += current

    def county_totals_for_reps(reps):
        if rep_county_totals is None:
            return None
        totals = defaultdict(lambda: [0.0, 0.0])
        for rep in reps:
            for county, (p, c) in rep_county_totals.get(rep, {}).items():
                totals[county][0] += p
                totals[county][1] += c
        return totals

    overall_insight = build_dm_level_insight(
        overall_brand, overall_ptype, overall_premise, overall_totals[0], overall_totals[1],
        county_totals=county_totals_for_reps(rep_totals.keys()))

    dm_rollup = []
    for dm, (prior, current) in dm_totals.items():
        insight = build_dm_level_insight(
            dm_brand[dm], dm_ptype[dm], dm_premise[dm], prior, current,
            county_totals=county_totals_for_reps(dm_reps[dm]))
        children = []
        for rep in dm_reps[dm]:
            rp, rc = rep_totals[rep]
            rep_trend = (rc / rp - 1) if rp else None
            rep_insight = build_dm_level_insight(
                rep_brand[rep], rep_ptype[rep], rep_premise[rep], rp, rc,
                county_totals=county_totals_for_reps([rep]))
            children.append({
                "rep": rep, "cePrior": round(rp, 1), "ceCurrent": round(rc, 1),
                "trendPct": round(rep_trend, 4) if rep_trend is not None else None,
                "insight": rep_insight,
            })
        children.sort(key=lambda c: -(c["ceCurrent"] or 0))
        dm_trend = (current / prior - 1) if prior else None
        dm_rollup.append({
            "manager": dm, "cePrior": round(prior, 1), "ceCurrent": round(current, 1),
            "trendPct": round(dm_trend, 4) if dm_trend is not None else None,
            "repCount": len(dm_reps[dm]), "insight": insight, "children": children,
        })
    dm_rollup.sort(key=lambda g: -(g["ceCurrent"] or 0))

    return {
        "rangePrior": range_prior, "rangeCurrent": range_current,
        "overall": overall_insight,
        "rollup": dm_rollup,
    }


# ---------------------------------------------------------------------------
# Raw Supplier -> Brand Family hierarchy, reconstructed straight from
# ytd_comparison.csv's own row order -- for the "Supplier + Brand" combo tab.
# ---------------------------------------------------------------------------
# Per Gavin, 2026-08-04: that tab should mirror the RDE export's own
# structure exactly (every supplier it tracks, every brand family under it,
# whether or not it has a workbook goal) rather than the curated with-goal
# subset the Vs. Goal / By Supplier tabs use. RDE flattens a 2-level
# Supplier -> Brand Family tree into one column with no indent markers, but
# it's still fully recoverable: a header row's own Case Equiv figures always
# equal the SUM of the brand-family rows immediately beneath it, up to the
# next header row. This reconstructs that tree by greedily finding, for each
# candidate header row, the smallest run of following rows whose Case Equiv
# sum matches it exactly (absorbing any trailing exact-zero row too, since a
# real zero-volume child can be genuinely ambiguous with the next header
# otherwise -- see "Coney Island" under Boston Beer Company).
HIERARCHY_TOLERANCE = 0.03
HIERARCHY_MAX_CHILDREN = 60


def build_raw_supplier_tree(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        ce_cols = [c for c in fieldnames if c.startswith("Case Equiv")
                   and not c.startswith("Case Equiv %") and not c.startswith("Case Equiv ±")]
        prior_col, current_col = ce_cols[0], ce_cols[1]
        rows = []
        for r in reader:
            name = (r.get("Supplier / Brand Family") or "").strip()
            if not name or name == "Total":
                continue
            rows.append((name, to_num(r[prior_col]), to_num(r[current_col])))

    groups = []
    unresolved = []
    i = 0
    while i < len(rows):
        name, prior, current = rows[i]
        matched_k = None
        csum_p = csum_c = 0.0
        for k in range(1, min(HIERARCHY_MAX_CHILDREN, len(rows) - i)):
            cp, cc = rows[i + k][1], rows[i + k][2]
            csum_p += cp
            csum_c += cc
            if abs(csum_p - prior) < HIERARCHY_TOLERANCE and abs(csum_c - current) < HIERARCHY_TOLERANCE:
                # Greedily absorb any immediately-following exact-zero row(s)
                # too -- they don't change the sum, so they're ambiguous
                # between "this group's trailing child" and "next header's
                # leading child"; a zero-volume item is unambiguously the
                # former.
                j = k
                while (i + j + 1 < len(rows)
                       and abs(rows[i + j + 1][1]) < HIERARCHY_TOLERANCE
                       and abs(rows[i + j + 1][2]) < HIERARCHY_TOLERANCE):
                    j += 1
                matched_k = j
                break
        if matched_k is None:
            unresolved.append((name, prior, current))
            i += 1
            continue
        children = rows[i + 1:i + 1 + matched_k]
        groups.append({"supplier": name, "ce_prior": prior, "ce_current": current, "children": children})
        i += 1 + matched_k

    if unresolved:
        raise SystemExit(
            f"Could not reconstruct the Supplier/Brand Family hierarchy for {len(unresolved)} row(s) in "
            f"{path.name} -- a header row's own Case Equiv figures no longer sum-match its following rows. "
            f"Investigate before trusting the Supplier + Brand tab: {unresolved[:10]}")

    return groups


# RDE's own Brand Family tagging conflates several of Food & Bev Enterprise
# LLC's workbook line-item brands: Aguila Light Import products get tagged
# "Aguila Import" (inflating that brand's real total), and Club Colombia
# Dorada/Roja + Pilsen Import products all get tagged a generic "Food & Bev"
# with no per-brand split at all. Confirmed with Kohler, 2026-07-29, via a
# product-level RDE export (denise_food_bev_product_detail.csv) where the
# Product Name text (unlike Brand Family) still distinguishes them -- this
# recovers the real per-brand split by matching on that text instead.
FOOD_BEV_SUPPLIER = "Food & Bev Enterprise LLC"
FOOD_BEV_BRAND_KEYWORDS = [
    # (keyword to find in Product Name, canonical workbook brand). Order
    # matters: "light" is checked before the bare "aguila" fallback so
    # Aguila Light Import products don't get counted as Aguila Import.
    ("dorada", "Club Colombia Dorada"),
    ("roja", "Club Colombia Roja"),
    ("pilsen", "Pilsen Import"),
    ("light", "Aguila Light"),
    ("aguila", "Aguila Import"),
]


def parse_product_detail_overrides(path, supplier=FOOD_BEV_SUPPLIER, keywords=FOOD_BEV_BRAND_KEYWORDS):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        ce_cols = [c for c in fieldnames if c.startswith("Case Equiv") and not c.startswith("Case Equiv %") and not c.startswith("Case Equiv ±")]
        prior_col, current_col = ce_cols[0], ce_cols[1]
        rows = list(reader)

    sums = defaultdict(lambda: {"ce_prior": 0.0, "ce_current": 0.0})
    for row in rows:
        if (row.get("Supplier") or "").strip() != supplier:
            continue
        product = (row.get("Product Name") or "").lower()
        brand = next((b for kw, b in keywords if kw in product), None)
        if brand is None:
            continue
        sums[brand]["ce_prior"] += to_num(row[prior_col])
        sums[brand]["ce_current"] += to_num(row[current_col])

    overrides = {}
    for brand, m in sums.items():
        pct_change = (m["ce_current"] / m["ce_prior"] - 1) if m["ce_prior"] else None
        overrides[brand] = {"ce_prior": m["ce_prior"], "ce_current": m["ce_current"], "pct_change": pct_change}
    return overrides


# Excluded outright (per Kohler, 2026-07-28) -- these are negative/near-zero
# credit-adjustment entries in the RDE export, not real current placements.
EXCLUDED_BRANDS = {"shipyard", "jersey girl", "soda birch", "whole hog"}

# Per Kohler, 2026-07-29: show these managers' goal-less-but-actively-selling
# brands on the main Vs. Goal table instead of "New in 2026", with the goal/
# gap/projection columns left blank (no goal to map them to). Scoped to named
# managers only, not a general policy change -- brands with zero 2025 volume
# still belong on the New tab regardless of manager, since there's no prior-
# year baseline to show a trend against either way.
FORCE_VS_GOAL_MANAGERS = {"Denise Montes"}


def main():
    brands, brands_lower, supplier_names, brand_manager_by_supplier, supplier_goals = load_workbook_taxonomy()
    print(f"Loaded {len(brands)} canonical brand families with 2026 goals from the workbook.")

    matched, unclassified, supplier_ytd, range_prior, range_current = parse_ytd_csv(
        CSV_YTD, brands, brands_lower, supplier_names)
    print(f"Matched {len(matched)} brand families in {CSV_YTD.name}; {len(unclassified)} unmatched (new SKUs).")
    print(f"YTD comparison window: {range_prior}  vs.  {range_current}")

    segment_package_trend = parse_segment_package_trend(CSV_SEGMENT_PACKAGE)
    if segment_package_trend:
        print(f"Segment Trend: {len(segment_package_trend['segments'])} segments from {CSV_SEGMENT_PACKAGE.name} "
              f"({segment_package_trend['rangePrior']}  vs.  {segment_package_trend['rangeCurrent']}), Cases-based.")

    brand_package_trend = parse_brand_package_trend(CSV_BRAND_PACKAGE)
    if brand_package_trend:
        print(f"Brand/Package trend popovers: {len(brand_package_trend['bySupplier'])} suppliers from "
              f"{CSV_BRAND_PACKAGE.name} ({brand_package_trend['rangePrior']}  vs.  {brand_package_trend['rangeCurrent']}).")

    # Per Gavin, 2026-08-10: merge County movers into the SAME insight
    # objects brand_package_trend already built (overall / bySupplier /
    # byBrand), adding countyGainers/countyDecliners keys to each, rather
    # than keeping geography as a separate structure -- the existing "i"
    # popovers just pick up one more section with no new UI plumbing this
    # way. Requires brand_package_trend to exist first (nothing to merge
    # county data INTO otherwise); geography-only with no brand/package
    # file present is not supported since combo_rollup's insight
    # attachment loop below reads from brand_package_trend either way.
    brand_geography_trend = parse_brand_geography_trend(CSV_BRAND_GEOGRAPHY)
    if brand_geography_trend and brand_package_trend:
        brand_package_trend["overall"]["countyGainers"] = brand_geography_trend["overallCountyGainers"]
        brand_package_trend["overall"]["countyDecliners"] = brand_geography_trend["overallCountyDecliners"]
        matched_supplier_counties = 0
        for supplier, insight in brand_package_trend["bySupplier"].items():
            geo = brand_geography_trend["bySupplier"].get(supplier)
            if geo:
                matched_supplier_counties += 1
                insight["countyGainers"] = geo["countyGainers"]
                insight["countyDecliners"] = geo["countyDecliners"]
        matched_brand_counties = 0
        for supplier, supplier_brand_insights in brand_package_trend["byBrand"].items():
            geo_brands = brand_geography_trend["byBrand"].get(supplier, {})
            for brand, insight in supplier_brand_insights.items():
                geo = geo_brands.get(brand)
                if geo:
                    matched_brand_counties += 1
                    insight["countyGainers"] = geo["countyGainers"]
                    insight["countyDecliners"] = geo["countyDecliners"]
        print(f"County trend-driver data: merged into {matched_supplier_counties} supplier and "
              f"{matched_brand_counties} brand-family insight objects from {CSV_BRAND_GEOGRAPHY.name}.")
    elif brand_geography_trend and not brand_package_trend:
        print(f"WARNING: {CSV_BRAND_GEOGRAPHY.name} is present but {CSV_BRAND_PACKAGE.name} isn't -- "
              f"county data has nothing to merge into, so no county sections will render.")

    district_manager_trend = parse_district_manager_trend(
        CSV_DISTRICT_MANAGER, brand_geography_trend["byRepCounty"] if brand_geography_trend else None)
    if district_manager_trend:
        total_reps = sum(g["repCount"] for g in district_manager_trend["rollup"])
        print(f"District Manager Trends: {len(district_manager_trend['rollup'])} districts / {total_reps} reps "
              f"from {CSV_DISTRICT_MANAGER.name} ({district_manager_trend['rangePrior']}  vs.  "
              f"{district_manager_trend['rangeCurrent']})"
              f"{', with county data joined in' if brand_geography_trend else ' (no county data -- brand_geography_trend.csv absent)'}.")

    # Per Gavin, 2026-08-10: switch the header's Package Trend panel from
    # Cases (segment_package_trend.csv) to Case Equivalents
    # (brand_package_trend.csv) -- CE is this whole page's native unit
    # everywhere else, and the two units disagreeing on the same package's
    # absolute count (while agreeing on its %) read as a bug. Segment Trend
    # itself stays Cases-based -- brand_package_trend.csv has no Segment/
    # Sub-Segment column to convert it with; that needs a re-pull of
    # segment_package_trend.csv using the Case Equiv formula instead of
    # Cases, not something derivable from data already in hand.
    if segment_package_trend and brand_package_trend:
        segment_package_trend["packageMovers"] = brand_package_trend["packageMoversCE"]
        segment_package_trend["packageMoversUnit"] = "CE"
    if segment_package_trend:
        pm = segment_package_trend["packageMovers"]
        print(f"Package Trend: {len(pm['up'])} up / {len(pm['down'])} down movers, "
              f"{segment_package_trend['packageMoversUnit']}-based "
              f"(>= {pm['minVolume']} {segment_package_trend['packageMoversUnit']}; {pm['newCount']} new, "
              f"{pm['discontinuedCount']} discontinued, {pm['belowMinCount']} below the volume floor, excluded from ranking).")

    if DENISE_PRODUCT_DETAIL.exists():
        overrides = parse_product_detail_overrides(DENISE_PRODUCT_DETAIL)
        for brand, metrics in overrides.items():
            matched[(FOOD_BEV_SUPPLIER, brand)] = metrics
        # The generic "Food & Bev" rollup row is now fully represented by the
        # disaggregated brands above -- drop it so its volume isn't double-counted.
        unclassified = [u for u in unclassified if not (u[0] == FOOD_BEV_SUPPLIER and u[1] == "Food & Bev")]
        print(f"Applied product-level brand split for {len(overrides)} {FOOD_BEV_SUPPLIER} brand(s) "
              f"from {DENISE_PRODUCT_DETAIL.name}: {sorted(overrides)}.")

    matched = {k: v for k, v in matched.items() if k[1].lower() not in EXCLUDED_BRANDS}
    unclassified = [u for u in unclassified if u[1].lower() not in EXCLUDED_BRANDS]

    with_goal = []
    no_goal = []

    for (supplier, brand), metrics in matched.items():
        base = brands.get(brand, {})
        brewery_pct = base.get("goal2026_brewery_pct")
        kohler_pct = base.get("goal2026_kohler_pct")
        finish_2025 = base.get("finish_2025_ce")
        manager = base.get("brand_manager") or brand_manager_by_supplier.get(supplier)
        trend = metrics["pct_change"]
        ce_prior, ce_current = metrics["ce_prior"], metrics["ce_current"]

        # 2026 projected finish = current YTD + (2025's full-year finish minus
        # its own YTD-comparable slice, i.e. the "remainder" period) grown at
        # this year's YTD trend rate. Same method used in ../2027-planning/.
        if finish_2025 is not None:
            remainder_2025 = finish_2025 - ce_prior
            proj_finish = ce_current + remainder_2025 * (1 + (trend or 0))
        else:
            proj_finish = None

        rec = {
            "brand": brand,
            "supplier": supplier,
            "brand_manager": manager,
            "ce_prior": ce_prior,
            "ce_current": ce_current,
            "trend_pct": trend,
            "finish_2025_ce": finish_2025,
            "proj_finish_2026_ce": proj_finish,
        }

        # No goal at all, OR zero recorded sales in the 2025 comparable
        # window -- either way there's no real prior-year baseline to
        # measure a 2026 trend against, so it belongs on the New tab
        # rather than the main vs-goal table (per Kohler, 2026-07-28).
        if (brewery_pct is None and kohler_pct is None) or ce_prior == 0:
            if manager in FORCE_VS_GOAL_MANAGERS and ce_prior != 0:
                rec["goal_brewery_pct"] = None
                rec["goal_kohler_pct"] = None
                rec["gap_brewery"] = None
                rec["gap_kohler"] = None
                with_goal.append(rec)
            else:
                no_goal.append(rec)
        else:
            rec["goal_brewery_pct"] = brewery_pct
            rec["goal_kohler_pct"] = kohler_pct
            rec["gap_brewery"] = (trend - brewery_pct) if (trend is not None and brewery_pct is not None) else None
            rec["gap_kohler"] = (trend - kohler_pct) if (trend is not None and kohler_pct is not None) else None
            with_goal.append(rec)

    # Some suppliers have workbook line-item brands that RDE's brand-family
    # tagging never breaks out individually -- their combined volume shows up
    # as one generic-named row (e.g. "Sinless Vodka Cocktail" covering both
    # Jim Beam Kentucky and Sinless). Relabel those rows to name the actual
    # planning-workbook brands they represent, so they're recognizable
    # instead of opaque. (Food & Bev Enterprise LLC's brands used to hit this
    # too, until the product-level override above started splitting them for
    # real instead of just relabeling the rollup.)
    matched_brand_names = {b for (_supplier, b) in matched.keys()}
    brands_by_supplier = defaultdict(list)
    for b_name, b_rec in brands.items():
        if b_rec["supplier"]:
            brands_by_supplier[b_rec["supplier"]].append(b_name)

    for supplier, name, metrics in unclassified:
        manager = brand_manager_by_supplier.get(supplier)
        unbroken_out = [b for b in brands_by_supplier.get(supplier, []) if b not in matched_brand_names]
        brand_label = f"{name} ({', '.join(sorted(unbroken_out))})" if unbroken_out else name
        rec = {
            "brand": brand_label,
            "supplier": supplier,
            "brand_manager": manager,
            "ce_prior": metrics["ce_prior"],
            "ce_current": metrics["ce_current"],
            "trend_pct": metrics["pct_change"],
        }
        if manager in FORCE_VS_GOAL_MANAGERS and metrics["ce_prior"] != 0:
            rec.update(finish_2025_ce=None, proj_finish_2026_ce=None,
                       goal_brewery_pct=None, goal_kohler_pct=None,
                       gap_brewery=None, gap_kohler=None)
            with_goal.append(rec)
        else:
            no_goal.append(rec)

    # Terminated: zero or negative 2026 YTD case volume, regardless of
    # whether the brand has a goal or prior-year sales -- it's not actively
    # selling right now, so it doesn't belong in either the vs-goal
    # comparison or the New-in-2026 white-space list. Pulled out of both
    # buckets rather than added on top, so a brand appears in exactly one tab.
    # ce_current is None for the goal-only per-brand rows added above (no
    # actual sales tracked against them individually) -- never terminated.
    terminated = [r for r in with_goal + no_goal if r["ce_current"] is not None and r["ce_current"] <= 0]
    with_goal = [r for r in with_goal if r["ce_current"] is None or r["ce_current"] > 0]
    no_goal = [r for r in no_goal if r["ce_current"] is None or r["ce_current"] > 0]

    with_goal.sort(key=lambda r: (r["gap_brewery"] if r["gap_brewery"] is not None else 999))
    no_goal.sort(key=lambda r: -r["ce_current"])
    terminated.sort(key=lambda r: r["ce_current"])

    managers = sorted({r["brand_manager"] for r in with_goal + no_goal + terminated if r.get("brand_manager")})
    suppliers = sorted({r["supplier"] for r in with_goal + no_goal if r.get("supplier")})

    behind_brewery = sum(1 for r in with_goal if r.get("gap_brewery") is not None and r["gap_brewery"] < 0)
    behind_kohler = sum(1 for r in with_goal if r.get("gap_kohler") is not None and r["gap_kohler"] < 0)

    # ------------------------------------------------------------------
    # Supplier-level rollup ("By Supplier" tab): same vs-goal math as brand
    # families, but at the supplier (Constellation, MolsonCoors, ...) level,
    # using each supplier's OWN Brewery/Kohler Goal % from the workbook's
    # grey header row and its own Case Equiv total from the YTD comparison.
    # ------------------------------------------------------------------
    supplier_rollup = []
    for supplier_name, metrics in supplier_ytd.items():
        goal = supplier_goals.get(supplier_name)
        if goal is None:
            continue
        ce_prior, ce_current, trend = metrics["ce_prior"], metrics["ce_current"], metrics["pct_change"]
        finish_2025 = goal.get("finish_2025_ce")
        brewery_pct = goal.get("goal2026_brewery_pct")
        kohler_pct = goal.get("goal2026_kohler_pct")

        if finish_2025 is not None:
            remainder_2025 = finish_2025 - ce_prior
            proj_finish = ce_current + remainder_2025 * (1 + (trend or 0))
        else:
            proj_finish = None

        supplier_rollup.append({
            "supplier": supplier_name,
            "brand_manager": goal.get("brand_manager"),
            "finish_2025_ce": finish_2025,
            "ce_prior": ce_prior,
            "ce_current": ce_current,
            "trend_pct": trend,
            "proj_finish_2026_ce": proj_finish,
            "goal_brewery_pct": brewery_pct,
            "goal_kohler_pct": kohler_pct,
            "gap_brewery": (trend - brewery_pct) if (trend is not None and brewery_pct is not None) else None,
            "gap_kohler": (trend - kohler_pct) if (trend is not None and kohler_pct is not None) else None,
        })

    # A handful of suppliers appear on brand rows (their "Supplier" column
    # value) but never got their own grey header/goal row built into the
    # workbook at all (e.g. "Food & Bev Enterprise LLC" for Denise Montes'
    # brands -- confirmed there's no such grey row anywhere in the workbook).
    # Rather than let their brands silently vanish from every supplier-level
    # view, synthesize a no-goal supplier entry by summing their own
    # with-goal children directly -- same "No Goal" treatment a goal-less
    # brand already gets, just at the supplier level.
    covered_suppliers = {r["supplier"] for r in supplier_rollup}
    orphan_suppliers = defaultdict(list)
    for r in with_goal:
        sup = r.get("supplier")
        if sup and sup not in covered_suppliers:
            orphan_suppliers[sup].append(r)

    for supplier_name, children in orphan_suppliers.items():
        ce_prior = sum(c["ce_prior"] for c in children if c["ce_prior"] is not None)
        ce_current = sum(c["ce_current"] for c in children if c["ce_current"] is not None)
        trend = (ce_current / ce_prior - 1) if ce_prior else None
        manager = brand_manager_by_supplier.get(supplier_name) or next(
            (c.get("brand_manager") for c in children if c.get("brand_manager")), None)
        supplier_rollup.append({
            "supplier": supplier_name,
            "brand_manager": manager,
            "finish_2025_ce": None,
            "ce_prior": ce_prior,
            "ce_current": ce_current,
            "trend_pct": trend,
            "proj_finish_2026_ce": None,
            "goal_brewery_pct": None,
            "goal_kohler_pct": None,
            "gap_brewery": None,
            "gap_kohler": None,
        })
    if orphan_suppliers:
        print(f"Synthesized {len(orphan_suppliers)} no-goal supplier rollup row(s) for suppliers with "
              f"brand-level goals but no supplier-level grey row in the workbook: {sorted(orphan_suppliers)}")

    supplier_rollup.sort(key=lambda r: (r["gap_brewery"] if r["gap_brewery"] is not None else 999))
    behind_brewery_supplier = sum(1 for r in supplier_rollup if r.get("gap_brewery") is not None and r["gap_brewery"] < 0)
    behind_kohler_supplier = sum(1 for r in supplier_rollup if r.get("gap_kohler") is not None and r["gap_kohler"] < 0)
    print(f"Supplier rollup: {len(supplier_rollup)} suppliers with both workbook goals and YTD data "
          f"(+ {len(orphan_suppliers)} synthesized no-goal supplier(s)).")

    # ------------------------------------------------------------------
    # "Supplier + Brand" combo tab: mirrors ytd_comparison.csv's own
    # Supplier -> Brand Family structure directly (every supplier RDE
    # tracks, every brand family under it -- not just the with-goal
    # subset above), attaching a workbook goal % wherever one exists and
    # leaving it blank otherwise. Per Gavin, 2026-08-04.
    # ------------------------------------------------------------------
    supplier_goals_lower = {k.lower(): v for k, v in supplier_goals.items()}
    raw_tree = build_raw_supplier_tree(CSV_YTD)

    def brand_goal_lookup(raw_name):
        canonical = NAME_ALIASES.get(raw_name.lower()) or brands_lower.get(raw_name.lower())
        return brands.get(canonical) if canonical else None

    combo_rollup = []
    for group in raw_tree:
        supplier_name = group["supplier"]

        if supplier_name == FOOD_BEV_SUPPLIER:
            # Per Gavin, 2026-08-04: leave Denise Montes' Food & Bev
            # Enterprise LLC brands exactly as already computed above (the
            # product-detail override split) -- don't rebuild them from
            # this raw CSV's generic, unsplit rollup.
            parent = next((r for r in supplier_rollup if r["supplier"] == FOOD_BEV_SUPPLIER), None)
            if parent is None:
                continue
            children_recs = [dict(r) for r in with_goal if r.get("supplier") == FOOD_BEV_SUPPLIER]
            combo_rollup.append({**dict(parent), "children": children_recs})
            continue

        goal = supplier_goals_lower.get(supplier_name.lower())
        ce_prior, ce_current = group["ce_prior"], group["ce_current"]
        trend = (ce_current / ce_prior - 1) if ce_prior else None
        finish_2025 = goal.get("finish_2025_ce") if goal else None
        brewery_pct = goal.get("goal2026_brewery_pct") if goal else None
        kohler_pct = goal.get("goal2026_kohler_pct") if goal else None
        manager = (goal.get("brand_manager") if goal else None) or brand_manager_by_supplier.get(supplier_name)
        proj_finish = (ce_current + (finish_2025 - ce_prior) * (1 + (trend or 0))) if finish_2025 is not None else None

        children_recs = []
        for child_name, c_prior, c_current in group["children"]:
            base = brand_goal_lookup(child_name) or {}
            c_trend = (c_current / c_prior - 1) if c_prior else None
            c_finish = base.get("finish_2025_ce")
            c_proj = (c_current + (c_finish - c_prior) * (1 + (c_trend or 0))) if c_finish is not None else None
            children_recs.append({
                "brand": child_name,
                "brand_manager": base.get("brand_manager") or manager,
                "finish_2025_ce": c_finish, "ce_prior": c_prior, "ce_current": c_current,
                "trend_pct": c_trend, "proj_finish_2026_ce": c_proj,
                "goal_brewery_pct": base.get("goal2026_brewery_pct"),
                "goal_kohler_pct": base.get("goal2026_kohler_pct"),
            })

        combo_rollup.append({
            "supplier": supplier_name, "brand_manager": manager,
            "finish_2025_ce": finish_2025, "ce_prior": ce_prior, "ce_current": ce_current,
            "trend_pct": trend, "proj_finish_2026_ce": proj_finish,
            "goal_brewery_pct": brewery_pct, "goal_kohler_pct": kohler_pct,
            "children": children_recs,
        })

    # Per Gavin, 2026-08-05: drop dead weight from the Supplier + Brand tab --
    # a whole supplier with 0 or negative 2026 YTD CE (e.g. the Buzbee's...
    # Point Brewing tail of the export), or an individual brand family with
    # 0/negative 2026 YTD CE even under an otherwise-healthy supplier (e.g.
    # Corona Refresca under Constellation). Same threshold the Terminated
    # Brands tab already uses. Buzbee's itself is +$0.57 CE (effectively
    # zero) but was named explicitly as the start of the range to drop.
    COMBO_MANUAL_EXCLUDE_SUPPLIERS = {"Buzbee's Beverages USA LLC"}
    existing_terminated_names = {r["brand"] for r in terminated}
    newly_terminated = []

    # A raw CSV row that never matched any single workbook brand shows up
    # here under its literal RDE name (e.g. "Monaco"), but it's ALREADY
    # represented in the with_goal/no_goal/terminated pool under a relabeled
    # compound name (e.g. "Monaco (Lech, Milwaukee's Best)" -- see the
    # unbroken_out relabeling below in this file). Match on the raw
    # (supplier, name) pair, not the display name, so this doesn't get
    # double-counted as a "new" addition to Terminated / New Brand Families.
    unclassified_pairs = {(u[0], u[1]) for u in unclassified}

    def already_represented(supplier, brand, name_set):
        return brand in name_set or (supplier, brand) in unclassified_pairs

    def is_dead(v):
        return v is not None and v <= 0

    kept_rollup = []
    for group in combo_rollup:
        if is_dead(group["ce_current"]) or group["supplier"] in COMBO_MANUAL_EXCLUDE_SUPPLIERS:
            for c in group["children"]:
                if c["brand"].lower() in EXCLUDED_BRANDS or already_represented(group["supplier"], c["brand"], existing_terminated_names):
                    continue
                newly_terminated.append({
                    "brand": c["brand"], "supplier": group["supplier"],
                    "brand_manager": c["brand_manager"], "ce_prior": c["ce_prior"], "ce_current": c["ce_current"],
                })
            continue
        kept_children = []
        for c in group["children"]:
            if is_dead(c["ce_current"]):
                if c["brand"].lower() not in EXCLUDED_BRANDS and not already_represented(group["supplier"], c["brand"], existing_terminated_names):
                    newly_terminated.append({
                        "brand": c["brand"], "supplier": group["supplier"],
                        "brand_manager": c["brand_manager"], "ce_prior": c["ce_prior"], "ce_current": c["ce_current"],
                    })
                continue
            kept_children.append(c)
        group["children"] = kept_children
        kept_rollup.append(group)

    dropped_supplier_count = len(combo_rollup) - len(kept_rollup)
    combo_rollup = kept_rollup
    if newly_terminated:
        terminated = terminated + newly_terminated
        terminated.sort(key=lambda r: r["ce_current"])

    # Per Gavin, 2026-08-05: same idea, mirrored for the OTHER end of the
    # lifecycle -- a whole supplier with 0 sales in 2025 (brand-new to the
    # portfolio, e.g. Carbliss under SN Food & Beverage LLC, Noca), or an
    # individual brand family with 0 2025 sales even under an established
    # supplier (e.g. Monaco under MolsonCoors), has no real prior-year
    # baseline to show a trend against -- move it to New Brand Families in
    # 2026 instead, same threshold that tab already uses.
    existing_new_names = {r["brand"] for r in no_goal}
    newly_new = []

    def is_new(v):
        return v is not None and v <= 0

    kept_rollup = []
    for group in combo_rollup:
        if is_new(group["ce_prior"]):
            for c in group["children"]:
                if c["brand"].lower() in EXCLUDED_BRANDS or already_represented(group["supplier"], c["brand"], existing_new_names):
                    continue
                newly_new.append({
                    "brand": c["brand"], "supplier": group["supplier"],
                    "brand_manager": c["brand_manager"], "ce_prior": c["ce_prior"], "ce_current": c["ce_current"],
                })
            continue
        kept_children = []
        for c in group["children"]:
            if is_new(c["ce_prior"]):
                if c["brand"].lower() not in EXCLUDED_BRANDS and not already_represented(group["supplier"], c["brand"], existing_new_names):
                    newly_new.append({
                        "brand": c["brand"], "supplier": group["supplier"],
                        "brand_manager": c["brand_manager"], "ce_prior": c["ce_prior"], "ce_current": c["ce_current"],
                    })
                continue
            kept_children.append(c)
        group["children"] = kept_children
        kept_rollup.append(group)

    dropped_new_supplier_count = len(combo_rollup) - len(kept_rollup)
    combo_rollup = kept_rollup
    if newly_new:
        no_goal = no_goal + newly_new
        no_goal.sort(key=lambda r: -(r["ce_current"] or 0))

    # Recompute now that newly_terminated/newly_new may have introduced
    # brand managers not already covered by the with_goal/no_goal/terminated
    # pool at the point managers/suppliers were first computed above.
    managers = sorted({r["brand_manager"] for r in with_goal + no_goal + terminated if r.get("brand_manager")})
    suppliers = sorted({r["supplier"] for r in with_goal + no_goal if r.get("supplier")})

    if brand_package_trend:
        matched_insight_count = 0
        matched_brand_insight_count = 0
        total_children_count = 0
        for group in combo_rollup:
            insight = brand_package_trend["bySupplier"].get(group["supplier"])
            if insight:
                matched_insight_count += 1
            group["insight"] = insight
            brand_insights = brand_package_trend["byBrand"].get(group["supplier"], {})
            for child in group["children"]:
                total_children_count += 1
                child_insight = brand_insights.get(child["brand"])
                if child_insight:
                    matched_brand_insight_count += 1
                child["insight"] = child_insight
        print(f"Matched trend-driver popovers to {matched_insight_count} / {len(combo_rollup)} suppliers and "
              f"{matched_brand_insight_count} / {total_children_count} brand families on the Supplier + Brand tab "
              f"(rows absent from {CSV_BRAND_PACKAGE.name} simply get no popover).")

    combo_rollup.sort(key=lambda r: -(r["ce_current"] or 0))
    total_combo_children = sum(len(g["children"]) for g in combo_rollup)
    print(f"Supplier + Brand combo rollup: {len(combo_rollup)} suppliers, {total_combo_children} brand families "
          f"(mirrors {CSV_YTD.name}'s own hierarchy; Food & Bev Enterprise LLC left as the existing override).")
    print(f"Dropped {dropped_supplier_count} supplier(s) with 0/negative 2026 CE and their brands from the combo "
          f"tab; moved {len(newly_terminated)} previously-unlisted brand(s) into Terminated Brands.")
    print(f"Dropped {dropped_new_supplier_count} supplier(s) with 0 2025 CE and their brands from the combo tab; "
          f"moved {len(newly_new)} previously-unlisted brand(s) into New Brand Families in 2026.")

    payload = {
        "generatedNote": "Built from 2026_planning_source.xlsx (goals) + ytd_comparison.csv (current trend). "
                          "See generate.py for methodology.",
        "meta": {
            "totalWithGoal": len(with_goal),
            "totalNoGoal": len(no_goal),
            "totalTerminated": len(terminated),
            "behindBrewery": behind_brewery,
            "behindKohler": behind_kohler,
            "totalSuppliers": len(supplier_rollup),
            "behindBrewerySupplier": behind_brewery_supplier,
            "behindKohlerSupplier": behind_kohler_supplier,
            "ytdRangePrior": range_prior,
            "ytdRangeCurrent": range_current,
        },
        "managers": managers,
        "suppliers": suppliers,
        "brands": with_goal,
        "newBrands": no_goal,
        "terminatedBrands": terminated,
        "supplierRollup": supplier_rollup,
        "comboRollup": combo_rollup,
        "segmentPackageTrend": segment_package_trend,
        "overallInsight": brand_package_trend["overall"] if brand_package_trend else None,
        "insightRange": {"prior": brand_package_trend["rangePrior"], "current": brand_package_trend["rangeCurrent"]} if brand_package_trend else None,
        "premiseSplit": brand_package_trend["premiseSplit"] if brand_package_trend else None,
        "dmTrend": district_manager_trend,
    }
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2))
    print(f"Wrote {len(with_goal)} brands with goals + {len(no_goal)} brands with no 2025 goal/sales "
          f"+ {len(terminated)} terminated brands to {OUT}")
    print(f"Behind Brewery goal: {behind_brewery} / {len(with_goal)}   Behind Kohler goal: {behind_kohler} / {len(with_goal)}")
    print(f"Supplier rollup behind Brewery: {behind_brewery_supplier} / {len(supplier_rollup)}   "
          f"behind Kohler: {behind_kohler_supplier} / {len(supplier_rollup)}")


if __name__ == "__main__":
    main()
