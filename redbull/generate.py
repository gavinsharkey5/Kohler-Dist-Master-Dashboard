#!/usr/bin/env python3
"""
Rebuilds data.csv from the RDE "Red Bull Tracker" export -- a raw
transaction-level report (one row per account x SKU x order date), not
the pre-aggregated (Customer Name, Category, Sales Rep, Bought) shape
index.html actually reads.

Classification (per Kohler): Core = Regular and Sugar Free Red Bull;
Core+ = every other flavor edition. Explicit product lists below, not a
keyword guess -- the script raises if an unrecognized product shows up
so a new flavor gets a deliberate Core/Core+ call instead of a silent
guess.

Usage:
    python3 generate.py RDE_Red_Bull_Tracker_Apr_1_Start.xlsx

Output: data.csv, tab-separated, one row per (Customer Name, Category,
Sales Rep) with at least one qualifying order -- matching the existing
file's shape exactly (index.html's ingestData() doesn't care about the
Bought column's value, just whether the row exists, so it's always 1).

goals.csv is untouched -- this export carries no goal information.
"""
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
OUT_CSV = HERE / "data.csv"

CORE_PRODUCTS = {
    '8710 Red Bull 1/24/8.4 oz Can',
    '8720 Red Bull Sugar Free 1/24/8.4 oz Can',
}
COREPLUS_PRODUCTS = {
    '8711 Red Bull Red Edition 1/24/8.4 oz Can',
    '8715 Red Bull Yellow Edition 1/24/8.4 oz Can',
    '8719 Red Bull Coconut 1/24/8.4 oz Can',
    '8723 Red Bull White Peach Edition 1/24/8.4 oz Can',
}


def classify(product):
    if product in CORE_PRODUCTS:
        return 'Core'
    if product in COREPLUS_PRODUCTS:
        return 'Core+'
    return None


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python3 generate.py RDE_Red_Bull_Tracker_Apr_1_Start.xlsx")
    src = Path(sys.argv[1])

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    unknown = set()
    # (rep, customer) -> ordered dict of categories seen (dict as insertion-ordered set)
    seen = {}
    order = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rep = row[idx['Sales Rep Assigned']]
        cust = row[idx['Customer Name']]
        product = row[idx['Product Num & Name']]
        if not rep or not cust:
            continue
        cat = classify(product)
        if cat is None:
            unknown.add(product)
            continue
        key = (rep, cust)
        if key not in seen:
            seen[key] = {}
            order.append(key)
        seen[key][cat] = True

    if unknown:
        raise SystemExit(
            f"Unclassified Red Bull product(s) -- add to CORE_PRODUCTS or COREPLUS_PRODUCTS "
            f"in this script after confirming with the user: {sorted(unknown)}"
        )

    lines = ["Customer Name\tCategory\tSales Rep\tBought"]
    for (rep, cust) in order:
        for cat in seen[(rep, cust)]:
            lines.append(f"{cust}\t{cat}\t{rep}\t1")
    OUT_CSV.write_text("\n".join(lines) + "\n")

    accounts = len(order)
    core = sum(1 for k in order if seen[k].get('Core'))
    coreplus = sum(1 for k in order if seen[k].get('Core+'))
    print(f"Wrote {len(lines) - 1} rows across {accounts} buying accounts "
          f"({core} Core, {coreplus} Core+)")


if __name__ == "__main__":
    main()
