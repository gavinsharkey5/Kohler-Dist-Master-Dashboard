#!/usr/bin/env python3
"""
Rebuilds MPOs/on-prem/data/2026-08/*.json from the August RDE CSV exports.

August's on-premise MPO objectives are entirely different brands/rules from
July's (Carbliss/Sapporo NA/Wine & Spirits), so this is a separate script
from generate.py rather than another branch of it -- see MONTH_KEY in each
script and the MONTHS array in index.html for how a month's script maps to
its data/<MONTH_KEY>/ folder.

Inputs (keep these filenames when re-exporting from RDE):
  angry_orchard_new_lines.csv     RDE "2 New Angry Orchard Draft Lines"
                                    export: Sales Rep Assigned, Brand
                                    Family, Customer Num, Customer Name,
                                    Date, Buyer Count, Placement Count,
                                    Cases -- one row per order occasion,
                                    windowed 5/1/2026-8/31/2026.
  molson_coors_peroni_banquet.csv RDE "Molson Coors ON (4) New Peroni
                                    Placements (4) New Banquet Placements
                                    90 Day Non Buy" export: same shape as
                                    above minus Buyer Count, Brand Family
                                    is "Peroni" or "Coors" (Coors = the
                                    Banquet objective's raw brand label
                                    in RDE), same 5/1-8/31 window.
  wine_spirits_yave_leyenda.csv   RDE "2 Yave Buying Accounts 2 Leyenda
                                    Buying Accounts" export: Sales Rep
                                    Assigned, Product Name, Product Num,
                                    Brand Family, Customer Num, Customer
                                    Name, Date, Placement Count, Cases --
                                    August-only window (8/1-8/31), no
                                    prior-month columns since this
                                    objective is a plain buyer count, not
                                    a new-vs-repeat classification.

Classification -- "90-Day Non-Buy" new placement (Angry Orchard, and each
brand within Molson Coors independently), per Kohler, 2026-08-04: a
customer's row on a given date is a NEW placement only if they have NO
purchase of that same brand before NEW_BUYER_WINDOW_START (i.e. in
May/June/July) AND DO have a purchase of it in August. A customer who
bought before August and buys again in August is a regular repeat
placement and does NOT count. Same date-based approach as on-prem's
Carbliss classification (see generate.py) -- every transaction row is
kept in the output (rep detail view shows full activity), with
NEW_PLACEMENT set to 1 on exactly the customer's first qualifying row in
the window and 0 on every other row for that customer+brand, so a repeat
purchase in August never double-counts toward a rep's placement goal.
Molson Coors classifies Peroni and Coors (Banquet) completely
independently, since a customer could be new-to-Peroni but a longtime
Coors buyer (or vice versa).

Wine & Spirits (Yave/Leyenda) is a simple distinct-buying-account count
per rep per brand family in August -- no new-vs-repeat distinction, per
Kohler: "sales reps need 2 buyers of each brand family to hit this."

Target accounts (Angry Orchard, Molson Coors Peroni/Banquet only -- per
Kohler, 2026-08-04, Wine & Spirits' Yave/Leyenda are sold in every county
so no territory filter applies there): a per-rep "who to go after" list,
answering which of a rep's OWN accounts don't carry the brand yet and are
worth chasing. Built by crossing three sources:
  1. sales_reps_customer_base.csv  -- the rep's full account base (every
     customer they cover), deduped by Customer Num since an account can
     appear more than once with a different Shipping Address. On-Premise
     only (per Kohler, 2026-08-05, via the export's own Premise column)
     -- this dashboard is on-prem, so off-premise accounts in a rep's
     book are never valid targets here even if they don't carry the
     brand yet.
  2. angry_orchard_new_lines.csv / molson_coors_peroni_banquet.csv -- ANY
     customer appearing here at all (new-placement flag or not) already
     has recent purchase history of that brand, so they're excluded --
     this is a "hasn't bought it recently" list, not just "hasn't been
     flagged new this month".
  3. kohler_brands_whitelist_blacklist.xlsx ("Master - US vs THEM" tab)
     -- per-brand, per-county sell authorization. A prospect in a THEM
     county is excluded entirely (Kohler doesn't hold the rights to sell
     that brand there); a prospect in a county with no whitelist row at
     all (e.g. Middlesex, which isn't part of the tracked core territory)
     is kept but marked TERRITORY_STATUS "UNKNOWN" rather than silently
     assumed sellable. Each account's county comes from the workbook's
     "Customers Table (Enc)" sheet (load_customer_area_overrides()), not
     the CSV's own Area column -- the CSV's Area falls back to a "Sales"
     placeholder for accounts missing geographic data on that export
     path (still ~5% of on-premise accounts even after Kohler's 2026-08-05
     refresh), and every one of those resolves to a real county in this
     sheet instead, closing most of the "Sales" UNKNOWN-territory gap.
Molson Coors' Peroni and Coors (Banquet) targets are computed
independently per brand, same as the placement classification.

Run: python3 generate_2026-08.py
"""
import csv
import json
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
DATA_DIR = HERE / "data"
MONTH_KEY = "2026-08"

ANGRY_ORCHARD_CSV = HERE / "angry_orchard_new_lines.csv"
MOLSON_COORS_CSV = HERE / "molson_coors_peroni_banquet.csv"
WINE_SPIRITS_CSV = HERE / "wine_spirits_yave_leyenda.csv"
CUSTOMER_BASE_CSV = HERE / "sales_reps_customer_base.csv"
WHITELIST_XLSX = HERE / "kohler_brands_whitelist_blacklist.xlsx"

NEW_BUYER_WINDOW_START = date(2026, 8, 1)
NEW_BUYER_WINDOW_END = date(2026, 8, 31)


def parse_date(raw):
    return datetime.strptime(raw.strip(), "%m/%d/%Y").date()


def load_csv(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def find_col(fieldnames, prefix):
    return next(c for c in fieldnames if c.startswith(prefix))


def classify_new_placements(rows, brand_key):
    """Shared 90-day-non-buy classifier. `brand_key(row)` returns the
    per-row key (e.g. Brand Family) that purchase history is scoped to --
    a customer's "new" status is evaluated independently per distinct key.
    Returns (output_rows, new_count, total_customer_brand_pairs).
    """
    parsed = [(parse_date(r["Date"]), r) for r in rows]

    by_cust_brand = {}
    for d, r in parsed:
        key = (r["Customer Num"], brand_key(r))
        by_cust_brand.setdefault(key, []).append(d)

    new_keys = set()
    for key, dates in by_cust_brand.items():
        bought_before = any(d < NEW_BUYER_WINDOW_START for d in dates)
        bought_in_window = any(NEW_BUYER_WINDOW_START <= d <= NEW_BUYER_WINDOW_END for d in dates)
        if not bought_before and bought_in_window:
            new_keys.add(key)

    parsed.sort(key=lambda item: item[0])  # earliest first, so the first qualifying row wins
    flagged = set()
    out = []
    for d, r in parsed:
        key = (r["Customer Num"], brand_key(r))
        is_new_row = 0
        if key in new_keys and key not in flagged and NEW_BUYER_WINDOW_START <= d <= NEW_BUYER_WINDOW_END:
            is_new_row = 1
            flagged.add(key)
        out.append((d, r, is_new_row))
    return out, len(new_keys), len(by_cust_brand)


def build_angry_orchard():
    rows = load_csv(ANGRY_ORCHARD_CSV)
    cases_col = find_col(rows[0].keys(), "Cases")
    placement_col = find_col(rows[0].keys(), "Placement Count")
    classified, new_count, total_pairs = classify_new_placements(rows, brand_key=lambda r: r["Brand Family"])
    out = [{
        "SALES_REP_ASSIGNED": r["Sales Rep Assigned"].strip(),
        "CUSTOMER_NUM": int(r["Customer Num"]),
        "CUSTOMER_NAME": r["Customer Name"].strip(),
        "BRAND_FAMILY": r["Brand Family"].strip(),
        "DATE": d.isoformat(),
        "PLACEMENT_COUNT": float(r[placement_col]),
        "CASES": float(r[cases_col]),
        "NEW_PLACEMENT": is_new,
    } for d, r, is_new in classified]
    out.sort(key=lambda row: row["DATE"], reverse=True)
    return out, new_count, total_pairs


def build_molson_coors():
    rows = load_csv(MOLSON_COORS_CSV)
    cases_col = find_col(rows[0].keys(), "Cases")
    placement_col = find_col(rows[0].keys(), "Placement Count")
    classified, new_count, total_pairs = classify_new_placements(rows, brand_key=lambda r: r["Brand Family"])
    out = [{
        "SALES_REP_ASSIGNED": r["Sales Rep Assigned"].strip(),
        "CUSTOMER_NUM": int(r["Customer Num"]),
        "CUSTOMER_NAME": r["Customer Name"].strip(),
        "BRAND_FAMILY": r["Brand Family"].strip(),
        "DATE": d.isoformat(),
        "PLACEMENT_COUNT": float(r[placement_col]),
        "CASES": float(r[cases_col]),
        "NEW_PLACEMENT": is_new,
    } for d, r, is_new in classified]
    out.sort(key=lambda row: row["DATE"], reverse=True)
    return out, new_count, total_pairs


def build_wine_spirits():
    rows = load_csv(WINE_SPIRITS_CSV)
    if not rows:
        return []
    cases_col = find_col(rows[0].keys(), "Cases")
    placement_col = find_col(rows[0].keys(), "Placement Count")
    out = []
    for r in rows:
        out.append({
            "SALES_REP_ASSIGNED": r["Sales Rep Assigned"].strip(),
            "PRODUCT_NAME": r["Product Name"].strip(),
            "PRODUCT_NUM": r["Product Num"].strip(),
            "BRAND_FAMILY": r["Brand Family"].strip(),
            "CUSTOMER_NUM": int(r["Customer Num"]),
            "CUSTOMER_NAME": r["Customer Name"].strip(),
            "DATE": parse_date(r["Date"]).isoformat(),
            "PLACEMENT_COUNT": float(r[placement_col]),
            "CASES": float(r[cases_col]),
        })
    out.sort(key=lambda row: row["DATE"], reverse=True)
    return out


def load_customer_area_overrides():
    """Customer Num -> authoritative Distribution Area, from the
    workbook's "Customers Table (Enc)" sheet. The customer-base CSV
    export's own Area column falls back to a "Sales" (or "Middlesex not
    in use") placeholder for accounts missing geographic data on that
    export path -- every one of those checked against this sheet (2026-
    08-05) resolves to a real county here instead, and every OTHER
    account's Area agrees with this sheet exactly, so this is preferred
    whenever present rather than only patching the placeholder rows."""
    wb = load_workbook(WHITELIST_XLSX, data_only=True)
    ws = wb["Customers Table (Enc)"]
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cust_id, dist_area = row[0], row[7]
        if cust_id is None or not dist_area:
            continue
        lookup[str(cust_id)] = dist_area.strip()
    return lookup


def load_customer_base(area_overrides):
    """Dedupes to one entry per (rep, customer) -- the same account can
    have multiple rows in the export (one per Shipping Address). On-
    Premise only (per Kohler, 2026-08-05) -- the on-prem dashboard's
    Target Accounts shouldn't suggest off-premise accounts."""
    rows = load_csv(CUSTOMER_BASE_CSV)
    by_rep = {}
    seen = set()
    for r in rows:
        rep = r["Sales Rep Assigned"].strip()
        cust_num = r["Customer Num"].strip()
        if not rep or not cust_num:
            continue
        if r["Premise"].strip() != "On Premise":
            continue
        key = (rep, cust_num)
        if key in seen:
            continue
        seen.add(key)
        area = area_overrides.get(cust_num) or r["Area"].strip()
        by_rep.setdefault(rep, []).append({
            "customer_num": cust_num,
            "customer_name": r["Customer Name"].strip(),
            "area": area,
        })
    return by_rep


def load_whitelist():
    """(brand family lowercased, county uppercased) -> 'US' or 'THEM',
    from the flat long-format sheet (one row per brand+county)."""
    wb = load_workbook(WHITELIST_XLSX, data_only=True)
    ws = wb["Master - US vs THEM"]
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        brand, county, determination = row[0], row[1], row[5]
        if not brand or not county or not determination:
            continue
        lookup[(brand.strip().lower(), county.strip().upper())] = determination.strip().upper()
    return lookup


def already_carrying(path, brand_filter=None):
    """Customer Nums with ANY row in a brand's raw export -- recent
    purchase history, whether or not that row was flagged NEW_PLACEMENT."""
    rows = load_csv(path)
    out = set()
    for r in rows:
        if brand_filter and r.get("Brand Family", "").strip().lower() != brand_filter.lower():
            continue
        out.add(r["Customer Num"].strip())
    return out


def build_targets(customer_base_by_rep, whitelist, carrying, brand_label):
    out = []
    unknown_territory = 0
    for rep, accounts in customer_base_by_rep.items():
        for a in accounts:
            if a["customer_num"] in carrying:
                continue
            status = whitelist.get((brand_label.lower(), a["area"].upper()), "UNKNOWN")
            if status == "THEM":
                continue
            if status == "UNKNOWN":
                unknown_territory += 1
            out.append({
                "SALES_REP_ASSIGNED": rep,
                "CUSTOMER_NUM": int(a["customer_num"]) if a["customer_num"].isdigit() else a["customer_num"],
                "CUSTOMER_NAME": a["customer_name"],
                "AREA": a["area"],
                "TERRITORY_STATUS": status,
            })
    out.sort(key=lambda row: (row["SALES_REP_ASSIGNED"], row["CUSTOMER_NAME"]))
    return out, unknown_territory


def main():
    angry_orchard_rows, ao_new, ao_total = build_angry_orchard()
    molson_coors_rows, mc_new, mc_total = build_molson_coors()
    wine_spirits_rows = build_wine_spirits()

    area_overrides = load_customer_area_overrides()
    customer_base_by_rep = load_customer_base(area_overrides)
    whitelist = load_whitelist()

    targets_angry_orchard, ao_unknown = build_targets(
        customer_base_by_rep, whitelist, already_carrying(ANGRY_ORCHARD_CSV), "Angry Orchard")

    targets_peroni, peroni_unknown = build_targets(
        customer_base_by_rep, whitelist, already_carrying(MOLSON_COORS_CSV, "Peroni"), "Peroni")
    for row in targets_peroni:
        row["BRAND_FAMILY"] = "Peroni"
    targets_coors, coors_unknown = build_targets(
        customer_base_by_rep, whitelist, already_carrying(MOLSON_COORS_CSV, "Coors"), "Coors")
    for row in targets_coors:
        row["BRAND_FAMILY"] = "Coors"
    targets_molson_coors = sorted(targets_peroni + targets_coors, key=lambda r: (r["SALES_REP_ASSIGNED"], r["BRAND_FAMILY"], r["CUSTOMER_NAME"]))

    month_dir = DATA_DIR / MONTH_KEY
    month_dir.mkdir(parents=True, exist_ok=True)
    (month_dir / "mpo_angry_orchard.json").write_text(json.dumps(angry_orchard_rows, indent=2))
    (month_dir / "mpo_molson_coors.json").write_text(json.dumps(molson_coors_rows, indent=2))
    (month_dir / "mpo_wine_spirits_yave_leyenda.json").write_text(json.dumps(wine_spirits_rows, indent=2))
    (month_dir / "mpo_targets_angry_orchard.json").write_text(json.dumps(targets_angry_orchard, indent=2))
    (month_dir / "mpo_targets_molson_coors.json").write_text(json.dumps(targets_molson_coors, indent=2))

    synced_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    (month_dir / "sync_meta.json").write_text(json.dumps({"synced_at": synced_at}, indent=2))

    print(f"Angry Orchard: {ao_new} new placements out of {ao_total} customer+brand pairs "
          f"({len(angry_orchard_rows)} transaction rows written)")
    print(f"Molson Coors (Peroni+Coors/Banquet combined): {mc_new} new placements out of {mc_total} "
          f"customer+brand pairs ({len(molson_coors_rows)} transaction rows written)")
    print(f"Wine & Spirits (Yave/Leyenda): {len(wine_spirits_rows)} rows written")
    print(f"Target accounts -- Angry Orchard: {len(targets_angry_orchard)} prospects across all reps "
          f"({ao_unknown} in unmapped/no-whitelist-data territory)")
    print(f"Target accounts -- Molson Coors: {len(targets_peroni)} Peroni + {len(targets_coors)} Coors/Banquet "
          f"prospects ({peroni_unknown}/{coors_unknown} in unmapped territory)")
    print(f"sync_meta.json timestamped {synced_at} in data/{MONTH_KEY}/")


if __name__ == "__main__":
    main()
