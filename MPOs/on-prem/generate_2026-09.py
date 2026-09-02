#!/usr/bin/env python3
"""Builds data/2026-09/*.json for the September 2026 on-premise MPO tab.

September's four objectives (25% each, from September_ON_PREM_2026_MPO.docx):

  1. Lofted Spirits  - (5) New Bardstown Menu Placements   bardstown_menu_promos.xlsx
  2. Molson Coors    - Fever Tree (3) New Placements       fever_tree.csv
  3. Spirits         - Carbliss (10) New On Premise
                       Buying Accounts                     carbliss.csv
  4. HUSA            - (1) New XX Draft Line                husa_xx_draft.csv

Objective 1 has no RDE export -- it is verified from iSellBeer promo photos
instead, which is a DIFFERENT KIND of source from the other three: a partial
weekly pull that must be MERGED onto a cumulative archive rather than
overwritten (repo CLAUDE.md). See build_bardstown_menu() and --merge-bardstown.

WHY THIS IS A SEPARATE SCRIPT FROM generate_2026-08.py: September's exports
changed shape in three ways that would each have broken the August code, and
the repo's convention is a script per month rather than one branching script.

  1. Customer Num and Customer Name arrive as ONE combined column,
     "Customer Num & Company" ("24038 J. Alexander's Restaurant"), where
     August had them separate. split_customer() pulls them apart on the
     leading digits.
  2. Fever Tree and Carbliss carry NO Date column at all. The client's
     buildNewAccountsDataset() returns null outright when it cannot find a
     date column, which would silently blank the objective, so those two
     datasets are stamped with a placeholder DATE of the current window's
     start -- the same trick off-prem's Corona Premier export already uses
     for the same reason. HUSA does carry real dates and keeps them.
  3. The premise column is "On-Off Premise", not "Premise".

NEW-PLACEMENT RULE is unchanged from August: a customer is NEW when the
current-period column (9/1-9/30) is populated and the base-period column
(6/1-8/31) is not. Classification is per (rep, customer) -- none of these
three objectives splits by brand, unlike August's Molson Coors.

NO TARGET ACCOUNTS this month. August built them for Angry Orchard and
Peroni/Banquet because Kohler had confirmed those sell only in the six
Core Market counties. Fever Tree, Carbliss and Dos Equis draft have no
confirmed territory scope, and the README is explicit that the pill and the
prospect list are claims a rep acts on and are never guessed. Add them here
once scope is confirmed.

Run: python3 generate_2026-09.py
"""
import csv
import importlib.util
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
DATA_DIR = HERE / "data"
MONTH_KEY = "2026-09"

FEVER_TREE_CSV = HERE / "fever_tree_new_placements.csv"
CARBLISS_CSV = HERE / "carbliss_new_on_prem_buyers.csv"
HUSA_CSV = HERE / "husa_xx_draft.csv"
# Cumulative iSellBeer promo ARCHIVE for objective 1, not a scratch copy of the
# latest pull -- see build_bardstown_menu().
BARDSTOWN_XLSX = HERE / "bardstown_menu_promos.xlsx"
# The off-prem Lytt POS tracker already solved partial-iSellBeer merging
# (hyperlinks preserved, header-name column matching, volatile counter columns
# ignored). Reused rather than reimplemented.
LYTT_POS_PY = HERE.parent / "off-prem" / "generate_lytt_pos.py"
CUSTOMER_BASE_CSV = HERE / "sales_reps_customer_base.csv"

ROSTER = ["Alex Rodriguez", "Alisa Acciardi", "Allison Scott", "Andrew Lundy",
          "Anthony Palmisano", "Brian Sengebush", "Chris Payton", "Dan Lagala",
          "Dave Ehlers", "Derrick Laws", "Dylan Rubino", "Hakan Sadik",
          "Jaime Colonna", "Javier Melo", "Jayson Romine", "Jim Heaney",
          "John O'Donoghue", "Klejdi Lamo", "Matt Powierski", "Michael Harboy",
          "Mike Ast", "Nick Melissari", "Pablo Lopez", "Paul Mclaughlin",
          "Phil Ernst", "Robin Feldman", "Shane Barreca"]


def load_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def to_num(raw):
    raw = (raw or "").strip()
    if raw == "":
        return None
    try:
        return int(raw)
    except ValueError:
        return float(raw)


def split_customer(raw):
    """"24038 J. Alexander's Restaurant" -> ("24038", "J. Alexander's Restaurant").

    September merged the two columns August had separate. Anything without a
    leading number keeps the whole string as the name and gets no number, so a
    format change shows up as a missing id rather than a crash.
    """
    m = re.match(r"\s*(\d+)\s+(.*)$", raw or "")
    if not m:
        return None, (raw or "").strip()
    return m.group(1), m.group(2).strip()


def find_period_cols(fieldnames, prefix):
    """The two dated columns sharing `prefix`, ordered base then current.

    Same approach as August's: picked apart by each header's embedded START
    date rather than its exact text, so a shifted day-of-month still resolves.
    """
    cols = [f for f in fieldnames if f.startswith(prefix) and re.search(r"\d{1,2}/\d{1,2}/\d{4}", f)]

    def start(col):
        mo, da, yr = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", col).groups()
        return datetime(int(yr), int(mo), int(da))

    cols.sort(key=start)
    if len(cols) != 2:
        raise SystemExit(f"Expected exactly 2 '{prefix}' columns (base + current), found: {cols}")
    return cols[0], cols[1]


def window_start(col):
    mo, da, yr = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", col).groups()
    return datetime(int(yr), int(mo), int(da)).date().isoformat()


def load_off_premise_only_ids():
    """Customer Nums that never appear as "On Premise" in the customer base.

    Per Kohler, 2026-08-07: off-premise accounts are never shown on this
    dashboard. Kept for September even though these three exports look
    on-premise already -- Fever Tree and Carbliss carry an On-Off Premise
    column and HUSA is draft kegs -- because the rule is about the account,
    not about what a given export happens to contain.
    """
    if not CUSTOMER_BASE_CSV.exists():
        print("  NOTE: no sales_reps_customer_base.csv -- off-premise exclusion skipped")
        return set()
    by_cust = {}
    for r in load_csv(CUSTOMER_BASE_CSV):
        cust = r["Customer Num"].strip()
        if cust:
            by_cust.setdefault(cust, set()).add(r["Premise"].strip())
    return {c for c, p in by_cust.items() if p == {"Off Premise"}}


def classify(rows, base_col, current_col, off_premise_ids):
    """(rep, customer) -> "new" / "rebuy" / "base_only", plus the parsed rows.

    NEW means the current window is populated and the base window is not.
    A populated cell counts even when its value is 0, matching August: the
    question is whether the account transacted in that window at all.
    """
    parsed, state = [], {}
    for r in rows:
        num, name = split_customer(r.get("Customer Num & Company"))
        if num and num in off_premise_ids:
            continue
        rep = r["Sales Rep Assigned"].strip()
        key = (rep, num or name)
        has_base = (r.get(base_col) or "").strip() != ""
        has_cur = (r.get(current_col) or "").strip() != ""
        st = state.setdefault(key, {"base": False, "current": False})
        st["base"] = st["base"] or has_base
        st["current"] = st["current"] or has_cur
        parsed.append({"row": r, "rep": rep, "num": num, "name": name,
                       "key": key, "has_base": has_base, "has_current": has_cur})
    status = {k: ("rebuy" if v["base"] and v["current"]
                  else "new" if v["current"]
                  else "base_only")
              for k, v in state.items()}
    return parsed, status


def emit(parsed, status, base_col, current_col, value_key, extra=None, date_col=None):
    """Shared row shape. One NEW_PLACEMENT=1 per new account, on its first
    current-period row, so a second visit never double-counts."""
    placeholder = window_start(current_col)
    flagged, out = set(), []
    for p in parsed:
        if p["has_current"]:
            period = "current"
        elif p["has_base"]:
            period = "base"
        else:
            continue
        is_new = 0
        if period == "current" and status[p["key"]] == "new" and p["key"] not in flagged:
            is_new = 1
            flagged.add(p["key"])
        raw_date = (p["row"].get(date_col) or "").strip() if date_col else ""
        if raw_date:
            date = datetime.strptime(raw_date, "%m/%d/%Y").date().isoformat()
        else:
            # No Date column on this export -- stamp the window start so the
            # client's date-column check passes. See the module docstring.
            date = placeholder
        row = {
            "SALES_REP_ASSIGNED": p["rep"],
            "CUSTOMER_NUM": int(p["num"]) if p["num"] else None,
            "CUSTOMER_NAME": p["name"],
            "BRAND_FAMILY": (p["row"].get("Brand Family") or "").strip(),
            "DATE": date,
            "PERIOD": period,
            value_key: to_num(p["row"][current_col if period == "current" else base_col]),
            "NEW_PLACEMENT": is_new,
        }
        if extra:
            row.update(extra(p["row"]))
        out.append(row)
    out.sort(key=lambda r: r["DATE"], reverse=True)
    return out


def build_fever_tree(off_premise_ids):
    rows = load_csv(FEVER_TREE_CSV)
    base, cur = find_period_cols(rows[0].keys(), "Placement Count")
    parsed, status = classify(rows, base, cur, off_premise_ids)
    out = emit(parsed, status, base, cur, "PLACEMENT_COUNT")
    return out, sum(1 for s in status.values() if s == "new"), len(status)


def build_carbliss(off_premise_ids):
    rows = load_csv(CARBLISS_CSV)
    base, cur = find_period_cols(rows[0].keys(), "Buyer Count")
    parsed, status = classify(rows, base, cur, off_premise_ids)
    out = emit(parsed, status, base, cur, "BUYER_COUNT")
    return out, sum(1 for s in status.values() if s == "new"), len(status)


def build_husa(off_premise_ids):
    """HUSA is the only September export with real dates and a Package column,
    so both are carried through -- the drill-down can show when a line went in
    and what size keg it was."""
    rows = load_csv(HUSA_CSV)
    base, cur = find_period_cols(rows[0].keys(), "Buyer Count")
    units_base, units_cur = find_period_cols(rows[0].keys(), "Units")
    parsed, status = classify(rows, base, cur, off_premise_ids)
    out = emit(parsed, status, base, cur, "BUYER_COUNT",
               extra=lambda r: {"PACKAGE": (r.get("Package") or "").strip(),
                                "UNITS": to_num(r.get(units_cur) if (r.get(units_cur) or "").strip()
                                                else r.get(units_base))},
               date_col="Date")
    return out, sum(1 for s in status.values() if s == "new"), len(status)


def _lytt_pos():
    spec = importlib.util.spec_from_file_location("lytt_pos", LYTT_POS_PY)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def build_bardstown_menu():
    """Objective 1 -- (5) New Bardstown Menu Placements, from iSellBeer promos.

    COUNTS DISTINCT SUBMISSIONS, NOT ROWS. One promo carries one row per brand
    on the menu -- the first pull is a single table tent at Hilton Hasbrouck
    Heights that lists two Bardstown SKUs, arriving as Promo # 1.1 and 1.2. That
    is one menu placement, not two, and it is counted as one: the same rule the
    display auction uses for photos ("one photo showing five Lytt items is ONE
    pic"). A submission is (photo taker + account + date/time).

    The sister program in incentive-tracking pays "per printed menu MENTION,
    multiple mentions on one menu means multiple payouts" -- a deliberately
    different unit. If this MPO objective turns out to be scored the same way,
    the fix is to drop the dedupe and flag every row; both counts are printed
    at build time so the gap is visible. Flagged for Gavin.

    Every submission counts as new: the promos export is a single window with no
    base period, so a menu placement submitted this month IS the new placement.
    """
    if not BARDSTOWN_XLSX.exists():
        print("  Bardstown menu: no bardstown_menu_promos.xlsx -- objective stays rules-only")
        return [], 0, 0
    ws = openpyxl.load_workbook(BARDSTOWN_XLSX)["Report"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header) if h}
    roster_by_lower = {r.lower(): r for r in ROSTER}

    seen, out, mentions = set(), [], 0
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        if not vals or not vals[idx["Date/Time"]]:
            continue
        raw_rep = str(vals[idx["Photo taker"]] or "").strip()
        # iSellBeer spells names its own way ("robin feldman"); the roster is
        # the RDE spelling. Unmatched names are kept as-is so they surface
        # rather than vanish.
        rep = roster_by_lower.get(raw_rep.lower(), raw_rep)
        dt = str(vals[idx["Date/Time"]]).strip()
        acct = str(vals[idx["Account #"]] or "").strip()
        key = (rep, acct, dt)
        mentions += 1
        photo_cell = row[idx["Photo"]]
        out.append({
            "SALES_REP_ASSIGNED": rep,
            "CUSTOMER_NUM": int(acct) if acct.isdigit() else None,
            "CUSTOMER_NAME": str(vals[idx["DBA"]] or "").strip(),
            "BRAND_FAMILY": str(vals[idx["Brand"]] or "").strip(),
            "DATE": datetime.strptime(dt.split()[0], "%m/%d/%Y").date().isoformat(),
            "PERIOD": "current",
            "PROMOTION_TYPE": str(vals[idx["Promotion type"]] or "").strip(),
            "ELEMENTS": str(vals[idx["Elements"]] or "").strip(),
            "PHOTO_URL": photo_cell.hyperlink.target if photo_cell.hyperlink else None,
            # One flag per distinct submission, on its first row.
            "NEW_PLACEMENT": 0 if key in seen else 1,
        })
        seen.add(key)
    out.sort(key=lambda r: r["DATE"], reverse=True)
    return out, len(seen), mentions


def main():
    if len(sys.argv) == 3 and sys.argv[1] == "--merge-bardstown":
        # Partial weekly iSellBeer pulls MERGE onto the archive (repo CLAUDE.md)
        # -- overwriting would drop every menu placement published before this
        # window. "Promo #" is a per-export counter like PODS' "POD #", so it is
        # excluded from the dedupe key or every overlapping row reads as new.
        _lytt_pos().merge_export(BARDSTOWN_XLSX, Path(sys.argv[2]),
                                 date_col="Date/Time", volatile_cols=("Promo #",))

    off_premise_ids = load_off_premise_only_ids()
    fever_rows, fever_new, fever_total = build_fever_tree(off_premise_ids)
    carb_rows, carb_new, carb_total = build_carbliss(off_premise_ids)
    husa_rows, husa_new, husa_total = build_husa(off_premise_ids)
    bard_rows, bard_placements, bard_mentions = build_bardstown_menu()

    month_dir = DATA_DIR / MONTH_KEY
    month_dir.mkdir(parents=True, exist_ok=True)
    (month_dir / "mpo_fever_tree.json").write_text(json.dumps(fever_rows, indent=2))
    (month_dir / "mpo_carbliss.json").write_text(json.dumps(carb_rows, indent=2))
    (month_dir / "mpo_husa_xx_draft.json").write_text(json.dumps(husa_rows, indent=2))
    (month_dir / "mpo_bardstown_menu.json").write_text(json.dumps(bard_rows, indent=2))

    synced_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    (month_dir / "sync_meta.json").write_text(json.dumps({"synced_at": synced_at}, indent=2))

    print(f"Off-premise-only customer IDs excluded: {len(off_premise_ids)}")
    print(f"Fever Tree (goal 3): {fever_new} new placements out of {fever_total} accounts "
          f"({len(fever_rows)} rows written) -- no Date column, window-start placeholder stamped")
    print(f"Carbliss (goal 10): {carb_new} new buying accounts out of {carb_total} accounts "
          f"({len(carb_rows)} rows written) -- no Date column, window-start placeholder stamped")
    print(f"HUSA XX draft (goal 1): {husa_new} new draft lines out of {husa_total} accounts "
          f"({len(husa_rows)} rows written)")
    print(f"Bardstown menu (goal 5): {bard_placements} distinct menu placements from "
          f"{bard_mentions} brand mentions across {len(bard_rows)} promo rows "
          f"-- counted per SUBMISSION; per-mention would read {bard_mentions} (unconfirmed)")
    print(f"sync_meta.json timestamped {synced_at} in data/{MONTH_KEY}/")


if __name__ == "__main__":
    main()
