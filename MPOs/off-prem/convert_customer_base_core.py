#!/usr/bin/env python3
"""Flattens RDE's "Sales Reps: Customer Base Core Off Prem" workbook into
sales_reps_customer_base_core.csv, the Core Market off-premise account base.

WHY A CONVERTER RATHER THAN A SAVE-AS. The workbook is HIERARCHICAL: column B
("Sales Rep Assigned / Customer Num") holds a rep NAME on a grouping row and a
customer NUMBER on each of that rep's account rows, with a "Total" row on top
and a Rank column down the side. Saving it straight to CSV would hand every
consumer a rep column full of customer numbers. So the rep is carried down onto
its accounts, Total/Rank are dropped, and the eleven columns come out in the
order the CSV has always had.

WHAT IT GUARDS. This file is a DENOMINATOR -- Keystone Ice's penetration is
scored against it and both Target Accounts lists are scoped by it -- so a
silently truncated or mis-parsed pull would move every rep's number without
erroring. The checks below refuse to write rather than publish a bad base:
every row must be Off Premise, every account must be unique per rep, every
area must be one of the six Core Market areas (or RDE's "Sales" placeholder,
which the consumers already fall back to County for), and the row count must
be in a sane band. Compare against the previous file before committing --
--dry-run prints the adds and drops without writing.

Run: python3 convert_customer_base_core.py <workbook.xlsx> [--dry-run]
"""
import csv
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
OUT_CSV = HERE / "sales_reps_customer_base_core.csv"
SHEET_PREFIX = "Sales Reps Customer Base Core"

# The header the rest of the repo reads. Order matters: generate_2026-08.py and
# generate_2026-09.py both index this file by name, but the incentive tracker's
# legacy copy is positional, so keep the columns where they are.
HEADER = ["Sales Rep Assigned", "Customer Num", "Customer Name", "Shipping Address",
          "Distribution Area", "Area", "County", "City", "Premise",
          "Buyer Count   2026", "Cases   2026"]

# Per Kohler, 2026-08-06: these accounts are only ever sold in the six Core
# Market areas. "Sales" is RDE's placeholder on rows with no geographic data --
# consumers fall back to the County column for those (see load_core_customer_base()).
CORE_AREAS = {"Bergen", "Passaic", "Passaic-FF", "Morris 1", "Morris 3", "Sussex", "Sales"}
MIN_ROWS, MAX_ROWS = 400, 700


def flatten(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if s.strip().startswith(SHEET_PREFIX)), None)
    if sheet is None:
        raise SystemExit(f"No '{SHEET_PREFIX}...' sheet in {path.name}; found {wb.sheetnames}")
    ws = wb[sheet]

    rows, rep, reps = [], None, []
    for r in ws.iter_rows(values_only=True):
        key = r[1]
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        if key == "Sales Rep Assigned / Customer Num":   # header
            continue
        if not key.isdigit():
            # A grouping row: a rep name, or the report-wide "Total" above them.
            if key != "Total":
                rep = key
                reps.append(rep)
            continue
        if rep is None:
            raise SystemExit(f"Customer row {key} appears before any rep grouping row "
                             f"-- the workbook's shape changed, check column B.")
        rows.append([rep, key, r[2] or "", r[3] or "", r[4] or "", r[5] or "",
                     r[6] or "", r[7] or "", r[8] or "", r[9], r[10]])
    return rows, reps


def check(rows):
    if not MIN_ROWS <= len(rows) <= MAX_ROWS:
        raise SystemExit(f"{len(rows)} account rows is outside the expected "
                         f"{MIN_ROWS}-{MAX_ROWS} band -- looks like a partial or wrong export.")
    bad_premise = {r[8] for r in rows} - {"Off Premise"}
    if bad_premise:
        raise SystemExit(f"This base is off-premise only, but found Premise values {bad_premise}. "
                         f"Wrong export? (The on-prem MPO tracker needs a BOTH-premise file.)")
    bad_area = {r[4] for r in rows} - CORE_AREAS
    if bad_area:
        raise SystemExit(f"Non-Core-Market Distribution Area values {bad_area} -- this file is "
                         f"supposed to be pre-filtered to the six Core Market areas.")
    seen = set()
    dupes = {(r[0], r[1]) for r in rows if (r[0], r[1]) in seen or seen.add((r[0], r[1]))}
    if dupes:
        raise SystemExit(f"{len(dupes)} rep+customer pairs appear twice, e.g. {list(dupes)[:3]}. "
                         f"Older pulls carried one row per shipping address; if that is back, "
                         f"dedupe here before writing.")


def previous_keys():
    if not OUT_CSV.exists():
        return None
    with open(OUT_CSV, newline="", encoding="utf-8-sig") as f:
        return {(r["Sales Rep Assigned"], r["Customer Num"]): r["Customer Name"]
                for r in csv.DictReader(f)}


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry = "--dry-run" in sys.argv
    if len(args) != 1:
        raise SystemExit(__doc__.strip().splitlines()[-1])
    rows, reps = flatten(Path(args[0]))
    check(rows)

    prev = previous_keys()
    if prev is not None:
        now = {(r[0], r[1]): r[2] for r in rows}
        added = [f"{k[0]} / {k[1]} {now[k]}" for k in now if k not in prev]
        dropped = [f"{k[0]} / {k[1]} {prev[k]}" for k in prev if k not in now]
        print(f"vs the current CSV: {len(added)} account(s) added, {len(dropped)} dropped, "
              f"{len(set(now) & set(prev))} unchanged")
        for a in added:
            print(f"  + {a}")
        for d in dropped:
            print(f"  - {d}")

    print(f"{len(rows)} account rows across {len(reps)} reps")
    if dry:
        print("--dry-run: nothing written")
        return
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADER)
        w.writerows(rows)
    print(f"wrote {OUT_CSV.relative_to(HERE.parent.parent)}")
    print("Now rerun the CURRENT month's generator (generate_2026-09.py). Do NOT rerun a "
          "closed month's -- its tab is a published snapshot, see README.txt.")


if __name__ == "__main__":
    main()
