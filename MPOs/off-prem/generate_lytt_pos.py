#!/usr/bin/env python3
"""Builds data/2026-08/mpo_lytt_photos.json -- the "Disruptors – (8) Lytt
POS Items Pics in iSellBeer" objective's dataset -- from three iSellBeer
photo exports saved under stable names in this folder:

  lytt_pos_displays.xlsx   iSellBeer "Report" tab export (Report_NN.xlsx),
                           Lytt-filtered: Photo Taker, Photo Taker's Role,
                           Account #, DBA, City, Brand, SKU, Quantity,
                           Date/Time, Photo (hyperlinked).
  lytt_pos_promos.xlsx     iSellBeer Promos export (Promos_Report_N.xlsx):
                           Photo taker, Photo taker's role, DBA, City,
                           Promotion type, Theme, Elements, Brand,
                           Quantity, Date/Time, Photo (hyperlinked).
  lytt_pos_pods.xlsx       iSellBeer PODS export (PODS_Report_N.xlsx):
                           Route / Sales Rep ("6 - James Heaney" format),
                           DBA, City, Brand, SKU, POD #, Photo -- only a
                           handful of POD rows carry a photo hyperlink;
                           the rest are distro records and are skipped.

Rules (confirmed with Gavin, 2026-08-19):
  - SALES REPS ONLY. Displays/promos rows are filtered by the export's own
    Role column; PODS has no role column but its "Route / Sales Rep" field
    is rep-only by construction. Names are canonicalized to the dashboard
    ROSTER spelling ("James Heaney" -> "Jim Heaney", "phil Ernst" ->
    "Phil Ernst", "Matthew Powierski" -> "Matt Powierski", "Daniel La
    Gala" -> "Dan Lagala", and PODS' "6 - " route prefix stripped); a name
    that still doesn't match the roster is warned about and skipped.
  - Only rows whose Photo cell carries a real hyperlink are written --
    a row without a clickable photo doesn't count for anything here.
  - The 8-pic target is counted in DISTINCT PHOTOS client-side (several
    SKU rows sharing one photo link = one pic -- confirmed 2026-08-19,
    "each distinct photo", NOT each row). Every photo-bearing row is
    still written so the dashboard can show which items are in each pic.
  - Only Lytt rows count (the exports are already Lytt-filtered at the
    source; a non-LYTT brand slipping in is warned about and dropped).

Refresh -- the stable workbooks above are the ARCHIVE, not a scratch copy.
Gavin pulls one week at a time (see the repo CLAUDE.md), so a fresh export
usually covers only its own window and saving it over the stable filename
would silently drop every earlier photo. Merge it in instead:

  python3 generate_lytt_pos.py --merge-displays Report_NN.xlsx
  python3 generate_lytt_pos.py --merge-promos Promos_Report_N.xlsx
  python3 generate_lytt_pos.py --merge-pods PODS_Report_N.xlsx

Either flag unions the incoming rows into the matching stable workbook
(deduped, re-sorted newest-first, "#" renumbered, hyperlinks and the
Filters tab's date span carried over) and then rebuilds the JSON as usual.
Re-merging an export already applied is a no-op. Only save an export
straight over a stable filename when it covers the WHOLE tracked period
(08/01/2026 onward). PODS used to be exactly that -- an undated full snapshot
-- but as of PODS_Report_12 (2026-08-24) it is a windowed pull like the others
(one day, and it now carries a Date/Time column), so it merges too rather than
being overwritten. Then commit and push. Does NOT touch
generate_2026-08.py's CSVs/outputs or sync_meta.json (the "Data refreshed"
pill tracks the RDE sync, not this photo feed).
"""
import copy
import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
OUT = HERE / "data" / "2026-08" / "mpo_lytt_photos.json"

DISPLAYS_XLSX = HERE / "lytt_pos_displays.xlsx"
PROMOS_XLSX = HERE / "lytt_pos_promos.xlsx"
PODS_XLSX = HERE / "lytt_pos_pods.xlsx"

# Same 27 names as index.html's ROSTER -- keep in sync if the roster changes.
ROSTER = [
    "Alex Rodriguez", "Alisa Acciardi", "Allison Scott", "Andrew Lundy",
    "Anthony Palmisano", "Brian Sengebush", "Chris Payton", "Dan Lagala",
    "Dave Ehlers", "Derrick Laws", "Dylan Rubino", "Hakan Sadik",
    "Jaime Colonna", "Javier Melo", "Jayson Romine", "Jim Heaney",
    "John O'Donoghue", "Klejdi Lamo", "Matt Powierski", "Michael Harboy",
    "Mike Ast", "Nick Melissari", "Pablo Lopez", "Paul Mclaughlin",
    "Phil Ernst", "Robin Feldman", "Shane Barreca",
]
ROSTER_BY_UPPER = {n.upper(): n for n in ROSTER}
# iSellBeer spellings that differ from the RDE/roster spelling.
NAME_FIXES = {
    "JAMES HEANEY": "Jim Heaney",
    "MATTHEW POWIERSKI": "Matt Powierski",
    "DANIEL LA GALA": "Dan Lagala",
    "NICHOLAS MELISSARI": "Nick Melissari",
    "PAUL MCLAUGHLIN": "Paul Mclaughlin",
}


def canon_rep(name):
    n = re.sub(r"^\d+\s*-\s*", "", str(name or "").strip())  # PODS "6 - " route prefix
    if not n:
        return None
    up = re.sub(r"\s+", " ", n).upper()
    return NAME_FIXES.get(up) or ROSTER_BY_UPPER.get(up)


def parse_dt(s):
    return datetime.datetime.strptime(str(s).strip(), "%m/%d/%Y %I:%M %p")


def merge_export(stable, incoming, date_col="Date/Time", volatile_cols=()):
    """Union a PARTIAL iSellBeer photo export into its cumulative stable
    workbook, keeping every row already published.

    Gavin pulls one week at a time to keep each upload small (see the repo
    CLAUDE.md), so an incoming export covers only its own window -- saving it
    over the stable filename would silently drop every earlier photo. This
    unions the two instead: the stable workbook IS the archive, so the JSON
    stays a purely derived artifact that can always be rebuilt from it.

    Columns are matched by HEADER NAME, not position, so an export that grows
    a column still merges: any column the incoming export has and the archive
    doesn't is appended to the archive and left blank on rows published before
    it existed (PODS grew a Date/Time this way on 2026-08-24). A column going
    the other way -- present in the archive, gone from the export -- is a real
    format regression and stops the merge instead.

    Rows are deduped on the columns the archive ALREADY had (plus the photo
    link), ignoring the "#" counter where the export has one, plus any
    volatile_cols the caller names. So re-merging an export already applied is a
    no-op, an overlapping export updates nothing it already has, and a row
    re-sent with a newly-added column populated matches its published copy
    instead of landing twice.

    volatile_cols exists for PODS' "POD #", found 2026-08-27: it is a sequence
    number scoped to the export's own window, not a property of the row, so the
    SAME purchase came back as 6.1 in one pull and 28.1 in the next and every
    overlapping row read as new (PODS_Report_15 reported 45 of 45 rows new and
    re-added 3 already-published photos). Same reason "#" is ignored; the
    archive keeps whichever value it already had, and nothing downstream reads
    either column.

    Rows carrying a date sort newest-first; undated ones (PODS rows published
    before it grew a Date/Time) keep their existing order at the bottom.
    """
    wb = openpyxl.load_workbook(stable)
    ws = wb["Report"]
    wb_in = openpyxl.load_workbook(incoming)
    ws_in = wb_in["Report"]

    headers = [c.value for c in ws[1]]
    headers_in = [c.value for c in ws_in[1]]
    dropped = [h for h in headers if h not in headers_in]
    if dropped:
        raise SystemExit(f"{incoming.name} no longer has column(s) {dropped} that "
                         f"{stable.name} carries -- the export format changed; "
                         f"reconcile by hand.")
    gained = [h for h in headers_in if h not in headers]
    final = headers + gained
    idx_date = final.index(date_col) if date_col in final else None

    def collect(sheet, hdrs):
        out = []
        for r in sheet.iter_rows(min_row=2):
            vals = {h: r[i].value for i, h in enumerate(hdrs) if h is not None}
            if all(v is None for v in vals.values()):
                continue
            links = {hdrs[i]: c.hyperlink.target for i, c in enumerate(r)
                     if c.hyperlink and hdrs[i] is not None}
            out.append((vals, links))
        return out

    published = collect(ws, headers)
    incoming_rows = collect(ws_in, headers_in)

    ignored = {"#", *volatile_cols}

    def key(row):
        vals, links = row
        return (tuple(vals.get(h) for h in headers if h not in ignored)
                + tuple(sorted(links.items())))

    seen = {key(r) for r in published}
    added = [r for r in incoming_rows if key(r) not in seen]
    rows = published + added

    def when(row):
        raw = row[0].get(date_col)
        return parse_dt(raw) if raw else None

    # Stable sort: dated rows newest-first, undated ones left in place after them.
    rows.sort(key=lambda r: (when(r) is not None, when(r) or datetime.datetime.min),
              reverse=True)

    if gained:
        print(f"  {incoming.name} added column(s) {gained}; widening "
              f"{stable.name} (blank on rows published before them).")
    print(f"Merge: {incoming.name} -> {stable.name}: {len(incoming_rows)} row(s) in, "
          f"{len(added)} new, {len(incoming_rows)-len(added)} already published; "
          f"{len(published)} kept -> {len(rows)} total.")

    pub_dates = [when(r) for r in published if when(r)]
    new_dates = [when(r) for r in added if when(r)]
    if pub_dates and new_dates:
        last_pub = max(pub_dates).date()
        first_new = min(new_dates).date()
        gap = [last_pub + datetime.timedelta(days=i) for i in range(1, (first_new - last_pub).days)]
        missed = [d for d in gap if d.weekday() < 5]
        if missed:
            print(f"  WARNING: no rows for {len(missed)} weekday(s) between the last published "
                  f"row ({last_pub:%m/%d/%Y}) and this export's first new one "
                  f"({first_new:%m/%d/%Y}): {', '.join(d.strftime('%m/%d') for d in missed)}. "
                  f"Photos submitted then are NOT on the board and won't arrive on their own -- "
                  f"re-pull from {missed[0]:%m/%d/%Y} if that gap wasn't just a quiet stretch.")

    # Style template per column, taken from the first published data row (the
    # incoming export's, for columns the archive is only now growing), so
    # merged-in rows keep the export's look -- notably the blue underlined
    # Photo link cell.
    tpl = {h: copy.copy(c) for h, c in zip(headers, ws[2])}
    if gained:
        tpl_in = {h: copy.copy(c) for h, c in zip(headers_in, ws_in[2])}
        hdr_style = copy.copy(ws.cell(row=1, column=1))
        for j, h in enumerate(final, start=1):
            if h in gained:
                hc = ws.cell(row=1, column=j)
                hc.value = h
                hc._style = hdr_style._style
                tpl[h] = tpl_in[h]

    idx_num = final.index("#") if "#" in final else None
    for i, (vals, links) in enumerate(rows, start=2):
        for j, h in enumerate(final, start=1):
            c = ws.cell(row=i, column=j)
            c.value = i - 1 if j - 1 == idx_num else vals.get(h)
            c._style = tpl[h]._style
            c.hyperlink = None
            if h in links:
                c.hyperlink = links[h]
    if ws.max_row > len(rows) + 1:
        ws.delete_rows(len(rows) + 2, ws.max_row - len(rows) - 1)

    # Keep the Filters tab honest: it now spans both exports' windows.
    if "Filters" in wb.sheetnames:
        fws = wb["Filters"]
        spans = {}
        for sheet in (wb["Filters"], wb_in["Filters"]) if "Filters" in wb_in.sheetnames else (wb["Filters"],):
            for row in sheet.iter_rows(values_only=True):
                if row and row[0] in ("Start Date", "End Date") and row[1]:
                    spans.setdefault(row[0], []).append(
                        datetime.datetime.strptime(str(row[1]).strip(), "%m/%d/%Y"))
        for row in fws.iter_rows():
            label = row[0].value
            if label == "Start Date" and spans.get("Start Date"):
                row[1].value = min(spans["Start Date"]).strftime("%m/%d/%Y")
            elif label == "End Date" and spans.get("End Date"):
                row[1].value = max(spans["End Date"]).strftime("%m/%d/%Y")

    wb.save(stable)


def header_map(ws):
    return {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value is not None}


def is_lytt(brand):
    return "LYTT" in str(brand or "").upper()


def main():
    args = sys.argv[1:]
    merges = {"--merge-displays": DISPLAYS_XLSX, "--merge-promos": PROMOS_XLSX,
              "--merge-pods": PODS_XLSX}
    while args:
        flag = args.pop(0)
        if flag not in merges or not args:
            raise SystemExit("Usage: python3 generate_lytt_pos.py "
                             "[--merge-displays Report_NN.xlsx] "
                             "[--merge-promos Promos_Report_N.xlsx] "
                             "[--merge-pods PODS_Report_N.xlsx]")
        # PODS' "POD #" renumbers per export window -- see merge_export().
        volatile = ("POD #",) if flag == "--merge-pods" else ()
        merge_export(merges[flag], Path(args.pop(0)), volatile_cols=volatile)

    rows_out = []
    skipped_names = set()

    def add(rep_raw, source, dba, city, brand, detail, qty, date, photo_cell):
        if photo_cell is None or not photo_cell.hyperlink:
            return
        if not is_lytt(brand):
            print(f"  WARNING: non-Lytt brand row dropped ({brand!r} at {dba!r})")
            return
        rep = canon_rep(rep_raw)
        if not rep:
            skipped_names.add(str(rep_raw))
            return
        rows_out.append({
            "REP": rep,
            "SOURCE": source,
            "CUSTOMER_NAME": str(dba or "").strip(),
            "CITY": str(city or "").strip(),
            "BRAND": str(brand or "").strip(),
            "DETAIL": str(detail or "").strip(),
            "QUANTITY": qty if qty is not None else "",
            "DATE": str(date or "").strip(),
            "PHOTO_URL": photo_cell.hyperlink.target,
        })

    # -- displays (Report_NN "Report" tab) --
    wb = openpyxl.load_workbook(DISPLAYS_XLSX)
    ws = wb["Report"]
    h = header_map(ws)
    n0 = len(rows_out)
    for r in ws.iter_rows(min_row=2):
        if r[h["#"]].value is None:
            continue
        if str(r[h["Photo Taker's Role"]].value or "").strip() != "Sales Rep":
            continue
        add(r[h["Photo Taker"]].value, "Display", r[h["DBA"]].value, r[h["City"]].value,
            r[h["Brand"]].value, r[h["SKU"]].value, r[h["Quantity"]].value,
            r[h["Date/Time"]].value, r[h["Photo"]])
    print(f"Displays: {len(rows_out)-n0} sales-rep photo rows")

    # -- promos --
    wb = openpyxl.load_workbook(PROMOS_XLSX)
    ws = wb["Report"]
    h = header_map(ws)
    n0 = len(rows_out)
    for r in ws.iter_rows(min_row=2):
        if r[h["Date/Time"]].value is None:
            continue
        if str(r[h["Photo taker's role"]].value or "").strip() != "Sales Rep":
            continue
        detail = " · ".join(x for x in [
            str(r[h["Promotion type"]].value or "").strip(),
            str(r[h["Theme"]].value or "").strip(),
            str(r[h["Elements"]].value or "").strip(),
        ] if x)
        add(r[h["Photo taker"]].value, "Promo", r[h["DBA"]].value, r[h["City"]].value,
            r[h["Brand"]].value, detail, r[h["Quantity"]].value,
            r[h["Date/Time"]].value, r[h["Photo"]])
    print(f"Promos: {len(rows_out)-n0} sales-rep photo rows")

    # -- PODS (photo-bearing rows only; no role column, rep-only by construction) --
    wb = openpyxl.load_workbook(PODS_XLSX)
    ws = wb["Report"]
    h = header_map(ws)
    n0 = len(rows_out)
    for r in ws.iter_rows(min_row=2):
        # PODS grew a Date/Time on 2026-08-24; rows archived before that have
        # none, and the dashboard renders a missing date as an em dash.
        date = r[h["Date/Time"]].value if "Date/Time" in h else ""
        add(r[h["Route / Sales Rep"]].value, "POD", r[h["DBA"]].value, r[h["City"]].value,
            r[h["Brand"]].value, r[h["SKU"]].value, None, date, r[h["Photo"]])
    print(f"PODS: {len(rows_out)-n0} photo rows (of {ws.max_row-1} total distro rows)")

    if skipped_names:
        print(f"  WARNING: rows skipped for non-roster names: {sorted(skipped_names)}")

    OUT.write_text(json.dumps(rows_out, indent=1))
    per_rep = {}
    for row in rows_out:
        per_rep.setdefault(row["REP"], set()).add(row["PHOTO_URL"])
    print(f"Wrote {len(rows_out)} rows to {OUT.name}; distinct photos per rep "
          f"(target 8): " + ", ".join(f"{k} {len(v)}" for k, v in
                                      sorted(per_rep.items(), key=lambda kv: -len(kv[1]))))


if __name__ == "__main__":
    main()
