# -*- coding: utf-8 -*-
"""
Rebuild Boston_Beer_Margin_Comparison.xlsx from the Boston Pricing Analysis file.

    python3 make_boston_margin_workbook.py [source.xlsx] [output.xlsx]
    python3 make_boston_margin_workbook.py --fix-views workbook.xlsx

Run --fix-views after ANY LibreOffice recalculation of this workbook. LibreOffice
writes a selection state for the frozen bottom-right pane whose active cell is A1,
a cell that is not in that pane; Excel then opens the sheet with the cursor
parked outside the active pane and the arrow keys appear to do nothing.

Everything that is NOT in the source file -- the DA amounts pulled out of the
free text in the price cells, and the Kohler realized-sales figures from the
7/1-8/31/2026 invoice transaction report -- is hardcoded below in P and ACT,
with the reasoning written onto the workbook's Assumptions tab. Refresh those
two tables when new data lands, then re-run.
"""
import re
import shutil
import sys
import zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.comments import Comment

def fix_sheet_views(path):
    """Give every frozen sheet one valid selection: the top-left cell of its active pane."""
    tmp = path + '.tmp'
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if re.match(r'xl/worksheets/sheet\d+\.xml$', item.filename):
                x = data.decode('utf-8')
                pane = re.search(r'<pane [^>]*/>', x)
                if pane:
                    tl = re.search(r'topLeftCell="([^"]+)"', pane.group(0)).group(1)
                    ap = re.search(r'activePane="([^"]+)"', pane.group(0)).group(1)
                    sel = f'<selection pane="{ap}" activeCell="{tl}" sqref="{tl}"/>'
                else:
                    sel = '<selection activeCell="A1" sqref="A1"/>'
                x = re.sub(r'(<selection [^>]*/>)+', sel, x, count=1)
                x = re.sub(r'<selection [^>]*/>', '', x.replace(sel, '\x00', 1)).replace('\x00', sel)
                data = x.encode('utf-8')
            elif item.filename == 'xl/workbook.xml':
                data = data.decode('utf-8').replace('<workbookProtection/>', '').encode('utf-8')
            zout.writestr(item, data)
    shutil.move(tmp, path)


if len(sys.argv) > 1 and sys.argv[1] == '--fix-views':
    fix_sheet_views(sys.argv[2])
    sys.exit(0)

SRC = sys.argv[1] if len(sys.argv) > 1 else 'Boston_Pricing_Analysis.xlsx'
OUT = sys.argv[2] if len(sys.argv) > 2 else 'Boston_Beer_Margin_Comparison.xlsx'

F = 'Arial'
BLUE  = Font(name=F, size=10, color='0000FF')          # editable input
BLACK = Font(name=F, size=10)                          # calculated
GREEN = Font(name=F, size=10, color='008000')          # link to another sheet
BOLD  = Font(name=F, size=10, bold=True)
TITLE = Font(name=F, size=14, bold=True, color='1F3864')
SUB   = Font(name=F, size=9, italic=True, color='595959')
HDRF  = Font(name=F, size=9, bold=True, color='FFFFFF')
SECF  = Font(name=F, size=11, bold=True, color='1F3864')

HDRFILL = PatternFill('solid', fgColor='1F3864')
INFILL  = PatternFill('solid', fgColor='EAF1FB')   # input cell background
BAND    = PatternFill('solid', fgColor='F5F5F5')
YEL     = PatternFill('solid', fgColor='FFFF00')
KOHL    = PatternFill('solid', fgColor='FFF2CC')   # Kohler row highlight

thin = Side(style='thin', color='BFBFBF')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
TOPRULE = Border(top=Side(style='medium', color='1F3864'))

MONEY = '$#,##0.00;($#,##0.00);"-"'
MONEY0= '$#,##0;($#,##0);"-"'
PCT   = '0.0%;(0.0%);"-"'
NUM   = '#,##0;(#,##0);"-"'

DISTS = ['Kohler', 'Peerless', 'Shore Point', 'Suggested (Boston Beer)']

# ---------------------------------------------------------------- source data
# Boston_Pricing_Analysis.xlsx, Sheet1. Prices are per keg.
# order: name, gallons, old laid-in, new laid-in,
#        current frontline {K,P,SP,Sug}, current best {K,P,SP,Sug},
#        current DA {K,P,SP,Sug}, new suggested FL, new suggested best, new DA
P = [
 ('Sam Adams 15.5',     15.5, 114.61, 117.11, [185,209,207,185], [175,188,188,169], [0,0,0,0],       190, 176, 0),
 ('Sam Adams 5.2',       5.2,  48.12,  49.62, [80,97.5,90,73],   [78,88.5,80,63],   [0,0,0,0],        79,  76, 0),
 ('Twisted Tea 15.5',   15.5, 114.11, 116.11, [185,209,207,184], [175,188,188,168], [0,0,0,0],       189, 184, 0),
 ('Twisted Tea 5.2',     5.2,  48.12,  49.62, [77,97.5,90,77],   [58,88.5,80,58],   [9,None,None,9],  79,  61, 9),
 ('Angry Orchard 15.5', 15.5, 114.61, 117.11, [182,200,207,185], [156,189,188,169], [0,0,0,0],       190, 176, 0),
 ('Angry Orchard 5.2',   5.2,  48.12,  49.62, [75,95,90,77],     [73,87,80,73],     [0,0,0,0],        79,  76, 0),
 ('Dogfish Head 15.5',  15.5, 121.95, 124.45, [185,209,203,189], [180,188,188,161], [0,0,0,0],       194, 167, 0),
 ('Dogfish Head 5.2',    5.2,  59.68,  61.18, [90,97.5,98,87],   [87,88.5,94,83],   [0,0,0,0],        92,  89, 0),
 ('Truly 15.5',         15.5, 114.11, 116.61, [168,None,207,184],[115,None,188,115],[35.40,None,None,35.40], 189, 120, 35.40),
 ('Truly 5.2',           5.2,  48.12,  49.62, [70,97.5,90,77],   [68,88.5,80,67],   [0,0,0,0],        79,  76, 0),
]

# Kohler realized sales, InvoiceTransReport 7/1/2026-8/31/2026 (units, wtd avg $/keg, accounts, %@frontline, %@best-or-lower)
ACT = {
 'Sam Adams 15.5':(1032,180.69,225,0.3372,0.1890),'Sam Adams 5.2':(617,78.62,148,0.3096,0.6872),
 'Twisted Tea 15.5':(28,179.00,13,0.3571,0.5714),'Twisted Tea 5.2':(75,64.41,17,0.2000,0.6133),
 'Angry Orchard 15.5':(151,168.38,42,0.4305,0.4503),'Angry Orchard 5.2':(86,73.84,25,0.4186,0.5814),
 'Dogfish Head 15.5':(259,182.14,72,0.4286,0.5714),'Dogfish Head 5.2':(243,88.16,63,0.3868,0.6132),
 'Truly 15.5':(42,120.05,16,0.0952,0.9048),'Truly 5.2':(33,68.12,14,0.0606,0.9394),
}

wb = openpyxl.Workbook()

def setc(ws, r, c, v, font=BLACK, fmt=None, fill=None, align=None, border=True, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if align or wrap: cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if border: cell.border = BOX
    return cell

def header_row(ws, r, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = HDRF; c.fill = HDRFILL; c.border = BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 34
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[gcl(i)].width = w

def titleblock(ws, title, sub):
    ws['A1'] = title; ws['A1'].font = TITLE
    ws['A2'] = sub;   ws['A2'].font = SUB
    ws.row_dimensions[1].height = 20

# =========================================================== 1. ASSUMPTIONS
a = wb.active; a.title = 'Assumptions'
titleblock(a, 'Boston Beer Margin Comparison — Read Me, Controls & Assumptions',
   'Built from Boston_Pricing_Analysis.xlsx (all pricing per keg) + InvoiceTransReport_GSHARKEY_20260903 (Kohler actuals 7/1/2026–8/31/2026).')
a.column_dimensions['A'].width = 46
a.column_dimensions['B'].width = 20
a.column_dimensions['C'].width = 104

a['A4'] = 'CONTROLS — change these and every margin in the workbook recalculates'
a['A4'].font = SECF
for lbl, val, note, row in [
  ('Apply DA to FRONTLINE margin?', 'N', 'Default N. In the source file the DA is annotated only on best-price cells, so it is treated as funding the discounted price, not everyday frontline business. Set to Y if the allowance is earned on every case/keg regardless of the price sold at.', 6),
  ('Apply DA to BEST-PRICE margin?', 'Y', 'Default Y. Effective cost on best-price business = laid-in cost − DA. This is the assumption that makes Truly 15.5 at $115 profitable (see note below).', 7),
]:
    setc(a, row, 1, lbl, BOLD)
    c = setc(a, row, 2, val, Font(name=F, size=10, bold=True, color='0000FF'), fill=YEL, align='center')
    setc(a, row, 3, note, Font(name=F, size=9), wrap=True)
    a.row_dimensions[row].height = 44

a['A9'] = 'COLOR LEGEND'; a['A9'].font = SECF
leg = [('Blue text on pale blue fill', 'EDITABLE INPUT — type over these (laid-in cost, DA, frontline price, best price)', BLUE, INFILL),
       ('Black text', 'CALCULATED — formula driven, do not type over', BLACK, None),
       ('Green text', 'LINK to another tab in this workbook', GREEN, None),
       ('Yellow fill', 'Key assumption / control cell', BOLD, YEL)]
r = 10
for k, v, fnt, fl in leg:
    setc(a, r, 1, k, fnt, fill=fl); setc(a, r, 2, '', BLACK, fill=fl)
    setc(a, r, 3, v, Font(name=F, size=9), wrap=True)
    r += 1

r += 1
a.cell(row=r, column=1, value='HOW MARGIN IS CALCULATED').font = SECF; r += 1
calcs = [
 ('Frontline margin $', 'Frontline price − frontline cost basis'),
 ('Frontline margin %', 'Frontline margin $ ÷ frontline price   (gross margin on selling price, not markup on cost)'),
 ('Best-price margin $', 'Best price − best-price cost basis'),
 ('Best-price margin %', 'Best-price margin $ ÷ best price'),
 ('Frontline cost basis', 'Laid-in cost — or laid-in − DA, if the FRONTLINE control above is set to Y'),
 ('Best-price cost basis', 'Laid-in − DA when the BEST-PRICE control is Y (default), otherwise plain laid-in cost'),
]
for k, v in calcs:
    setc(a, r, 1, k, BOLD); setc(a, r, 3, v, Font(name=F, size=9), wrap=True); r += 1

r += 1
a.cell(row=r, column=1, value='ASSUMPTIONS & JUDGEMENT CALLS (read before using the numbers)').font = SECF; r += 1
notes = [
 ('DA is a cost reducer, not a price reducer',
  'The source file records DA as free text inside the price cell — Twisted Tea 5.2 best price "58 ($9 DA)" and Truly 15.5 "$115 ($35.40 DA)". Those are read as: the number shown IS the selling price, and the DA lowers our cost. Truly 15.5 proves it — $115 against a $114.11 laid-in is $0.89 of margin (0.8%), which is not a real price. Net of the $35.40 DA the cost is $78.71 and the margin is $36.29 / 31.6%, right in line with the rest of the book. The DA has been split into its own editable column so it is no longer buried in text.'),
 ('DA applies only where the source annotated it',
  'Current book: Twisted Tea 5.2 ($9) and Truly 15.5 ($35.40), on Kohler and on the Boston Beer suggested price. New book: the same two SKUs carry the same allowances. Peerless and Shore Point DA is UNKNOWN — entered as 0, so their best-price margins are shown before any allowance they may receive and are therefore conservative (understated). Enter their DA if you learn it.'),
 ('One laid-in cost per product, applied to every distributor',
  'The source file carries a single "Laid In" column, not one per wholesaler. Competitor margins are therefore modelled on KOHLER\'S laid-in cost. Freight, fuel and any supplier-specific terms will move a competitor\'s real cost, so treat Peerless / Shore Point margin figures as directional, not as their books. The cost cell is editable per row if you want to model a different competitor cost.'),
 ('"Suggested" is a benchmark, not a competitor',
  'The Suggested columns are Boston Beer\'s suggested pricing. It is carried as a fourth row per product so you can see how far each wholesaler sits off the supplier\'s intent, but it is not a distributor in the market.'),
 ('Pack size = keg size',
  '15.5 = half barrel, 5.2 = sixtel. Every price and cost in the source is per keg, so no case-pack conversion is applied anywhere. $/gallon columns are added so a 15.5 and a 5.2 can be compared on the same basis (price ÷ gallons).'),
 ('Truly 15.5 — Peerless does not carry it',
  'The source shows "NA". Those cells are left blank and no margin is computed, rather than being scored as a zero.'),
 ('New-scenario prices are SEEDED, not sourced',
  'The source file supplies new laid-in cost and new SUGGESTED prices only — the Kohler and competitor new-price cells were blank question marks. On the New Pricing tab, Kohler / Peerless / Shore Point frontline and best prices are seeded with their CURRENT prices, i.e. "hold price and eat the cost increase". They are blue editable cells — overwrite them with what you actually intend to charge and every margin follows.'),
 ('Twisted Tea 15.5 new suggested best price',
  'The source shows a new suggested best of $184 against a new suggested frontline of $189 — only $5 of promotional room, versus $16 in the current book. Carried through as written; worth confirming with the supplier.'),
 ('Kohler Actuals tab',
  'Weighted-average realized price = Σ(units × unit price) ÷ units, from the invoice transaction report for 7/1/2026–8/31/2026. Zero-price lines (18 lines / 21 kegs of samples and comps) and backorder-only lines are excluded. The realized prices confirm the Kohler columns in the pricing file: every frontline and best price in the analysis appears in the invoice data as an actual charged price.'),
]
for k, v in notes:
    setc(a, r, 1, k, BOLD, wrap=True); setc(a, r, 3, v, Font(name=F, size=9), wrap=True)
    a.row_dimensions[r].height = 14 * (1 + len(v)//110)
    r += 1

r += 1
a.cell(row=r, column=1, value='TABS').font = SECF; r += 1
for k, v in [('Current Pricing','Today\'s laid-in cost and today\'s prices — every product × every distributor.'),
             ('New Pricing','New laid-in cost, new suggested pricing, and your editable new prices. Shows the margin change vs today.'),
             ('Summary','Kohler vs Peerless vs Shore Point vs Suggested, by product, on one screen.'),
             ('Kohler Actuals','What Kohler actually charged and shipped, 7/1–8/31/2026, per product.'),
             ('Source Data','The original Boston_Pricing_Analysis.xlsx sheet, untouched.')]:
    setc(a, r, 1, k, BOLD); setc(a, r, 3, v, Font(name=F, size=9)); r += 1

CTRL_FL  = "Assumptions!$B$6"
CTRL_BP  = "Assumptions!$B$7"

# ================================================= 2. CURRENT PRICING
cp = wb.create_sheet('Current Pricing')
titleblock(cp, 'Current Pricing & Margin — by Product and Distributor',
   'Laid-in cost and prices per keg. Blue cells are editable; everything else recalculates. Competitor margins use Kohler\'s laid-in cost (see Assumptions).')
CP_HDR = ['Product','Keg Size (Gal)','Distributor','Old Laid-In Cost ($/keg)','DA ($/keg)',
          'Frontline Price ($/keg)','Best Price ($/keg)','Frontline Cost Basis ($)',
          'Frontline Margin $','Frontline Margin %','Best-Price Cost Basis ($)',
          'Best-Price Margin $','Best-Price Margin %','Best vs Frontline ($)',
          'Frontline $/Gal','Best $/Gal','Notes']
CP_W = [21,9,20,12,9,12,12,12,11,11,12,11,11,11,10,10,42]
header_row(cp, 4, CP_HDR, CP_W)
CP_R0 = 5
for pi, (name, gal, old, new, fl, bp, da, snfl, snbp, snda) in enumerate(P):
    band = BAND if pi % 2 else None
    for di, dist in enumerate(DISTS):
        r = CP_R0 + pi*4 + di
        fill = KOHL if di == 0 else band
        setc(cp, r, 1, name, BOLD if di == 0 else BLACK, fill=fill)
        setc(cp, r, 2, gal, BLACK, NUM if gal == int(gal) else '0.0', fill=fill, align='center')
        setc(cp, r, 3, dist, BOLD if di == 0 else BLACK, fill=fill)
        setc(cp, r, 4, old, BLUE, MONEY, INFILL)
        setc(cp, r, 5, (da[di] if da[di] is not None else 0), BLUE, MONEY, INFILL)
        setc(cp, r, 6, fl[di], BLUE, MONEY, INFILL)
        setc(cp, r, 7, bp[di], BLUE, MONEY, INFILL)
        setc(cp, r, 8, f'=IF($D{r}="","",IF(UPPER({CTRL_FL})="Y",$D{r}-$E{r},$D{r}))', BLACK, MONEY, fill)
        setc(cp, r, 9, f'=IF(OR($F{r}="",$H{r}=""),"",$F{r}-$H{r})', BLACK, MONEY, fill)
        setc(cp, r,10, f'=IF(OR($I{r}="",$F{r}=""),"",IFERROR($I{r}/$F{r},""))', BLACK, PCT, fill)
        setc(cp, r,11, f'=IF($D{r}="","",IF(UPPER({CTRL_BP})="Y",$D{r}-$E{r},$D{r}))', BLACK, MONEY, fill)
        setc(cp, r,12, f'=IF(OR($G{r}="",$K{r}=""),"",$G{r}-$K{r})', BLACK, MONEY, fill)
        setc(cp, r,13, f'=IF(OR($L{r}="",$G{r}=""),"",IFERROR($L{r}/$G{r},""))', BLACK, PCT, fill)
        setc(cp, r,14, f'=IF(OR($F{r}="",$G{r}=""),"",$G{r}-$F{r})', BLACK, MONEY, fill)
        setc(cp, r,15, f'=IF($F{r}="","",IFERROR($F{r}/$B{r},""))', BLACK, MONEY, fill)
        setc(cp, r,16, f'=IF($G{r}="","",IFERROR($G{r}/$B{r},""))', BLACK, MONEY, fill)
        note = ''
        if di == 3: note = 'Boston Beer suggested price — benchmark, not a distributor.'
        if da[di]: note = f'Source shows best price with a ${da[di]:,.2f} DA; DA broken out to col E.'
        if da[di] is None and fl[di] is not None: note = 'Competitor DA unknown — entered as 0; best-price margin is conservative.'
        if fl[di] is None: note = 'Source shows "NA" — Peerless does not carry this SKU.'
        setc(cp, r,17, note, Font(name=F, size=8, italic=True, color='7F7F7F'), fill=fill, wrap=True)
        if fl[di] is None:
            for col in (6,7,5):
                cp.cell(row=r, column=col).value = None
    cp.cell(row=CP_R0+pi*4, column=1).border = Border(left=thin, right=thin, bottom=thin, top=Side(style='medium', color='1F3864'))
cp.freeze_panes = 'D5'
cp.auto_filter.ref = f'A4:Q{CP_R0+len(P)*4-1}'
cp.cell(row=CP_R0+len(P)*4+1, column=1,
    value='Source: Boston_Pricing_Analysis.xlsx, Sheet1 rows 1–11 (current book). DA amounts extracted from the free text in the source best-price cells.').font = SUB

# ================================================= 3. NEW PRICING
npg = wb.create_sheet('New Pricing')
titleblock(npg, 'New Pricing & Margin — by Product and Distributor',
   'New laid-in cost from the source file. Kohler / Peerless / Shore Point new prices are SEEDED with current prices (= hold price, absorb the increase) — overwrite the blue cells with your intended pricing.')
NP_HDR = ['Product','Keg Size (Gal)','Distributor','Old Laid-In ($/keg)','New Laid-In ($/keg)',
          'Δ Laid-In ($)','DA ($/keg)','Frontline Price ($/keg)','Best Price ($/keg)',
          'Frontline Cost Basis ($)','Frontline Margin $','Frontline Margin %',
          'Best-Price Cost Basis ($)','Best-Price Margin $','Best-Price Margin %',
          'Δ Frontline Margin % vs Current','Δ Best-Price Margin % vs Current',
          'Best vs Frontline ($)','Notes']
NP_W = [21,9,20,11,11,9,9,12,12,12,11,11,12,11,11,13,13,11,42]
header_row(npg, 4, NP_HDR, NP_W)
NP_R0 = 5
for pi, (name, gal, old, new, fl, bp, da, snfl, snbp, snda) in enumerate(P):
    band = BAND if pi % 2 else None
    for di, dist in enumerate(DISTS):
        r  = NP_R0 + pi*4 + di
        cr = CP_R0 + pi*4 + di          # matching row on Current Pricing
        fill = KOHL if di == 0 else band
        setc(npg, r, 1, name, BOLD if di == 0 else BLACK, fill=fill)
        setc(npg, r, 2, gal, BLACK, NUM if gal == int(gal) else '0.0', fill=fill, align='center')
        setc(npg, r, 3, dist, BOLD if di == 0 else BLACK, fill=fill)
        setc(npg, r, 4, f"='Current Pricing'!$D${cr}", GREEN, MONEY, fill)
        setc(npg, r, 5, new, BLUE, MONEY, INFILL)
        setc(npg, r, 6, f'=IF(OR($D{r}="",$E{r}=""),"",$E{r}-$D{r})', BLACK, MONEY, fill)
        if di == 3:                      # Suggested — new prices come from the source
            setc(npg, r, 7, snda, BLUE, MONEY, INFILL)
            setc(npg, r, 8, snfl, BLUE, MONEY, INFILL)
            setc(npg, r, 9, snbp, BLUE, MONEY, INFILL)
        else:
            setc(npg, r, 7, (da[di] if da[di] is not None else 0), BLUE, MONEY, INFILL)
            setc(npg, r, 8, fl[di], BLUE, MONEY, INFILL)
            setc(npg, r, 9, bp[di], BLUE, MONEY, INFILL)
        setc(npg, r,10, f'=IF($E{r}="","",IF(UPPER({CTRL_FL})="Y",$E{r}-$G{r},$E{r}))', BLACK, MONEY, fill)
        setc(npg, r,11, f'=IF(OR($H{r}="",$J{r}=""),"",$H{r}-$J{r})', BLACK, MONEY, fill)
        setc(npg, r,12, f'=IF(OR($K{r}="",$H{r}=""),"",IFERROR($K{r}/$H{r},""))', BLACK, PCT, fill)
        setc(npg, r,13, f'=IF($E{r}="","",IF(UPPER({CTRL_BP})="Y",$E{r}-$G{r},$E{r}))', BLACK, MONEY, fill)
        setc(npg, r,14, f'=IF(OR($I{r}="",$M{r}=""),"",$I{r}-$M{r})', BLACK, MONEY, fill)
        setc(npg, r,15, f'=IF(OR($N{r}="",$I{r}=""),"",IFERROR($N{r}/$I{r},""))', BLACK, PCT, fill)
        setc(npg, r,16, f'=IF(OR($L{r}="",\'Current Pricing\'!$J${cr}=""),"",$L{r}-\'Current Pricing\'!$J${cr})', BLACK, PCT, fill)
        setc(npg, r,17, f'=IF(OR($O{r}="",\'Current Pricing\'!$M${cr}=""),"",$O{r}-\'Current Pricing\'!$M${cr})', BLACK, PCT, fill)
        setc(npg, r,18, f'=IF(OR($H{r}="",$I{r}=""),"",$I{r}-$H{r})', BLACK, MONEY, fill)
        if di == 3:
            note = 'New suggested frontline / best price from the source file.'
            if snda: note += f' Source annotates a ${snda:,.2f} DA — broken out to col G.'
        elif fl[di] is None:
            note = 'Peerless does not carry this SKU ("NA" in source).'
        else:
            note = 'SEEDED with current price — source gave no new price for this distributor. Edit.'
            if da[di]: note += f' Current ${da[di]:,.2f} DA carried forward.'
        setc(npg, r,19, note, Font(name=F, size=8, italic=True, color='7F7F7F'), fill=fill, wrap=True)
        if fl[di] is None and di != 3:
            for col in (7,8,9):
                npg.cell(row=r, column=col).value = None
    npg.cell(row=NP_R0+pi*4, column=1).border = Border(left=thin, right=thin, bottom=thin, top=Side(style='medium', color='1F3864'))
npg.freeze_panes = 'D5'
npg.auto_filter.ref = f'A4:S{NP_R0+len(P)*4-1}'
npg.cell(row=NP_R0+len(P)*4+1, column=1,
    value='Source: Boston_Pricing_Analysis.xlsx, Sheet1 rows 13–23 (new book). Only new laid-in cost and new SUGGESTED prices exist in the source; distributor new prices are seeded and editable.').font = SUB

# ================================================= 4. SUMMARY
sm = wb.create_sheet('Summary')
titleblock(sm, 'Summary — Kohler vs the Market, by Product',
   'Every figure links live to Current Pricing / New Pricing. Change a price on those tabs and this tab moves with it.')
SW = [21,8,11,10,11,10,11,10,11,10,11,11,11,11]
for i, w in enumerate(SW, 1): sm.column_dimensions[gcl(i)].width = w

def cpref(col, pi, di): return f"'Current Pricing'!${col}${CP_R0+pi*4+di}"
def npref(col, pi, di): return f"'New Pricing'!${col}${NP_R0+pi*4+di}"

def block(start, heading, subhead, price_col, mgn_col, ref):
    sm.cell(row=start, column=1, value=heading).font = SECF
    sm.cell(row=start+1, column=1, value=subhead).font = SUB
    hdr = ['Product','Keg Size','Kohler Price','Kohler Mgn %','Peerless Price','Peerless Mgn %',
           'Shore Point Price','Shore Pt Mgn %','Suggested Price','Suggested Mgn %',
           'Kohler vs Peerless ($)','Kohler vs Shore Pt ($)','Kohler vs Suggested ($)',
           'Kohler Mgn % vs Competitor Avg']
    hr = start+2
    for i, h in enumerate(hdr, 1):
        c = sm.cell(row=hr, column=i, value=h)
        c.font = HDRF; c.fill = HDRFILL; c.border = BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sm.row_dimensions[hr].height = 34
    for pi, (name, gal, *_rest) in enumerate(P):
        r = hr+1+pi
        fill = BAND if pi % 2 else None
        setc(sm, r, 1, name, BOLD, fill=fill)
        setc(sm, r, 2, gal, BLACK, '0.0', fill=fill, align='center')
        for j, di in enumerate([0,1,2,3]):
            pc = 3+j*2; mc = 4+j*2
            setc(sm, r, pc, f'=IF({ref(price_col,pi,di)}="","",{ref(price_col,pi,di)})', GREEN, MONEY,
                 KOHL if di == 0 else fill)
            setc(sm, r, mc, f'=IF({ref(mgn_col,pi,di)}="","",{ref(mgn_col,pi,di)})', GREEN, PCT,
                 KOHL if di == 0 else fill)
        setc(sm, r, 11, f'=IF(OR($C{r}="",$E{r}=""),"",$C{r}-$E{r})', BLACK, MONEY, fill)
        setc(sm, r, 12, f'=IF(OR($C{r}="",$G{r}=""),"",$C{r}-$G{r})', BLACK, MONEY, fill)
        setc(sm, r, 13, f'=IF(OR($C{r}="",$I{r}=""),"",$C{r}-$I{r})', BLACK, MONEY, fill)
        setc(sm, r, 14, f'=IF($D{r}="","",IFERROR($D{r}-AVERAGE($F{r},$H{r}),""))',
             BLACK, PCT, fill)
    return hr+1+len(P)

end1 = block(4,  'A.  FRONTLINE — CURRENT BOOK',
             'Frontline price and frontline margin % at today\'s laid-in cost. Positive "Kohler vs" = Kohler is priced ABOVE that distributor.',
             'F','J', cpref)
end2 = block(end1+2, 'B.  BEST PRICE — CURRENT BOOK',
             'Best (promotional) price and margin %, net of DA where one applies. Peerless / Shore Point DA is unknown and set to 0, so their margins here are understated.',
             'G','M', cpref)
end3 = block(end2+2, 'C.  FRONTLINE — NEW BOOK',
             'Same view at the NEW laid-in cost, using whatever prices are entered on the New Pricing tab (seeded with current prices until you change them).',
             'H','L', npref)
end4 = block(end3+2, 'D.  BEST PRICE — NEW BOOK',
             'Best price and margin % at the new laid-in cost.',
             'I','O', npref)

# Block E — Kohler margin impact
r0 = end4+2
sm.cell(row=r0, column=1, value='E.  KOHLER MARGIN IMPACT — NEW BOOK vs CURRENT BOOK').font = SECF
sm.cell(row=r0+1, column=1, value='What the cost increase does to Kohler at the prices currently entered on the New Pricing tab. Negative = margin lost.').font = SUB
hdr = ['Product','Keg Size','Δ Laid-In ($)','Current FL Price','New FL Price','Current FL Mgn %','New FL Mgn %',
       'Δ FL Mgn %','Current Best Price','New Best Price','Current Best Mgn %','New Best Mgn %','Δ Best Mgn %',
       'FL Price Needed to Hold Current FL Mgn %']
hr = r0+2
for i, h in enumerate(hdr, 1):
    c = sm.cell(row=hr, column=i, value=h)
    c.font = HDRF; c.fill = HDRFILL; c.border = BOX
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
sm.row_dimensions[hr].height = 34
for pi, (name, gal, *_r) in enumerate(P):
    r = hr+1+pi
    fill = BAND if pi % 2 else None
    setc(sm, r, 1, name, BOLD, fill=fill)
    setc(sm, r, 2, gal, BLACK, '0.0', fill=fill, align='center')
    setc(sm, r, 3, f'={npref("F",pi,0)}', GREEN, MONEY, fill)
    setc(sm, r, 4, f'={cpref("F",pi,0)}', GREEN, MONEY, fill)
    setc(sm, r, 5, f'={npref("H",pi,0)}', GREEN, MONEY, fill)
    setc(sm, r, 6, f'={cpref("J",pi,0)}', GREEN, PCT, fill)
    setc(sm, r, 7, f'={npref("L",pi,0)}', GREEN, PCT, fill)
    setc(sm, r, 8, f'=IF(OR($F{r}="",$G{r}=""),"",$G{r}-$F{r})', BLACK, PCT, fill)
    setc(sm, r, 9, f'={cpref("G",pi,0)}', GREEN, MONEY, fill)
    setc(sm, r,10, f'={npref("I",pi,0)}', GREEN, MONEY, fill)
    setc(sm, r,11, f'={cpref("M",pi,0)}', GREEN, PCT, fill)
    setc(sm, r,12, f'={npref("O",pi,0)}', GREEN, PCT, fill)
    setc(sm, r,13, f'=IF(OR($K{r}="",$L{r}=""),"",$L{r}-$K{r})', BLACK, PCT, fill)
    setc(sm, r,14, f'=IF(OR($F{r}="",{npref("J",pi,0)}=""),"",IFERROR({npref("J",pi,0)}/(1-$F{r}),""))',
         BLACK, MONEY, fill)
sm.freeze_panes = 'C5'

# ================================================= 5. KOHLER ACTUALS
ka = wb.create_sheet('Kohler Actuals')
titleblock(ka, 'Kohler Actuals — What We Really Charged, 7/1/2026 – 8/31/2026',
   'From InvoiceTransReport_GSHARKEY_20260903. Reality check on the pricing file: the realized prices match the Kohler columns in the analysis exactly.')
KA_HDR = ['Product','Keg Size (Gal)','Kegs Shipped','Accounts Bought','Weighted Avg Realized $/keg',
          'Frontline Price','Best Price','Realized vs Frontline ($)','% of Kegs at Frontline',
          '% of Kegs at Best Price or Lower','DA ($/keg)','Weighted DA Credit ($/keg)',
          'Effective Cost — Current Laid-In','Realized Margin $ (current cost)','Realized Margin % (current cost)',
          'Effective Cost — New Laid-In','Realized Margin $ (new cost, price held)',
          'Realized Margin % (new cost, price held)','Margin $ at Risk on This Volume']
header_row(ka, 4, KA_HDR, [21,9,10,10,13,11,11,11,11,12,9,11,12,12,12,12,12,12,13])
KA_R0 = 5
for pi, (name, gal, *_r) in enumerate(P):
    r = KA_R0+pi
    u, w, ac, pfl, pbp = ACT[name]
    fill = BAND if pi % 2 else None
    setc(ka, r, 1, name, BOLD, fill=fill)
    setc(ka, r, 2, gal, BLACK, '0.0', fill=fill, align='center')
    setc(ka, r, 3, u, BLACK, NUM, fill)
    setc(ka, r, 4, ac, BLACK, NUM, fill)
    setc(ka, r, 5, w, BLACK, MONEY, fill)
    setc(ka, r, 6, f'={cpref("F",pi,0)}', GREEN, MONEY, fill)
    setc(ka, r, 7, f'={cpref("G",pi,0)}', GREEN, MONEY, fill)
    setc(ka, r, 8, f'=IF($F{r}="","",$E{r}-$F{r})', BLACK, MONEY, fill)
    setc(ka, r, 9, pfl, BLACK, PCT, fill)
    setc(ka, r,10, pbp, BLACK, PCT, fill)
    setc(ka, r,11, f'={cpref("E",pi,0)}', GREEN, MONEY, fill)
    setc(ka, r,12, f'=IF($K{r}="","",$K{r}*$J{r})', BLACK, MONEY, fill)
    setc(ka, r,13, f'=IF({cpref("D",pi,0)}="","",{cpref("D",pi,0)}-$L{r})', BLACK, MONEY, fill)
    setc(ka, r,14, f'=IF($M{r}="","",$E{r}-$M{r})', BLACK, MONEY, fill)
    setc(ka, r,15, f'=IF($N{r}="","",IFERROR($N{r}/$E{r},""))', BLACK, PCT, fill)
    setc(ka, r,16, f'=IF({npref("E",pi,0)}="","",{npref("E",pi,0)}-$L{r})', BLACK, MONEY, fill)
    setc(ka, r,17, f'=IF($P{r}="","",$E{r}-$P{r})', BLACK, MONEY, fill)
    setc(ka, r,18, f'=IF($Q{r}="","",IFERROR($Q{r}/$E{r},""))', BLACK, PCT, fill)
    setc(ka, r,19, f'=IF(OR($N{r}="",$Q{r}=""),"",($Q{r}-$N{r})*$C{r})', BLACK, MONEY0, fill)
tr = KA_R0+len(P)
setc(ka, tr, 1, 'TOTAL', BOLD)
setc(ka, tr, 3, f'=SUM(C{KA_R0}:C{tr-1})', BOLD, NUM)
for c in [2,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]: setc(ka, tr, c, '', BOLD)
setc(ka, tr,19, f'=SUM(S{KA_R0}:S{tr-1})', Font(name=F, size=10, bold=True, color='C00000'), MONEY0)
for c in range(1, 20):
    ka.cell(row=tr, column=c).border = Border(top=Side(style='medium', color='1F3864'), bottom=thin, left=thin, right=thin)
ka.cell(row=tr+2, column=1, value='Weighted average realized price excludes zero-price lines (18 lines / 21 kegs of samples and comps) and backorder-only lines.').font = SUB
ka.cell(row=tr+3, column=1, value='Weighted DA credit = DA x % of kegs sold at best price or lower — i.e. the allowance is credited only against the share of volume that actually moved on the promoted price. Without it, Truly 15.5 reads as a 5% margin product, which is an artefact of ignoring a $35.40 allowance on 90% of its volume.').font = SUB
ka.cell(row=tr+4, column=1, value='Margin $ at Risk = (new-cost margin - current-cost margin) x kegs shipped in this two-month window. It is what holding price through the cost increase would have cost on July-August volume, not a forecast.').font = SUB
ka.cell(row=tr+5, column=1, value='Sam Adams 15.5 shows only 34% at frontline and 19% at best price because a large share ships at $179 and $181 — intermediate prices between the two, which the pricing analysis does not capture.').font = SUB
ka.freeze_panes = 'C5'

# ================================================= 6. SOURCE DATA
src = openpyxl.load_workbook(SRC, data_only=True)['Sheet1']
sd = wb.create_sheet('Source Data')
titleblock(sd, 'Source Data — Boston_Pricing_Analysis.xlsx (unmodified)',
   'Verbatim copy of the original Sheet1, including the DA notes recorded as free text inside price cells. Nothing on this tab feeds the calculations.')
for i, row in enumerate(src.iter_rows(min_row=1, max_row=src.max_row, max_col=src.max_column, values_only=True)):
    for j, v in enumerate(row):
        if v is None: continue
        c = sd.cell(row=4+i, column=1+j, value=v)
        c.font = BOLD if (i in (0, 12) or j == 0) else BLACK
        c.border = BOX
        if isinstance(v, (int, float)): c.number_format = MONEY
        if i in (0, 12):
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
sd.column_dimensions['A'].width = 22
for j in range(2, 11): sd.column_dimensions[gcl(j)].width = 15
sd.row_dimensions[4].height = 40
sd.row_dimensions[16].height = 40

wb.save(OUT)
print('saved', OUT)
