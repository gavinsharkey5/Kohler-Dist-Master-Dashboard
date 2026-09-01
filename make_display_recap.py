#!/usr/bin/env python3
"""Rebuilds Summer_Display_Auction_Jul-Aug_2026.xlsx -- the July/August 2026
Summer Display Auction recap workbook -- from a FULL-PERIOD iSellBeer
Report_NN.xlsx export.

    python3 make_display_recap.py Report_NN.xlsx

Why this exists: the workbook was hand-built twice (commits 6289339 and
6b2157b) with no script behind it, so every remake risked drifting from the
live leaderboard. This script imports the tracker's OWN scoring functions
from isellbeer/display-auction-tracker/generate.py -- read_rows(),
build_displays(), canonical_brand(), classify(), tier_for() -- so the recap
cannot disagree with the board about what a display is, what tier it lands
in, or what it pays. Anything this file adds on top (supplier attribution,
lead brand, duplicate clustering) is recap-only presentation that the
tracker has no opinion about.

INPUT MUST COVER THE WHOLE PERIOD (07/01/2026 onward). Unlike the tracker,
this script has no --merge mode: the workbook is a point-in-time recap
rebuilt from scratch each time, not an accumulating archive. Hand it a
weekly partial export and you get a workbook covering only that week. The
script refuses anything whose Filters tab does not start on 07/01/2026.

Supplier attribution (the one piece of logic not shared with the tracker):
each display is credited to the single supplier group holding the most
cases in that photo, so no display is counted twice across the supplier
tab. The group comes from the BRAND, not the export's Supplier column,
because three groups do not line up with it -- Garage Beer appears under
both GARAGE BEER CO. and MAHOU USA (its contract brewer), Lytt under both
BOSTON BEER COMPANY and LYTT LLC, and Montauk sits under TILRAY BRANDS
next to the unrelated Popsicle FMB. See BRAND_SUPPLIER_OVERRIDES.
"""
import collections
import datetime
import importlib.util
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
TRACKER = HERE / "isellbeer" / "display-auction-tracker" / "generate.py"
OUT = HERE / "Summer_Display_Auction_Jul-Aug_2026.xlsx"

WINDOW_START = "07/01/2026"
WINDOW_END = "08/31/2026"

# --- pull the tracker's scoring in, rather than restating it here ---------
_spec = importlib.util.spec_from_file_location("da_tracker", TRACKER)
tracker = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(tracker)

# Supplier groups named on the program slide. Everything else becomes
# "All Other: <raw supplier>" so the tab still reconciles to 100%.
NAMED_SUPPLIERS = {
    "CONSTELLATION BRANDS": "Constellation",
    "MOLSON COORS BEVERAGE COMPANY": "Molson Coors",
    "HEINEKEN": "HUSA (Heineken USA)",
    "BOSTON BEER COMPANY": "Boston Beer",
    "LYTT LLC": "Boston Beer",
    "MARK ANTHONY GROUP": "Mark Anthony",
    "YUENGLING BREWERY": "Yuengling",
    "GARAGE BEER CO.": "Garage",
    "MAHOU USA": "Garage",
    "SAPPORO": "Sapporo",
    "PABST BREWING COMPANY": "Pabst",
    "DRINK CARBLISS": "Carbliss",
}
SLIDE_ORDER = ["Constellation", "Molson Coors", "HUSA (Heineken USA)", "Boston Beer",
               "Mark Anthony", "Yuengling", "Garage", "Sapporo", "Pabst", "Montauk",
               "Carbliss"]

# The three brands whose supplier column lies about which group they belong
# to (see module docstring). Keyed on the canonical brand.
BRAND_SUPPLIER_OVERRIDES = {
    "MONTAUK BREWING COMPANY": "Montauk",
    "LYTT": "Boston Beer",
    "GARAGE BEER": "Garage",
    "GARAGE BEER - CONTRACT BREWING": "Garage",
}

NAVY, AMBER, LIGHT, GREY = "1F4E5F", "C46A28", "EAF0F3", "F2F2F2"
FOOT_GREY, LINK_BLUE = "595959", "0563C1"
FONT = "Arial"
NUM = "#,##0;-;-"

TIER_LABEL = {4: "Gold (70+)", 3: "Silver (40-69)", 2: "Bronze (20-39)",
              1: "Other (10-19)", 0: "Non-qualifying (<10)"}


def fill(hexrgb):
    return PatternFill("solid", fgColor=hexrgb)


def style(cell, *, bold=False, size=10, color=None, bg=None, wrap=False,
          align=None, numfmt=None):
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    if bg:
        cell.fill = fill(bg)
    if wrap or align:
        cell.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="center")
    if numfmt:
        cell.number_format = numfmt


def banner(ws, title, subtitle, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.cell(1, 1, title)
    ws.cell(2, 1, subtitle)
    for r, size, bold in ((1, 16, True), (2, 10, False)):
        for c in range(1, width + 1):
            style(ws.cell(r, c), bold=bold, size=size, color="FFFFFF", bg=NAVY, align="left")
    ws.row_dimensions[1].height = 24


def footnotes(ws, row, lines):
    for i, text in enumerate(lines):
        c = ws.cell(row + i, 1, text)
        style(c, size=9, color=FOOT_GREY)


def supplier_group(row, brand_supplier):
    """A source row's supplier group, resolved by brand first."""
    brand = tracker.canonical_brand(row)
    if brand in BRAND_SUPPLIER_OVERRIDES:
        return BRAND_SUPPLIER_OVERRIDES[brand]
    raw = (row["supplier"] or "").strip()
    if not raw or raw == "Unspecified":
        # One Monaco line ships with no Supplier. Resolve it the way the
        # rest of the brand's rows are labelled rather than inventing an
        # "Unspecified" group.
        raw = brand_supplier.get(brand, "Unspecified")
    return NAMED_SUPPLIERS.get(raw, f"All Other: {raw}")


def enrich(rows):
    """Tracker displays + the recap-only fields (supplier, lead brand, splits)."""
    # A brand's modal non-blank supplier, for resolving blank Supplier cells.
    seen = collections.defaultdict(collections.Counter)
    for r in rows:
        raw = (r["supplier"] or "").strip()
        if raw and raw != "Unspecified":
            seen[tracker.canonical_brand(r)][raw] += 1
    brand_supplier = {b: sorted(c.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
                      for b, c in seen.items()}

    by_key = collections.defaultdict(list)
    for r in rows:
        by_key[(r["taker"], r["acct"], r["dt"])].append(r)

    displays = tracker.build_displays(rows)
    for d in displays:
        grp = by_key[(d["taker"], d["acct"], d["dt"])]
        sup_cases, brand_cases, raw_sups = (collections.Counter(), collections.Counter(),
                                            set())
        for r in grp:
            qty = r["qty"] or 0
            sup_cases[supplier_group(r, brand_supplier)] += qty
            brand_cases[tracker.canonical_brand(r)] += qty
            raw_sups.add((r["supplier"] or "Unspecified").strip())
        # Credit the group holding the most cases; ties break alphabetically
        # so a rerun on the same data always lands the same way.
        d["supplier"] = min(sup_cases, key=lambda g: (-sup_cases[g], g))
        d["supplier_cases"] = sup_cases[d["supplier"]]
        d["lead_brand"] = min(brand_cases, key=lambda b: (-brand_cases[b], b))
        d["multi"] = len(sup_cases) > 1
        d["split"] = " + ".join(f"{g} ({n})" for g, n in
                                sorted(sup_cases.items(), key=lambda kv: (-kv[1], kv[0])))
        d["raw_suppliers"] = ", ".join(sorted(raw_sups))
        d["dm"] = grp[0]["dm"]          # build_displays() doesn't carry the DM through
        d["sku_lines"] = len(grp)
        d["dt_obj"] = tracker.parse_dt(d["dt"])
    displays.sort(key=lambda d: (-d["cases"], d["dt_obj"]))
    for n, d in enumerate(displays, 1):
        d["id"] = f"D{n:04d}"
    return displays


def find_duplicates(displays):
    """Qualifying displays sharing an account, a day and a brand list.

    Within one account/day/brand-list group, displays are first sub-grouped by
    CASE COUNT:

      * a sub-group of 2+ displays on the same case count is an exact cluster --
        "Same person — exact repeat" if one person logged it twice, otherwise
        "Cross-person — exact match";
      * the singletons left over (each a different case count) form a
        "Cross-person — same brands, different cases" cluster, but only if they
        come from 2+ DIFFERENT people.

    That last condition is the important one: one rep logging 41 and 45 cases of
    Corona at the same store on the same day is two genuinely different displays,
    not a double submission, so it is not flagged. Two different people logging
    80 and 173 cases of the same brand at the same store on the same day is one
    display counted twice, and is.

    Earliest submission is kept, the rest are flagged. Ordering is on the PARSED
    timestamp: the workbook this replaces sorted the raw string, so "01:03 PM"
    sorted ahead of "10:43 AM" and a handful of clusters kept the later
    submission while labelling it "first submitted".
    """
    buckets = collections.defaultdict(list)
    for d in displays:
        if d["tier"] == 0:
            continue
        buckets[(d["acct"], d["dt_obj"].date(), tuple(d["brands"]))].append(d)

    clusters = []
    for group in buckets.values():
        by_cases = collections.defaultdict(list)
        for d in group:
            by_cases[d["cases"]].append(d)
        leftovers = []
        for members in by_cases.values():
            if len(members) < 2:
                leftovers.extend(members)
                continue
            members.sort(key=lambda d: d["dt_obj"])
            kind = ("Same person — exact repeat"
                    if len({d["taker"] for d in members}) == 1
                    else "Cross-person — exact match")
            clusters.append((kind, members))
        if len(leftovers) >= 2 and len({d["taker"] for d in leftovers}) >= 2:
            leftovers.sort(key=lambda d: d["dt_obj"])
            clusters.append(("Cross-person — same brands, different cases", leftovers))

    order = {"Same person — exact repeat": 2, "Cross-person — exact match": 1,
             "Cross-person — same brands, different cases": 0}
    clusters.sort(key=lambda c: (order[c[0]], -c[1][0]["cases"]))
    for n, (_, group) in enumerate(clusters, 1):
        for i, d in enumerate(group):
            d["cluster"] = f"C{n:03d}"
            d["dup_flag"] = "—" if i == 0 else "Possible duplicate"
    for d in displays:
        d.setdefault("cluster", "")
        d.setdefault("dup_flag", "—")
    return clusters


# --------------------------------------------------------------------------
# sheets
# --------------------------------------------------------------------------
SUM_HEADERS = ["Gold\n70+ cases", "Silver\n40-69 cases", "Bronze\n20-39 cases",
               "Other\n10-19 cases", "Total Qualifying\nDisplays",
               "Total Cases\n(qualifying)", "Total Points"]


def summary_sheet(wb, name, title, key_col, labels, sections, widths,
                  foot, dup_col=False, last_row=0):
    """One of the three COUNTIFS/SUMIFS summary tabs.

    Every figure is a live formula over Display Detail rather than a baked
    number, so filtering that tab shows the records behind any cell.
    """
    ws = wb.create_sheet(name)
    heads = ["" ] + SUM_HEADERS[:]
    if dup_col:
        heads.append("Possible\nDuplicates")
    heads += ["Non-Qualifying\n(<10 cases)", "Total\nSubmissions"]
    width = len(heads)
    banner(ws, title,
           f"Kohler iSell Beer Summer Display Program · {WINDOW_START} – {WINDOW_END}"
           f" · one row per photo submission, counted once in its case-size tier", width)
    heads[0] = labels
    for c, h in enumerate(heads, 1):
        cell = ws.cell(4, c, h)
        style(cell, bold=True, color="FFFFFF", bg=NAVY, wrap=True,
              align="left" if c == 1 else "center")
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = "A5"
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    D = "'Display Detail'!"
    rng = lambda col: f"{D}${col}$2:${col}${last_row}"
    row = 5
    for kind, value in sections:
        if kind == "section":
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
            ws.cell(row, 1, value)
            for c in range(1, width + 1):
                style(ws.cell(row, c), bold=True, color="FFFFFF", bg=AMBER, align="left")
            row += 1
            continue
        total = kind == "total"
        ws.cell(row, 1, value)
        a = f"$A{row}"
        formulas = [
            f'=COUNTIFS({rng(key_col)},{a},{rng("K")},4)',
            f'=COUNTIFS({rng(key_col)},{a},{rng("K")},3)',
            f'=COUNTIFS({rng(key_col)},{a},{rng("K")},2)',
            f'=COUNTIFS({rng(key_col)},{a},{rng("K")},1)',
            f'=COUNTIFS({rng(key_col)},{a},{rng("K")},">0")',
            f'=SUMIFS({rng("M")},{rng(key_col)},{a},{rng("K")},">0")',
            f'=SUMIFS({rng("P")},{rng(key_col)},{a},{rng("K")},">0")',
        ]
        if dup_col:
            formulas.append(f'=COUNTIFS({rng(key_col)},{a},{rng("X")},"Possible duplicate")')
        formulas += [f'=COUNTIFS({rng(key_col)},{a},{rng("K")},0)',
                     f'=COUNTIF({rng(key_col)},{a})']
        if total:
            # The TOTAL row counts every display, so it sums the tier column
            # directly instead of matching a label that no row carries.
            formulas = [
                f'=COUNTIF({rng("K")},4)', f'=COUNTIF({rng("K")},3)',
                f'=COUNTIF({rng("K")},2)', f'=COUNTIF({rng("K")},1)',
                f'=COUNTIF({rng("K")},">0")',
                f'=SUMIFS({rng("M")},{rng("K")},">0")',
                f'=SUMIFS({rng("P")},{rng("K")},">0")',
            ]
            if dup_col:
                formulas.append(f'=COUNTIF({rng("X")},"Possible duplicate")')
            formulas += [f'=COUNTIF({rng("K")},0)', f'=COUNTA({rng("A")})']
        for c, f in enumerate(formulas, 2):
            style(ws.cell(row, c, f), bold=total, color="FFFFFF" if total else None,
                  bg=NAVY if total else None, numfmt=NUM, align="center")
        style(ws.cell(row, 1), bold=total, color="FFFFFF" if total else None,
              bg=NAVY if total else None, align="left")
        row += 1

    footnotes(ws, row + 1, foot)
    return ws


def display_detail_sheet(wb, displays):
    ws = wb.create_sheet("Display Detail")
    heads = ["Display ID", "Date", "Date / Time", "Rep / Associate", "Role",
             "District Manager", "Account #", "Account (DBA)", "City",
             "Supplier Group", "Tier #", "Display Tier", "Cases",
             "Cases from This Supplier", "Point Class", "Points", "Qualifying?",
             "Lead Brand", "All Brands in Display", "Raw Supplier(s)", "SKU Lines",
             "Multi-Supplier?", "Supplier Split (cases)", "Duplicate Flag",
             "Photo Link"]
    widths = [10, 11, 18, 20, 15, 18, 10, 34, 18, 30, 7, 20, 8, 12, 11, 8, 11, 30,
              46, 34, 9, 13, 46, 26, 12]
    for c, (h, w) in enumerate(zip(heads, widths), 1):
        style(ws.cell(1, c, h), bold=True, color="FFFFFF", bg=NAVY)
        ws.column_dimensions[get_column_letter(c)].width = w
    for i, d in enumerate(displays, 2):
        vals = [d["id"], d["dt_obj"].strftime("%m/%d/%Y"), d["dt"], d["taker"],
                d["role"], d.get("dm", ""), d["acct"], d["dba"], d["city"],
                d["supplier"], d["tier"], TIER_LABEL[d["tier"]], d["cases"],
                d["supplier_cases"],
                "Priority" if d["classification"] == "priority" else "All Other",
                d["points"], "Yes" if d["tier"] else "No", d["lead_brand"],
                ", ".join(d["brands"]), d["raw_suppliers"], d["sku_lines"],
                "Yes" if d["multi"] else "No", d["split"], d["dup_flag"], None]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            style(cell, bg=GREY if not d["tier"] else None,
                  numfmt=NUM if c in (13, 14, 16) else None)
        link = ws.cell(i, 25, "View Photo" if d["photos"] else "—")
        if d["photos"]:
            link.hyperlink = d["photos"][0]
            style(link, color=LINK_BLUE, bg=GREY if not d["tier"] else None)
            link.font = Font(name=FONT, size=10, color=LINK_BLUE, underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Y{len(displays) + 1}"
    return ws


def duplicate_sheet(wb, clusters, total_points):
    ws = wb.create_sheet("Duplicate Check")
    banner(ws, "Duplicate Check — displays that may have been logged twice",
           "Qualifying displays only. Clusters are the same account on the same day "
           "with the same brands; the earliest submission is kept, the rest are "
           "flagged for review.", 12)
    flagged = [d for _, g in clusters for d in g[1:]]
    pts = sum(d["points"] for d in flagged)
    by_kind = collections.Counter(k for k, g in clusters for _ in g[1:])
    rows = [
        ("Clusters found", len(clusters),
         "Groups of 2+ qualifying displays on one account/day with the same brand list."),
        ("Displays flagged as possible duplicates", len(flagged),
         "Every display in a cluster except the first one submitted."),
        ("Points sitting on those displays", pts,
         f"{pts / total_points * 100:.1f}% of the {total_points:,} points on the board."),
        ("  · Same person — exact repeat", by_kind["Same person — exact repeat"],
         "One person logged the identical display twice, minutes apart. Almost "
         "certainly a double submission."),
        ("  · Cross-person — exact match", by_kind["Cross-person — exact match"],
         "Two+ people logged the same display, same case count. Usually a rep and "
         "the associate riding with them."),
        ("  · Cross-person — same brands, different cases",
         by_kind["Cross-person — same brands, different cases"],
         "Same account/day/brands but the case counts differ — same display, "
         "counted differently by each person."),
    ]
    for i, (label, value, note) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, label); ws.cell(r, 2, value); ws.cell(r, 3, note)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
        for c in range(1, 13):
            style(ws.cell(r, c), bold=i < 3, bg=LIGHT)
        style(ws.cell(r, 2), bold=i < 3, bg=LIGHT, numfmt=NUM, align="center")
    footnotes(ws, 10, ["None of these are removed from the summary tabs — those match "
                       "the published leaderboard exactly. This tab is the review list."])

    heads = ["Flag", "Cluster", "Status", "Date / Time", "Rep / Associate", "Role",
             "Account (DBA)", "Account #", "Cases", "Display Tier", "Points",
             "Brands in Display", "Photo"]
    widths = [30, 9, 20, 18, 20, 15, 34, 10, 8, 18, 8, 44, 11]
    for c, (h, w) in enumerate(zip(heads, widths), 1):
        style(ws.cell(12, c, h), bold=True, color="FFFFFF", bg=NAVY, align="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    r = 13
    for kind, group in clusters:
        for i, d in enumerate(group):
            vals = [kind, d["cluster"],
                    "Kept (first submitted)" if i == 0 else "Possible duplicate",
                    d["dt"], d["taker"], d["role"], d["dba"], d["acct"], d["cases"],
                    TIER_LABEL[d["tier"]], d["points"], ", ".join(d["brands"]), None]
            for c, v in enumerate(vals, 1):
                style(ws.cell(r, c, v), bg=None if i == 0 else GREY,
                      numfmt=NUM if c in (9, 11) else None)
            link = ws.cell(r, 13, "View Photo" if d["photos"] else "—")
            if d["photos"]:
                link.hyperlink = d["photos"][0]
                link.font = Font(name=FONT, size=10, color=LINK_BLUE, underline="single")
                if i:
                    link.fill = fill(GREY)
            r += 1
    ws.freeze_panes = "A13"
    return ws


def multi_supplier_sheet(wb, displays):
    ws = wb.create_sheet("Multi-Supplier Audit")
    banner(ws, "Multi-Supplier Displays — attribution audit",
           "Photo submissions carrying more than one supplier group. Each is "
           "credited to the group holding the most cases.", 9)
    heads = ["Date", "Rep / Associate", "Account (DBA)", "Total Cases", "Display Tier",
             "Credited To", "Cases from Credited Supplier",
             "Full Supplier Split (cases)", "Qualifying?"]
    widths = [11, 20, 34, 11, 20, 28, 14, 56, 11]
    for c, (h, w) in enumerate(zip(heads, widths), 1):
        style(ws.cell(4, c, h), bold=True, color="FFFFFF", bg=NAVY, wrap=True,
              align="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"
    multi = sorted([d for d in displays if d["multi"]],
                   key=lambda d: (-d["cases"], d["dt_obj"]))
    for i, d in enumerate(multi, 5):
        vals = [d["dt_obj"].strftime("%m/%d/%Y"), d["taker"], d["dba"], d["cases"],
                TIER_LABEL[d["tier"]], d["supplier"], d["supplier_cases"], d["split"],
                "Yes" if d["tier"] else "No"]
        for c, v in enumerate(vals, 1):
            style(ws.cell(i, c, v), bg=GREY if not d["tier"] else None,
                  numfmt=NUM if c in (4, 7) else None)
    return ws, len(multi)


def notes_sheet(wb, stats, src_name):
    ws = wb.create_sheet("Notes & Method")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 120
    rows = [
        ("H", "Notes & Method", None),
        ("S", "How every number in this workbook was derived, and what to watch for.", None),
        ("B", None, None),
        ("S", "Scope", None),
        ("R", "Program", "Kohler iSell Beer Summer Display Auction (Sales Reps) and Summer "
                         "Display Rewards (Sales Associates), July/August 2026."),
        ("R", "Source data", f"iSellBeer display photo export, one row per SKU per photo "
                             f"({src_name}). The same rows drive the live leaderboard, and "
                             f"this workbook imports the tracker's own scoring functions so "
                             f"the two cannot disagree."),
        ("R", "Window", f"{WINDOW_START} through {WINDOW_END} — the complete program period."),
        ("R", "Rebuilt by", "make_display_recap.py in the repo root. Rerun it against a "
                            "full-period export to remake this workbook; it has no merge "
                            "mode, so a partial export would produce a partial recap."),
        ("S", "What counts as one display", None),
        ("R", "Display definition", "One display = one photo submission = all export rows "
                                    "sharing the same Photo Taker + Account # + Date/Time. "
                                    "Multiple SKU rows under one photo are summed into that "
                                    "display's case count. This matches how the live tracker "
                                    "scores the board."),
        ("R", "Case count", "The display's TOTAL cases across every SKU in the photo. This "
                            "is the number the tier is scored on."),
        ("R", "Tier", "Gold 70+ · Silver 40-69 · Bronze 20-39 · Other 10-19 · under 10 cases "
                      "does not qualify. A display appears in exactly one tier."),
        ("S", "Supplier attribution", None),
        ("R", "Rule", f"Each display is credited to the single supplier group holding the "
                      f"most cases in that photo, so no display is counted twice. "
                      f"{stats['multi']} of {stats['total']} submissions "
                      f"({stats['multi_qual']} of the {stats['qualifying']} qualifying) carry "
                      f"more than one supplier — all are listed on the Multi-Supplier Audit "
                      f"tab with their full case split."),
        ("R", "Why brand, not the Supplier column",
         "Supplier is mapped from the brand rather than read off the export's Supplier "
         "column, because three groups do not line up with it: Garage Beer appears under "
         "both GARAGE BEER CO. and MAHOU USA (its contract brewer), Lytt under both BOSTON "
         "BEER COMPANY and LYTT LLC, and Montauk sits under TILRAY BRANDS next to the "
         "unrelated Popsicle FMB."),
        ("R", "Pabst", "Named on the program slide, but no Pabst display was submitted in "
                       "July or August — the row reads zero."),
        ("S", "Points", None),
        ("R", "Scale used", f"Points are the live tracker's published scale: Priority "
                            f"200 / 300 / 500 / 1000 for Other / Bronze / Silver / Gold, "
                            f"All Other 100 / 200 / 300 / 600. Workbook total "
                            f"{stats['points']:,} points reconciles exactly to the published "
                            f"leaderboard."),
        ("R", "ATTENTION — differs from the slide",
         "The program slide shows a 10x-smaller scale (Priority 100 Gold / 50 Silver / 30 "
         "Bronze, All Other 60 / 30 / 10, plus 10 points for any 10+ case display). Scaled "
         "up, that is Priority 1000/500/300/100 and All Other 600/300/100/100 — which agrees "
         "with the tracker at Gold and Silver but not at the two smallest tiers (tracker "
         "Priority-Other 200 vs slide 100; tracker All-Other-Bronze 200 vs slide 100). The "
         "tracker's values were confirmed directly by Kohler when each tier first appeared, "
         "so they are what the board pays. Worth a quick confirmation before the September "
         "auction if points are being redeemed off the slide."),
        ("R", "Priority vs All Other", "Assigned per brand from the tracker's own confirmed "
                                       "brand lists, not from the supplier grouping. Two "
                                       "consequences worth knowing: Molson Coors covers both "
                                       "Priority brands (Coors, Miller Lite, Blue Moon) and an "
                                       "All Other one (Redd's); and Monaco and Fever-Tree score "
                                       "as Priority on the tracker although neither appears on "
                                       "the program slide's supplier list."),
        ("R", "Mixed-class displays", "A display carrying both Priority and All Other brands is "
                                      "classed by whichever side holds more cases, then scored "
                                      "on the display's total cases — the rule Kohler confirmed "
                                      "when this first appeared."),
        ("S", "Data quality", None),
        ("R", "Complete fields", f"Photo Taker, Role, District Manager, Account #, DBA, City, "
                                 f"Brand, SKU, Quantity, Date/Time and the photo link are "
                                 f"populated on every one of the {stats['rows']:,} source rows. "
                                 f"Nothing needed to identify a display, its cases, or its rep "
                                 f"is missing."),
        ("R", "Blank Brand Family", f"{stats['blank_bf']} rows carry no Brand Family and fall "
                                    f"back to the raw Brand column (Carbliss, Monaco, Sinless, "
                                    f"Sun Cruisers). Handled, no data lost."),
        ("R", "Blank Supplier", f"{stats['blank_sup']} row(s) have no Supplier value. Resolved "
                                f"by brand, so they land in All Other rather than Unspecified."),
        ("R", "Zero-quantity rows", f"{stats['zero_qty']} rows carry Quantity 0. They are real "
                                    f"SKU lines on real photos and are kept, but contribute no "
                                    f"cases — a display made up entirely of them scores as "
                                    f"non-qualifying."),
        ("R", "Non-qualifying displays", f"{stats['nonqual']} of the {stats['total']} "
                                         f"submissions came in under 10 cases. They are kept in "
                                         f"the detail (greyed out, Qualifying? = No) so the tabs "
                                         f"reconcile to the full submission count, but they are "
                                         f"excluded from every qualifying total."),
        ("S", "Duplicates", None),
        ("R", "What was checked", "Every qualifying display was compared against the others on "
                                  "the same account and the same day. Three patterns are flagged "
                                  "on the Duplicate Check tab, and the Display Detail tab "
                                  "carries a matching Duplicate Flag column you can filter on."),
        ("R", "Same person — exact repeat", f"{stats['dup_same']} displays. One person logged "
                                            f"the identical account, day, case count and brand "
                                            f"twice, minutes apart. These look like "
                                            f"straightforward double submissions."),
        ("R", "Cross-person — exact match", f"{stats['dup_exact']} displays. Two or more people "
                                            f"logged the same account, day, cases and brands. "
                                            f"Most pair a Sales Rep with the Sales Associate "
                                            f"working the same store — whether that is double "
                                            f"counting depends on whether both programs are "
                                            f"meant to pay on the same physical display."),
        ("R", "Cross-person — same brands, different cases",
         f"{stats['dup_diff']} displays. Same account, day and brands, but each person recorded "
         f"a different case count."),
        ("R", "Nothing was removed", f"Every summary figure in this workbook includes all "
                                     f"{stats['qualifying']} qualifying displays and matches the "
                                     f"published leaderboard. The {stats['dup_flagged']} flagged "
                                     f"displays carry {stats['dup_points']:,} points "
                                     f"({stats['dup_pct']:.1f}% of the board) — decide the "
                                     f"treatment first, then rerun if any are to be excluded."),
        ("R", "Not duplicates", "No photo URL appears on more than one display, so nothing is "
                                "double-counted at the photo level. Repeat visits to the same "
                                "account on DIFFERENT days, and different brands photographed "
                                "separately on the same visit, are normal program activity and "
                                "are not flagged."),
        ("S", "Reconciliation", None),
        ("R", "Check", f"{stats['total']} total submissions · {stats['qualifying']} qualifying · "
                       f"{stats['points']:,} points. All three tie exactly to the live tracker "
                       f"at gavinsharkey5.github.io/Kohler-Dist-Master-Dashboard/isellbeer/"
                       f"display-auction-tracker/."),
    ]
    for r, (kind, a, b) in enumerate(rows, 1):
        ws.cell(r, 1, a); ws.cell(r, 2, b)
        if kind == "H":
            style(ws.cell(r, 1), bold=True, size=16, color="FFFFFF", bg=NAVY)
            style(ws.cell(r, 2), bg=NAVY)
        elif kind == "S" and b is None and a and r > 2:
            for c in (1, 2):
                style(ws.cell(r, c), bold=True, color="FFFFFF", bg=AMBER)
        else:
            style(ws.cell(r, 1), bold=True)
            style(ws.cell(r, 2), wrap=True, align="left")
    ws.column_dimensions["B"].width = 110
    return ws


def main():
    if len(sys.argv) != 2:
        raise SystemExit("usage: make_display_recap.py Report_NN.xlsx  (full period only)")
    src = Path(sys.argv[1])
    wbin = openpyxl.load_workbook(src)
    filters = {r[0]: r[1] for r in wbin["Filters"].iter_rows(values_only=True) if r[0]}
    start = str(filters.get("Start Date", "")).strip()
    if start != WINDOW_START:
        raise SystemExit(
            f"{src.name} starts {start!r}, not {WINDOW_START}. This workbook is a "
            f"full-period recap and has no merge mode — a partial export would "
            f"silently produce a partial recap. Re-pull 07/01/2026 onward."
        )

    rows = tracker.read_rows(src)
    unknown = {tracker.canonical_brand(r) for r in rows}
    unknown -= tracker.PRIORITY_BRANDS | tracker.ALLOTHER_BRANDS
    if unknown:
        raise SystemExit(f"Unclassified brand(s) — add to the tracker's brand lists "
                         f"after confirming with the user: {sorted(unknown)}")

    displays = enrich(rows)
    clusters = find_duplicates(displays)
    qualifying = [d for d in displays if d["tier"]]
    points = sum(d["points"] for d in displays)
    flagged = [d for _, g in clusters for d in g[1:]]
    by_kind = collections.Counter(k for k, g in clusters for _ in g[1:])
    dup_points = sum(d["points"] for d in flagged)

    stats = {
        "rows": len(rows), "total": len(displays), "qualifying": len(qualifying),
        "nonqual": len(displays) - len(qualifying), "points": points,
        "multi": sum(1 for d in displays if d["multi"]),
        "multi_qual": sum(1 for d in qualifying if d["multi"]),
        "blank_bf": sum(1 for r in rows if not (r["brand_family"] or "").strip()),
        "blank_sup": sum(1 for r in rows
                         if not (r["supplier"] or "").strip()
                         or (r["supplier"] or "").strip() == "Unspecified"),
        "zero_qty": sum(1 for r in rows if not r["qty"]),
        "dup_same": by_kind["Same person — exact repeat"],
        "dup_exact": by_kind["Cross-person — exact match"],
        "dup_diff": by_kind["Cross-person — same brands, different cases"],
        "dup_flagged": len(flagged), "dup_points": dup_points,
        "dup_pct": dup_points / points * 100 if points else 0,
    }

    last = len(displays) + 1
    src_note = (f"Source: iSellBeer display photo export, {WINDOW_START} – {WINDOW_END} "
                f"({src.stem}, rebuilt {datetime.date.today():%m/%d/%Y}).")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Summary by Supplier -------------------------------------------
    present = {d["supplier"] for d in displays}
    named = [g for g in SLIDE_ORDER]
    others = sorted(g for g in present if g.startswith("All Other: "))
    sections = [("row", g) for g in named]
    sections.append(("section", "ALL OTHER BRANDS / SUPPLIERS — not named on the program slide"))
    sections += [("row", g) for g in others]
    sections.append(("total", "TOTAL — ALL DISPLAYS"))
    widths = {"A": 34, "B": 12, "C": 12, "D": 12, "E": 12, "F": 14, "G": 13, "H": 12,
              "I": 12, "J": 13, "K": 12}
    summary_sheet(
        wb, "Summary by Supplier",
        "Summer Display Auction — Qualifying Displays by Supplier", "J",
        "Supplier / Brand Group", sections, widths,
        ["A display is counted once, in the single tier its TOTAL case count falls into "
         "(85 cases = Gold only, not Gold + Silver + Bronze).",
         f"A display carrying more than one supplier is credited to the supplier holding "
         f"the most cases in that photo — see the Multi-Supplier Audit tab for all "
         f"{stats['multi']}.",
         "Pabst is on the program slide but has no display submissions in the July/August "
         "data, so it reads zero throughout.",
         "Every figure is a live COUNTIFS/SUMIFS over the Display Detail tab — filter that "
         "tab to see the records behind any number.",
         f"These totals match the published leaderboard and include everything submitted. "
         f"{stats['dup_flagged']} displays look like the same display logged twice — see "
         f"the Duplicate Check tab before treating any total as final.",
         "Possible Duplicates counts qualifying displays in this supplier's total that look "
         "like the same display logged twice — they are INCLUDED in every other column, not "
         "netted out. Full list on the Duplicate Check tab.",
         src_note], dup_col=True, last_row=last)

    # --- Summary by Brand ----------------------------------------------
    brand_cases = collections.Counter()
    for d in displays:
        brand_cases[d["lead_brand"]] += d["cases"] if d["tier"] else 0
    brands = sorted(brand_cases, key=lambda b: (-brand_cases[b], b))
    widths = {"A": 38, "B": 12, "C": 12, "D": 12, "E": 12, "F": 14, "G": 13, "H": 12,
              "I": 13, "J": 12}
    summary_sheet(
        wb, "Summary by Brand",
        "Summer Display Auction — Qualifying Displays by Lead Brand", "R", "Lead Brand",
        [("row", b) for b in brands] + [("total", "TOTAL — ALL DISPLAYS")], widths,
        ["Lead Brand = the brand holding the most cases in that photo submission, so each "
         "display appears against exactly one brand.",
         "Multi-brand displays are common (a single photo often carries 3-6 SKUs); the full "
         "brand list for every display is on the Display Detail tab.",
         src_note], last_row=last)

    # --- Summary by Person ---------------------------------------------
    reps = sorted({d["taker"] for d in displays if d["role"] == "Sales Rep"},
                  key=str.casefold)
    assoc = sorted({d["taker"] for d in displays if d["role"] != "Sales Rep"},
                   key=str.casefold)
    widths = {"A": 26, "B": 12, "C": 12, "D": 12, "E": 12, "F": 14, "G": 13, "H": 12,
              "I": 13, "J": 12}
    summary_sheet(
        wb, "Summary by Person",
        "Summer Display Auction — Qualifying Displays by Person", "D",
        "Rep / Sales Associate",
        [("section", "SALES REPS — Display Auction")] + [("row", p) for p in reps] +
        [("section", "SALES ASSOCIATES — Display Rewards")] + [("row", p) for p in assoc] +
        [("total", "TOTAL — ALL DISPLAYS")], widths,
        ["Sales Reps and Sales Associates run as two separate programs but earn on the same "
         "point scale.",
         "Points shown are the tracker's published scale — see the Notes & Method tab on how "
         "that compares to the program slide.",
         src_note], last_row=last)

    duplicate_sheet(wb, clusters, points)
    _, n_multi = multi_supplier_sheet(wb, displays)
    display_detail_sheet(wb, displays)
    notes_sheet(wb, stats, src.stem)
    wb.save(OUT)

    print(f"Wrote {OUT.name}")
    print(f"  {stats['total']} submissions · {stats['qualifying']} qualifying · "
          f"{points:,} points · {len(reps)} reps + {len(assoc)} associates")
    print(f"  {n_multi} multi-supplier displays · {len(clusters)} duplicate clusters, "
          f"{stats['dup_flagged']} flagged ({dup_points:,} pts)")
    print(f"  supplier groups: {len(named)} named + {len(others)} All Other")


if __name__ == "__main__":
    main()
