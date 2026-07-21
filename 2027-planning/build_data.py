#!/usr/bin/env python3
"""
Builds data.json for the 2027 Mod Year Review / Planning dashboard from:

  2026_planning_source.xlsx   The 2026 Planning by Brand workbook (RDE export
                               format). Gives us the canonical brand-family ->
                               supplier / segment / sub-segment / brand manager
                               taxonomy, plus the 2026 Brewery & Kohler goal %
                               and CE for every brand (columns K-P of the
                               '2026 Planning by Brand' tab).
  fy2024_fy2025_full_year.csv RDE "Comparison" export: CE / $Vol / Gross for
                               1/1/2024-12/31/2024 and 1/1/2025-12/31/2025.
  fy2025_ytd_comparable.csv   Same export, 1/1/2025-7/21/2025 -- the period
                               matched to fy2026_ytd.csv below.
  fy2026_ytd.csv               Same export, 1/1/2026-7/21/2026 (this year's
                               actuals so far).

RDE "Comparison" exports use a nested hierarchy with no indentation to mark
it: each supplier group starts with a row giving the SUPPLIER's total, then
one row per brand family under that supplier. There is no column flagging
which rows are which -- we instead build a canonical brand_name -> supplier
lookup from the workbook's grey/formula rows (which the workbook itself uses
for exactly this purpose, see is_grey below) and match every CSV row against
it by brand name. Rows that don't match any known brand family (e.g. brand
new SKUs added after the 1/8/2026 planning doc was built) are kept in an
"Unclassified" bucket per supplier so dollars still reconcile to the file's
own Total row, but are flagged as not yet in the taxonomy.

Run: python3 build_data.py
"""
import csv
import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
WORKBOOK = HERE / "2026_planning_source.xlsx"
CSV_FY24_25 = HERE / "fy2024_fy2025_full_year.csv"
CSV_FY25_YTD = HERE / "fy2025_ytd_comparable.csv"
CSV_FY26_YTD = HERE / "fy2026_ytd.csv"
OUT = HERE / "data" / "data.json"


def to_num(raw):
    if raw is None:
        return 0.0
    raw = str(raw).strip()
    if raw == "" or raw == "-":
        return 0.0
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()").replace("$", "").replace(",", "").replace("%", "")
    if raw == "":
        return 0.0
    val = float(raw)
    return -val if neg else val


# ---------------------------------------------------------------------------
# 1. Canonical taxonomy + 2026 goals, from the workbook
# ---------------------------------------------------------------------------

# A handful of brand families are labeled differently between the workbook
# (hand-maintained) and the RDE "Comparison" export (system-generated) for
# the same real product. Confirmed by volume/segment match, not just name
# similarity: "Pabst Brand" in the workbook is Pabst Brewing Company's
# flagship SKU, which RDE exports as "Pabst Blue Ribbon".
NAME_ALIASES = {
    "pabst blue ribbon": "Pabst Brand",
}


def load_workbook_taxonomy():
    wb_values = openpyxl.load_workbook(WORKBOOK, data_only=True)
    wb_formulas = openpyxl.load_workbook(WORKBOOK, data_only=False)
    wsv = wb_values["2026 Planning by Brand"]
    wsf = wb_formulas["2026 Planning by Brand"]

    brands = {}          # brand_name (as in workbook) -> record
    brands_lower = {}    # lowercased brand_name -> workbook brand_name
    supplier_names = set()
    brand_manager_by_supplier = {}

    for r in range(3, wsv.max_row + 1):
        brand = wsv.cell(r, 1).value
        if brand is None or str(brand).strip() == "":
            continue
        is_grey = wsf.cell(r, 1).fill.patternType == "solid"
        supplier = wsv.cell(r, 2).value

        if is_grey:
            # Supplier subtotal row -- 'brand' here is actually the supplier name.
            supplier_names.add(str(brand).strip())
            continue

        segment = wsv.cell(r, 4).value
        subsegment = wsv.cell(r, 5).value
        manager = wsv.cell(r, 3).value
        est_2025_finish = wsv.cell(r, 8).value
        brewery_pct = wsv.cell(r, 11).value
        brewery_ce_change = wsv.cell(r, 12).value
        brewery_goal_ce = wsv.cell(r, 13).value
        kohler_pct = wsv.cell(r, 14).value
        kohler_ce_change = wsv.cell(r, 15).value
        kohler_goal_ce = wsv.cell(r, 16).value

        name = str(brand).strip()
        brands[name] = {
            "brand": name,
            "supplier": (str(supplier).strip() if supplier else None),
            "segment": segment,
            "subsegment": subsegment,
            "brand_manager": manager,
            "finish_2025_ce": est_2025_finish if isinstance(est_2025_finish, (int, float)) else None,
            "goal2026_brewery_pct": brewery_pct if isinstance(brewery_pct, (int, float)) else None,
            "goal2026_brewery_ce_change": brewery_ce_change if isinstance(brewery_ce_change, (int, float)) else None,
            "goal2026_brewery_ce": brewery_goal_ce if isinstance(brewery_goal_ce, (int, float)) else None,
            "goal2026_kohler_pct": kohler_pct if isinstance(kohler_pct, (int, float)) else None,
            "goal2026_kohler_ce_change": kohler_ce_change if isinstance(kohler_ce_change, (int, float)) else None,
            "goal2026_kohler_ce": kohler_goal_ce if isinstance(kohler_goal_ce, (int, float)) else None,
        }
        if supplier:
            # Some suppliers only ever appear as the Supplier column on their
            # brands' leaf rows, with no grey subtotal row of their own in
            # this workbook (e.g. Food & Bev Enterprise LLC) -- register them
            # too so their header row in the CSV is recognized.
            supplier_names.add(str(supplier).strip())
        if supplier and manager:
            brand_manager_by_supplier.setdefault(str(supplier).strip(), str(manager).strip())
        brands_lower[name.lower()] = name

    return brands, brands_lower, supplier_names, brand_manager_by_supplier


# ---------------------------------------------------------------------------
# 2. Parse the hierarchical RDE "Comparison" CSVs against that taxonomy
# ---------------------------------------------------------------------------

def parse_comparison_csv(path, metric_cols, brands, brands_lower, supplier_names):
    """metric_cols: dict of output_key -> column header substring to match."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        col_map = {}
        for out_key, needle in metric_cols.items():
            match = next((c for c in fieldnames if c.startswith(needle)), None)
            if not match:
                raise SystemExit(f"{path}: couldn't find a column starting with {needle!r} in {fieldnames}")
            col_map[out_key] = match

        rows = []
        for r in reader:
            name = (r.get("Supplier / Brand Family") or "").strip()
            if not name or name == "Total":
                continue
            rows.append((name, {k: to_num(r[c]) for k, c in col_map.items()}))

    supplier_names_lower = {s.lower() for s in supplier_names}

    results = {}       # (supplier, brand) -> metrics dict
    current_supplier = None
    # The name that set current_supplier on the immediately preceding row, if
    # any -- lets us tell apart a supplier's subtotal header row from a brand
    # family that happens to share the supplier's exact name (e.g. Sapporo /
    # Sapporo, Heineken USA / Heineken USA): the header always comes first,
    # and if a leaf with the identical name immediately follows, that second
    # occurrence is real brand data, not another header.
    prev_header_name = None
    unclassified = []
    supplier_conflicts = []

    for raw_name, metrics in rows:
        name_l = raw_name.lower()

        if prev_header_name is not None and name_l == prev_header_name.lower():
            canonical = brands_lower.get(name_l, raw_name)
            results[(current_supplier, canonical)] = metrics
            prev_header_name = None
            continue

        if name_l in supplier_names_lower:
            current_supplier = next(s for s in supplier_names if s.lower() == name_l)
            prev_header_name = current_supplier
            continue

        prev_header_name = None
        alias = NAME_ALIASES.get(name_l)
        canonical = alias or brands_lower.get(name_l)
        if canonical:
            # The workbook is hand-maintained and can drift from who RDE
            # currently attributes a brand to; trust the workbook's supplier
            # (it's what the 2026 goals were set against) but fall back to
            # RDE's live grouping when the workbook doesn't have one.
            workbook_supplier = brands[canonical]["supplier"]
            supplier = workbook_supplier or current_supplier
            if workbook_supplier and current_supplier and workbook_supplier != current_supplier:
                supplier_conflicts.append((canonical, workbook_supplier, current_supplier))
            results[(supplier, canonical)] = metrics
            continue

        # Not a known brand or known supplier header -> new/unclassified SKU.
        unclassified.append((current_supplier, raw_name, metrics))

    return results, unclassified, supplier_conflicts


METRIC_SETS = {
    "fy2024_fy2025": {
        "ce_2024": "Case Equiv 1/1/2024",
        "ce_2025": "Case Equiv 1/1/2025",
        "vol_2024": "$Vol 1/1/2024",
        "vol_2025": "$Vol 1/1/2025",
        "gross_2024": "Gross 1/1/2024",
        "gross_2025": "Gross 1/1/2025",
    },
    "fy2025_ytd": {
        "ce_2025_ytd": "Case Equiv 1/1/2025",
        "vol_2025_ytd": "$Vol 1/1/2025",
        "gross_2025_ytd": "Gross 1/1/2025",
    },
    "fy2026_ytd": {
        "ce_2026_ytd": "Case Equiv 1/1/2026",
        "vol_2026_ytd": "$Vol 1/1/2026",
        "gross_2026_ytd": "Gross 1/1/2026",
    },
}


def main():
    brands, brands_lower, supplier_names, brand_manager_by_supplier = load_workbook_taxonomy()
    print(f"Loaded {len(brands)} canonical brand families under {len(supplier_names)} suppliers from workbook.")

    fy2425, unclass_2425, conflicts_2425 = parse_comparison_csv(
        CSV_FY24_25, METRIC_SETS["fy2024_fy2025"], brands, brands_lower, supplier_names)
    fy25ytd, unclass_25ytd, conflicts_25ytd = parse_comparison_csv(
        CSV_FY25_YTD, METRIC_SETS["fy2025_ytd"], brands, brands_lower, supplier_names)
    fy26ytd, unclass_26ytd, conflicts_26ytd = parse_comparison_csv(
        CSV_FY26_YTD, METRIC_SETS["fy2026_ytd"], brands, brands_lower, supplier_names)

    conflicts = {c[0]: c for c in conflicts_2425 + conflicts_25ytd + conflicts_26ytd}
    if conflicts:
        print(f"\n{len(conflicts)} brand(s) where the workbook's supplier disagrees with RDE's current grouping "
              f"(RDE's grouping was used):")
        for brand, (_, wb_sup, rde_sup) in conflicts.items():
            print(f"  {brand}: workbook says {wb_sup!r}, RDE groups it under {rde_sup!r}")

    # --- Reconciliation: matched brand-level sums should equal each file's own Total row ---
    def csv_total(path, needle):
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            col = next(c for c in reader.fieldnames if c.startswith(needle))
            for r in reader:
                if (r.get("Supplier / Brand Family") or "").strip() == "Total":
                    return to_num(r[col])
        return None

    def check(label, matched, unclassified, total, key):
        matched_sum = sum(v[key] for v in matched.values())
        unclass_sum = sum(v[key] for _, _, v in unclassified)
        print(f"[{label}] matched={matched_sum:,.2f} + unclassified={unclass_sum:,.2f} "
              f"= {matched_sum + unclass_sum:,.2f} vs file total={total:,.2f} "
              f"(diff={matched_sum + unclass_sum - total:,.2f}); {len(unclassified)} unclassified rows")

    check("2025 full year CE", fy2425, unclass_2425, csv_total(CSV_FY24_25, "Case Equiv 1/1/2025"), "ce_2025")
    check("2025 YTD CE", fy25ytd, unclass_25ytd, csv_total(CSV_FY25_YTD, "Case Equiv 1/1/2025"), "ce_2025_ytd")
    check("2026 YTD CE", fy26ytd, unclass_26ytd, csv_total(CSV_FY26_YTD, "Case Equiv 1/1/2026"), "ce_2026_ytd")

    if unclass_2425:
        print("Unclassified (2024/2025 full year):", [n for _, n, _ in unclass_2425][:20])
    if unclass_25ytd:
        print("Unclassified (2025 YTD):", [n for _, n, _ in unclass_25ytd][:20])
    if unclass_26ytd:
        print("Unclassified (2026 YTD):", [n for _, n, _ in unclass_26ytd][:20])

    # Fold unclassified rows in as their own records (supplier from RDE context,
    # no workbook taxonomy/goals) so their dollars aren't dropped from the
    # 2026 projection or 2027 base -- just flagged as not yet in the workbook.
    for supplier, name, metrics in unclass_2425:
        fy2425[(supplier, name)] = metrics
    for supplier, name, metrics in unclass_25ytd:
        fy25ytd[(supplier, name)] = metrics
    for supplier, name, metrics in unclass_26ytd:
        fy26ytd[(supplier, name)] = metrics

    # --- Merge everything by (supplier, brand) key ---
    all_keys = set(fy2425) | set(fy25ytd) | set(fy26ytd)
    records = []
    for supplier, brand in sorted(all_keys, key=lambda k: (k[0] or "", k[1])):
        base = brands.get(brand, {})
        m24 = fy2425.get((supplier, brand), {})
        m25y = fy25ytd.get((supplier, brand), {})
        m26y = fy26ytd.get((supplier, brand), {})

        rec = {
            "supplier": supplier,
            "brand": brand,
            "segment": base.get("segment"),
            "subsegment": base.get("subsegment"),
            "brand_manager": base.get("brand_manager") or brand_manager_by_supplier.get(supplier),
            "in_taxonomy": brand in brands,
            "ce_2024": m24.get("ce_2024", 0.0),
            "ce_2025": m24.get("ce_2025", 0.0),
            "vol_2024": m24.get("vol_2024", 0.0),
            "vol_2025": m24.get("vol_2025", 0.0),
            "gross_2024": m24.get("gross_2024", 0.0),
            "gross_2025": m24.get("gross_2025", 0.0),
            "ce_2025_ytd": m25y.get("ce_2025_ytd", 0.0),
            "vol_2025_ytd": m25y.get("vol_2025_ytd", 0.0),
            "gross_2025_ytd": m25y.get("gross_2025_ytd", 0.0),
            "ce_2026_ytd": m26y.get("ce_2026_ytd", 0.0),
            "vol_2026_ytd": m26y.get("vol_2026_ytd", 0.0),
            "gross_2026_ytd": m26y.get("gross_2026_ytd", 0.0),
            "goal2026_brewery_pct": base.get("goal2026_brewery_pct"),
            "goal2026_brewery_ce": base.get("goal2026_brewery_ce"),
            "goal2026_kohler_pct": base.get("goal2026_kohler_pct"),
            "goal2026_kohler_ce": base.get("goal2026_kohler_ce"),
        }

        # 2025 remainder (7/22-12/31) actuals, derived (not exported directly).
        rec["ce_2025_remainder"] = rec["ce_2025"] - rec["ce_2025_ytd"]
        rec["vol_2025_remainder"] = rec["vol_2025"] - rec["vol_2025_ytd"]
        rec["gross_2025_remainder"] = rec["gross_2025"] - rec["gross_2025_ytd"]

        # 2026 YTD YoY growth rate per metric (guard divide-by-zero).
        def growth(cur, prior):
            return (cur / prior - 1.0) if prior else None

        g_ce = growth(rec["ce_2026_ytd"], rec["ce_2025_ytd"])
        g_vol = growth(rec["vol_2026_ytd"], rec["vol_2025_ytd"])
        g_gross = growth(rec["gross_2026_ytd"], rec["gross_2025_ytd"])
        rec["growth_2026_ce"] = g_ce
        rec["growth_2026_vol"] = g_vol
        rec["growth_2026_gross"] = g_gross

        # Aug-Dec 2026 (7/22-12/31) projection = 2025 remainder x (1 + YTD growth rate).
        # If there's no 2025 YTD base to compute a growth rate from, fall back to
        # flat (0% growth) on the 2025 remainder rather than dropping the brand.
        rec["ce_2026_proj_remainder"] = rec["ce_2025_remainder"] * (1 + (g_ce or 0))
        rec["vol_2026_proj_remainder"] = rec["vol_2025_remainder"] * (1 + (g_vol or 0))
        rec["gross_2026_proj_remainder"] = rec["gross_2025_remainder"] * (1 + (g_gross or 0))

        rec["ce_2026_proj_finish"] = rec["ce_2026_ytd"] + rec["ce_2026_proj_remainder"]
        rec["vol_2026_proj_finish"] = rec["vol_2026_ytd"] + rec["vol_2026_proj_remainder"]
        rec["gross_2026_proj_finish"] = rec["gross_2026_ytd"] + rec["gross_2026_proj_remainder"]

        # Trailing price/margin per case, blended across 2024-2026 YTD actuals,
        # used as the default (editable) basis for 2027 $ and GP targets.
        total_ce = rec["ce_2024"] + rec["ce_2025"] + rec["ce_2026_ytd"]
        total_vol = rec["vol_2024"] + rec["vol_2025"] + rec["vol_2026_ytd"]
        total_gross = rec["gross_2024"] + rec["gross_2025"] + rec["gross_2026_ytd"]
        rec["price_per_ce"] = (total_vol / total_ce) if total_ce else None
        rec["margin_per_ce"] = (total_gross / total_ce) if total_ce else None

        # 2027 planning defaults: carry the 2026 Brewery/Kohler goal % forward
        # as the starting editable input (user can change in the dashboard).
        rec["goal2027_brewery_pct"] = rec.get("goal2026_brewery_pct")
        rec["goal2027_kohler_pct"] = rec.get("goal2026_kohler_pct")

        records.append(rec)

    OUT.parent.mkdir(exist_ok=True)
    payload = {
        "generated_note": "Built from 2026_planning_source.xlsx + 3 RDE Comparison CSV exports. "
                           "See build_data.py for methodology.",
        "records": records,
    }
    OUT.write_text(json.dumps(payload, indent=2))
    print(f"\nWrote {len(records)} brand records to {OUT}")


if __name__ == "__main__":
    main()
