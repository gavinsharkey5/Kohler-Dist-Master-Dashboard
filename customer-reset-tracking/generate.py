#!/usr/bin/env python3
"""
Rebuilds the embedded data in index.html for the Customer Reset Tracking
dashboard: does resetting an off-premise account's shelf/cooler space (the
"SFS" program) actually lift its sales? Evaluates TWO cohorts side by side
-- the 2025 resets and the 2026 resets -- each against its own prior-year
baseline.

Methodology (per Kohler, 2026-08-11 -- monthly, not daily, since that's the
grain the RDE/Fusion exports below actually come in; kept deliberately
simple):
  - Sales inputs are monthly account totals (Customer Num x Year Month x
    Cases), not a dated transaction ledger, so windows are whole calendar
    months, not a rolling N-day window from the exact reset date.
  - The dashboard's account table has two date-view modes:
      - YTD (default): January 1 through the latest fully-elapsed month,
        this year vs. last year -- computed directly from each cohort's
        own monthly sales file (sales_2026.csv / sales_2025.csv). See
        "2026 YTD source, and the fix" below for a prior version of this
        script that briefly overrode the 2026 cohort's YTD with a
        separate Fusion export; that override has been removed.
      - 3-Month Window: two different reset-anchored comparisons, shown
        side by side:
          - Year-over-year (pre3/post3/lift3): the reset's own month plus
            the following two calendar months (e.g. a March reset ->
            March+April+May), vs. that exact same 3-month window one year
            earlier. Controls for normal seasonal variation.
          - Same-year before -> after (preAdj/liftAdj, sharing post3 as
            its "after" side): the 3 calendar months immediately before
            the reset's month, vs. the reset month + the following two
            months, both within the same year. Shows the immediate change
            around the reset, but does NOT control for seasonality -- in
            this data it reads much higher than the YoY number in both
            cohorts, most likely normal seasonal lift (resets cluster
            earlier in the year; volume generally climbs into summer),
            not the reset's own effect. Computed and still in the data
            (cases3Adj / preAdj / liftAdj) for reference, but no longer
            displayed anywhere on the page as of 2026-08-11 (dashboard
            was doing too much) -- see README.
    Every account/brand row carries a human label for whichever window(s)
    it represents (e.g. "Mar-May 2026") so the exact months being compared
    are always visible, not just implied by the reset date.
  - Lift % = (post total - pre total) / pre total, on Cases. An account
    with zero prior-year cases in a window has no baseline to divide by
    -- flagged as "no prior baseline" rather than an invented or infinite
    percentage.
  - A "Misc" Brand Family bucket in both source exports (opaque, not tied
    to any real brand or supplier -- see load_monthly_sales() below) is
    removed from the sales data before anything else runs, so it never
    enters any total or any Brand Family breakdown on this page.
  - Each account row can be expanded to show its own Brand Family
    breakdown, for whichever window is currently selected -- always
    computed from that cohort's own monthly sales file (sales_2026.csv /
    sales_2025.csv, the only source with brand-level detail), so it
    always sums exactly to that account's own top-line total.

2026 YTD source, and the fix (2026-08-10 -> 2026-08-11): Kohler initially
supplied fusion_ytd_2026.csv, a Fusion "Case Equivalent" export giving
Jan 1-Jul 31 totals for both 2025 and 2026 directly, at the account level,
which briefly overrode this script's own computed YTD numbers for the
2026 cohort only (per Kohler's direction at the time). Its numbers did
NOT match what this script computes by summing sales_2026.csv's monthly
Cases over the same Jan-Jul window (differences ran up to double digits
at some accounts) -- and it turned out that mismatch was the bug: Kohler
had pulled fusion_ytd_2026.csv with the wrong date range on their end.
Kohler confirmed this by sending two corrected RDE exports (one per
cohort, Jan-Jul both years each) -- cross-checked here and their numbers
match sales_2026.csv / sales_2025.csv's own existing Jan-Jul subset
EXACTLY (0 of 73 2026 accounts differ, 0 of 68 2025 accounts differ; not
new data, just confirmation). So fusion_ytd_2026.csv is now known-wrong
and no longer read -- the 2026 cohort's YTD is back to being computed
directly from sales_2026.csv, exactly like the 2025 cohort's always was.
This also means the account-level Brand Family breakdown (always computed
from the same monthly sales file) now sums exactly to the top-line total
for BOTH cohorts, removing the mismatch caveat that used to be on the
2026 cohort's account table. fusion_ytd_2026.csv has been deleted from
this folder; the two corrected RDE exports Kohler sent to confirm the fix
were not saved (their Jan-Jul rows are already fully present in
sales_2026.csv / sales_2025.csv, so keeping them would just be a
duplicate of already-tracked data).

Inputs (keep these filenames when refreshing -- see README for the full
refresh steps):
  reset_accounts_2026.xlsx  2026 reset roster -- Kohler Account #, TD Linx
                             #, Account Name, City, Segmentation (A/B/C),
                             Reset Date. Join key: Kohler Account # (==
                             sales_2026.csv's Customer Num).
  sales_2026.csv             RDE export, one row per (Customer, Brand
                              Family, Year Month) since Jan 2025, Cases only
                              -- replaces the old sales_2026-MM_batch.csv
                              files (those were a daily ledger covering only
                              Jan/Feb/Mar; this is a monthly aggregate
                              covering the full roster and all five cohort
                              months at once). Drives BOTH the account
                              table's YTD default view and its 3-Month
                              Window view for the 2026 cohort.
  reset_accounts_2025.xlsx  2025 reset roster -- TD Linx #, Account Name,
                             City, Segmentation, Reset Date, and a
                             "Customer ID" column that's ALREADY the sales
                             join key (Kohler manually matched TD Linx -> a
                             Customer Num for this file; see its own "Match
                             Basis" column) -- use that directly, there's no
                             "Kohler Account #" column in this file.
  sales_2025.csv             Same shape as sales_2026.csv (Customer Num x
                              Brand Family x Year Month x Cases), one year
                              back (Jan 2024 - Dec 2025), from a different
                              source system ("Fusion" vs. RDE) -- column
                              names are otherwise identical.

  (reset_history_2024.xlsx / reset_history_2025.xlsx are no longer read --
  they only fed the first-time-vs-repeat split, which has been removed.
  Kept on disk for provenance.)

Run: python3 generate.py
"""
import calendar
import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
HTML = HERE / "index.html"

WINDOW_MONTHS = 3


def to_num(raw):
    if raw is None:
        return None
    raw = str(raw).strip()
    if raw == "":
        return None
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()").replace("$", "").replace(",", "").replace("%", "")
    if raw == "":
        return None
    val = float(raw)
    return -val if neg else val


def load_roster_2026():
    wb = load_workbook(HERE / "reset_accounts_2026.xlsx", data_only=True)
    ws = wb["Sheet1"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    accounts = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[idx["Kohler Account #"]]:
            continue
        accounts.append({
            "account": str(r[idx["Kohler Account #"]]),
            "tdLinx": r[idx["TD Linx #"]],
            "name": (r[idx["Account Name"]] or "").strip(),
            "city": (r[idx["City"]] or "").strip().title(),
            "segment": (r[idx["Segmentation"]] or "").strip() or None,
            "resetDate": r[idx["RESET DATE"]].date(),
        })
    return accounts


def load_roster_2025():
    """Header is on row 2 here (row 1 is blank) -- see the file itself."""
    wb = load_workbook(HERE / "reset_accounts_2025.xlsx", data_only=True)
    ws = wb["Sheet1"]
    header = [c.value for c in ws[2]]
    idx = {h: i for i, h in enumerate(header)}
    accounts = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r[idx["Customer ID"]] or not r[idx["RESET DATE"]]:
            continue
        accounts.append({
            "account": str(r[idx["Customer ID"]]),
            "tdLinx": r[idx["TD Linx #"]],
            "name": (r[idx["Account Name"]] or "").strip(),
            "city": (r[idx["City"]] or "").strip().title(),
            "segment": (r[idx["Segmentation"]] or "").strip() or None,
            "resetDate": r[idx["RESET DATE"]].date(),
        })
    return accounts


def parse_year_month(s):
    y, m = s.strip().split("/")
    return int(y), int(m)


def load_monthly_sales(path):
    """Returns (totals, by_brand, brands_by_account, misc_cases).
    totals: {(account, year, month): total cases across every brand}.
    by_brand: {(account, brand, year, month): cases for that brand}.
    brands_by_account: {account: set of brand names ever seen for it}.

    Each row's two "Cases <year>" columns are mutually exclusive (whichever
    one matches the row's own Year Month is populated, same pattern as the
    prior daily export) -- summed rather than picked, defensively, in case
    a future export ever populates both.

    Rows with Brand Family "Misc" (Supplier is always "Misc" too, on these
    rows specifically -- checked, never a real supplier paired with a
    "Misc" brand family or vice versa) are dropped: they're an opaque,
    unattributed catch-all -- not tied to any specific brand's shelf/cooler
    placement, which is what a reset actually affects -- and in this data
    they're wildly lumpy per account/month (single rows worth tens of
    thousands of cases in one month, then zero for months at a stretch),
    which was single-handedly producing account-level lift swings well into
    quadruple digits. They're a large share of raw volume (roughly
    30% of total cases in both exports) -- excluded rather than silently
    kept, so the number is visible: see the printed summary below and the
    dashboard's own caveats."""
    totals = defaultdict(float)
    by_brand = defaultdict(float)
    brands_by_account = defaultdict(set)
    misc_cases = 0.0
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        cases_cols = [c for c in reader.fieldnames
                      if c.strip().startswith("Cases") and "Unit" not in c and "Percentage" not in c]
        for r in reader:
            ym = r.get("Year Month")
            account = (r.get("Customer Num") or "").strip()
            if not ym or not account:
                continue
            year, month = parse_year_month(ym)
            cases = 0.0
            found = False
            for c in cases_cols:
                v = to_num(r.get(c))
                if v is not None:
                    cases += v
                    found = True
            if not found:
                continue
            if (r.get("Brand Family") or "").strip().lower() == "misc":
                misc_cases += cases
                continue
            brand = (r.get("Brand Family") or "").strip() or "(Unlabeled)"
            totals[(account, year, month)] += cases
            by_brand[(account, brand, year, month)] += cases
            brands_by_account[account].add(brand)
    return totals, by_brand, brands_by_account, misc_cases


def format_date_range(d0, d1):
    if d0.year == d1.year:
        return f"{d0.strftime('%b')} {d0.day}–{d1.strftime('%b')} {d1.day}, {d1.year}"
    return f"{d0.strftime('%b')} {d0.day}, {d0.year}–{d1.strftime('%b')} {d1.day}, {d1.year}"


def month_range(year, month, count):
    out = []
    for i in range(count):
        total = (year * 12 + (month - 1)) + i
        out.append((total // 12, total % 12 + 1))
    return out


def month_range_before(year, month, count):
    """The `count` calendar months immediately BEFORE (year, month), same
    year unless the window crosses a January boundary."""
    start_total = (year * 12 + (month - 1)) - count
    return month_range(start_total // 12, start_total % 12 + 1, count)


def month_range_label(months):
    """e.g. [(2026,3),(2026,4),(2026,5)] -> 'Mar–May 2026'."""
    first_y, first_m = months[0]
    last_y, last_m = months[-1]
    first_txt = date(first_y, first_m, 1).strftime("%b")
    last_txt = date(last_y, last_m, 1).strftime("%b")
    if first_y == last_y:
        return f"{first_txt}–{last_txt} {last_y}"
    return f"{first_txt} {first_y}–{last_txt} {last_y}"


def full_month_span_label(months):
    """Like month_range_label, but with day precision (first of the first
    month through the last calendar day of the last month) -- used for the
    YTD label when there's no Fusion file to source an exact label from."""
    first_y, first_m = months[0]
    last_y, last_m = months[-1]
    last_day = calendar.monthrange(last_y, last_m)[1]
    return format_date_range(date(first_y, first_m, 1), date(last_y, last_m, last_day))


def sum_months(cases_by_ym, account, months):
    return round(sum(cases_by_ym.get((account, y, m), 0.0) for y, m in months), 2)


def sum_months_brand(by_brand, account, brand, months):
    return round(sum(by_brand.get((account, brand, y, m), 0.0) for y, m in months), 2)


def lift_pct(post, pre):
    # A non-positive baseline (zero, or negative -- a window where returns/
    # credits outweighed purchases, seen at a couple of small accounts) has
    # no meaningful "% growth" reading -- "no prior baseline" rather than a
    # sign-flipped or wildly negative-looking percentage.
    if pre is None or pre <= 0:
        return None
    return round((post - pre) / pre * 100, 1)


def latest_complete_month(cases_by_ym):
    """The newest (year, month) with data, EXCLUDING the current real-world
    month if the export happens to include a still-in-progress month (its
    partial total would otherwise understate a YTD comparison)."""
    yms = sorted({(y, m) for (_, y, m) in cases_by_ym.keys()})
    if not yms:
        return None
    latest = yms[-1]
    today = date.today()
    if latest == (today.year, today.month):
        return yms[-2] if len(yms) >= 2 else None
    return latest


def evaluate_cohort(roster, cases_by_ym, by_brand, brands_by_account, reset_year):
    sales_accounts = {a for (a, _, _) in cases_by_ym.keys()}
    cutoff = latest_complete_month(cases_by_ym)
    cutoff_month = cutoff[1] if cutoff and cutoff[0] == reset_year else 12
    ytd_post_months = [(reset_year, m) for m in range(1, cutoff_month + 1)]
    ytd_pre_months = [(reset_year - 1, m) for m in range(1, cutoff_month + 1)]
    ytd_label = f"Jan–{date(reset_year, cutoff_month, 1).strftime('%b')} {reset_year}"
    ytd_pre_label = full_month_span_label(ytd_pre_months)
    ytd_post_label = full_month_span_label(ytd_post_months)

    evaluated, pending = [], []
    for a in roster:
        reset_date = a["resetDate"]
        entry = {
            "account": a["account"], "name": a["name"].title(), "city": a["city"],
            "segment": a["segment"], "resetDate": reset_date.isoformat(),
            "resetMonth": reset_date.strftime("%B %Y"),
        }
        if a["account"] not in sales_accounts:
            entry["status"] = "pending"
            pending.append(entry)
            continue

        post_months = month_range(reset_date.year, reset_date.month, WINDOW_MONTHS)
        pre_months = month_range(reset_date.year - 1, reset_date.month, WINDOW_MONTHS)
        adj_pre_months = month_range_before(reset_date.year, reset_date.month, WINDOW_MONTHS)
        post3 = sum_months(cases_by_ym, a["account"], post_months)
        pre3 = sum_months(cases_by_ym, a["account"], pre_months)
        preAdj = sum_months(cases_by_ym, a["account"], adj_pre_months)

        ytdPost = sum_months(cases_by_ym, a["account"], ytd_post_months)
        ytdPre = sum_months(cases_by_ym, a["account"], ytd_pre_months)

        brand_list = []
        for brand in sorted(brands_by_account.get(a["account"], ())):
            b_post3 = sum_months_brand(by_brand, a["account"], brand, post_months)
            b_pre3 = sum_months_brand(by_brand, a["account"], brand, pre_months)
            b_preAdj = sum_months_brand(by_brand, a["account"], brand, adj_pre_months)
            b_ytdPost = sum_months_brand(by_brand, a["account"], brand, ytd_post_months)
            b_ytdPre = sum_months_brand(by_brand, a["account"], brand, ytd_pre_months)
            if not any((b_post3, b_pre3, b_preAdj, b_ytdPost, b_ytdPre)):
                continue
            brand_list.append({
                "name": brand,
                "post3": b_post3, "pre3": b_pre3, "lift3": lift_pct(b_post3, b_pre3),
                "preAdj": b_preAdj, "liftAdj": lift_pct(b_post3, b_preAdj),
                "ytdPost": b_ytdPost, "ytdPre": b_ytdPre, "liftYtd": lift_pct(b_ytdPost, b_ytdPre),
            })
        brand_list.sort(key=lambda b: -b["post3"])

        entry.update({
            "status": "evaluated",
            "post3": post3, "post3Label": month_range_label(post_months),
            "pre3": pre3, "pre3Label": month_range_label(pre_months), "lift3": lift_pct(post3, pre3),
            "preAdj": preAdj, "preAdjLabel": month_range_label(adj_pre_months), "liftAdj": lift_pct(post3, preAdj),
            "ytdPost": ytdPost, "ytdPre": ytdPre, "liftYtd": lift_pct(ytdPost, ytdPre),
            "brands": brand_list,
        })
        evaluated.append(entry)

    evaluated.sort(key=lambda e: (e["lift3"] is None, -(e["lift3"] or 0)))

    def blended(group, post_key, pre_key):
        pre_sum = sum(e[pre_key] for e in group)
        post_sum = sum(e[post_key] for e in group)
        return {"pre": round(pre_sum, 2), "post": round(post_sum, 2), "liftPct": lift_pct(post_sum, pre_sum)}

    def cohort_summary(group):
        return {
            "accountCount": len(group),
            "cases3": blended(group, "post3", "pre3"),
            "cases3Adj": blended(group, "post3", "preAdj"),
            "casesYtd": blended(group, "ytdPost", "ytdPre"),
            "upCount": sum(1 for e in group if e["lift3"] is not None and e["lift3"] > 0),
            "downCount": sum(1 for e in group if e["lift3"] is not None and e["lift3"] < 0),
            "noBaselineCount": sum(1 for e in group if e["lift3"] is None),
            "upCountYtd": sum(1 for e in group if e["liftYtd"] is not None and e["liftYtd"] > 0),
            "downCountYtd": sum(1 for e in group if e["liftYtd"] is not None and e["liftYtd"] < 0),
            "noBaselineCountYtd": sum(1 for e in group if e["liftYtd"] is None),
        }

    overall = cohort_summary(evaluated)
    by_month = {}
    for e in evaluated:
        by_month.setdefault(e["resetMonth"], []).append(e)
    by_month_summary = {m: cohort_summary(g) for m, g in sorted(by_month.items(), key=lambda kv: kv[1][0]["resetDate"])}
    by_segment = {}
    for e in evaluated:
        if e["segment"]:
            by_segment.setdefault(e["segment"], []).append(e)
    by_segment_summary = {s: cohort_summary(g) for s, g in sorted(by_segment.items())}

    return {
        "resetYear": reset_year,
        "ytdLabel": ytd_label,
        "ytdPreLabel": ytd_pre_label,
        "ytdPostLabel": ytd_post_label,
        "rosterTotal": len(roster),
        "evaluatedCount": len(evaluated),
        "pendingCount": len(pending),
        "overall": overall,
        "byMonth": by_month_summary,
        "bySegment": by_segment_summary,
        "accounts": evaluated,
        "pendingAccounts": pending,
    }


def main():
    sales_2026, by_brand_2026, brands_by_account_2026, misc_2026 = load_monthly_sales(HERE / "sales_2026.csv")
    sales_2025, by_brand_2025, brands_by_account_2025, misc_2025 = load_monthly_sales(HERE / "sales_2025.csv")

    cohort_2026 = evaluate_cohort(load_roster_2026(), sales_2026, by_brand_2026, brands_by_account_2026, 2026)
    cohort_2025 = evaluate_cohort(load_roster_2025(), sales_2025, by_brand_2025, brands_by_account_2025, 2025)

    payload = {
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "windowMonths": WINDOW_MONTHS,
        "cohorts": {"2026": cohort_2026, "2025": cohort_2025},
    }

    data_json = json.dumps(payload, separators=(",", ":"))
    html = HTML.read_text(encoding="utf-8")
    new_html, n = re.subn(
        r'(<script id="reset-data" type="application/json">).*?(</script>)',
        lambda m: m.group(1) + data_json + m.group(2),
        html, count=1, flags=re.S,
    )
    assert n == 1, 'reset-data script tag not found in index.html'
    HTML.write_text(new_html, encoding="utf-8")

    for year, misc in (("2026", misc_2026), ("2025", misc_2025)):
        print(f"{year} sales file: removed {round(misc):,} Brand Family \"Misc\" cases (opaque, unattributed -- see generate.py)")
    for year, c in (("2026", cohort_2026), ("2025", cohort_2025)):
        o = c["overall"]
        print(f"{year} cohort: {c['evaluatedCount']} of {c['rosterTotal']} accounts evaluated "
              f"({c['pendingCount']} pending), YTD = {c['ytdPreLabel']} vs. {c['ytdPostLabel']}")
        print(f"  Overall: {o['upCount']} up / {o['downCount']} down / {o['noBaselineCount']} no baseline")
        print(f"  Blended 3-month Cases lift (YoY): {o['cases3']['liftPct']}%")
        print(f"  Blended 3-month Cases lift (before->after, same year): {o['cases3Adj']['liftPct']}%")
        print(f"  Blended YTD lift: {o['casesYtd']['liftPct']}%")


if __name__ == "__main__":
    main()
