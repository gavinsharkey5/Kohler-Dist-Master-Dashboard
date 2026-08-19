#!/usr/bin/env python3
"""Tier 1 Display Account Program photo recap.

Rebuilds the embedded JSON in index.html from report.xlsx (an iSellBeer
display-photo export, same shape as the display-auction-tracker's
DisplayPhotoReport: one row per SKU on a display photo, with the photo
itself as a hyperlink in the Photo column). Multiple SKU rows share one
photo URL, so photos are deduped by URL and their brand families
aggregated.

The account list, program names, and store contacts come from Ashley
Furman's 7/28/2026 email "Tier 1 Display Account Program" (20 accounts,
program window May-September 2026). The export in report.xlsx is already
filtered to exactly those 20 accounts (see its Filters sheet).

To refresh: export a new report from iSellBeer with the same filters,
save it over report.xlsx, then run: python3 generate.py
Requires: openpyxl (pip install openpyxl).
"""
import datetime
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
REPORT = HERE / "report.xlsx"
INDEX = HERE / "index.html"

# (account #, program name per Ashley's email, store contact)
TIER1_ACCOUNTS = [
    ("21033", "The Liquor Shop", "Danny"),
    ("20106", "Super Wine", "Adjim"),
    ("25008", "Clifton Commons", "Al"),
    ("24011", "Shopper Vineyard Clifton", "Ed"),
    ("21044", "Farmboy", "Rubin / Peter"),
    ("21026", "C Town Supermarket", "Henry"),
    ("28012", "Lucky 7", "Bobby"),
    ("27031", "Home Liquors Passaic", "Gee"),
    ("23004", "Shop Rite Clifton", "Gurerimo"),
    ("14007", "Total Wine Totowa", "Ashley"),
    ("58001", "Total Wine River Edge", "Mike"),
    ("77009", "M & M Liquor Elmwood Park", "SK"),
    ("11003", "Lincoln Park Fine Wines", "Mitch"),
    ("12005", "Shays Hamburgh Tpk", "Darsh"),
    ("76010", "USA Wine Saddle Brook", "Vshal / Tom"),
    ("40022", "Buy Rite Hackensack", "Veral"),
    ("29011", "World of Wine", "Harsh"),
    ("43013", "Buy Rite Fairview", "Harold"),
    ("31060", "Bottle Republic", "Drew / Kayton"),
    ("80005", "Shop Rite Paramus", ""),
]

DT_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4}) (\d{2}):(\d{2}) (AM|PM)")


def parse_dt(s):
    m = DT_RE.match(s or "")
    if not m:
        return None
    mo, da, yr, hh, mi, ap = m.groups()
    hh = int(hh) % 12 + (12 if ap == "PM" else 0)
    return datetime.datetime(int(yr), int(mo), int(da), hh, int(mi))


def main():
    wb = openpyxl.load_workbook(REPORT)
    ws = wb["Report"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    photo_col = header.index("Photo")

    rows = []
    for r in ws.iter_rows(min_row=2):
        d = dict(zip(header, [c.value for c in r]))
        d["PhotoURL"] = r[photo_col].hyperlink.target if r[photo_col].hyperlink else None
        if d["PhotoURL"]:
            rows.append(d)

    # Dedupe SKU rows into photos, per account.
    photos_by_acct = defaultdict(dict)  # acct -> url -> photo dict
    meta_by_acct = {}
    for d in rows:
        acct = str(d["Account #"]).strip()
        meta_by_acct.setdefault(acct, {
            "dba": str(d["DBA"] or "").strip(),
            "address": str(d["Address"] or "").strip(),
            "city": str(d["City"] or "").strip().title(),
        })
        p = photos_by_acct[acct].setdefault(d["PhotoURL"], {
            "url": d["PhotoURL"],
            "dt": str(d["Date/Time"] or "").strip(),
            "taker": str(d["Photo Taker"] or "").strip().title(),
            "brands": [],
            "cases": 0,
        })
        fam = str(d["Brand Family"] or "").strip()
        if fam and fam not in p["brands"]:
            p["brands"].append(fam)
        try:
            p["cases"] += int(d["Quantity"] or 0)
        except (TypeError, ValueError):
            pass

    accounts = []
    total_photos = 0
    all_dts = []
    for acct, name, contact in TIER1_ACCOUNTS:
        photos = sorted(photos_by_acct.get(acct, {}).values(),
                        key=lambda p: parse_dt(p["dt"]) or datetime.datetime.min)
        meta = meta_by_acct.get(acct, {"dba": "", "address": "", "city": ""})
        by_month = []
        for p in photos:
            dt = parse_dt(p["dt"])
            label = dt.strftime("%B %Y") if dt else "Undated"
            if not by_month or by_month[-1]["month"] != label:
                by_month.append({"month": label, "photos": []})
            by_month[-1]["photos"].append(p)
            if dt:
                all_dts.append(dt)
        brand_families = sorted({b for p in photos for b in p["brands"]})
        total_photos += len(photos)
        accounts.append({
            "acct": acct, "name": name, "contact": contact,
            "dba": meta["dba"], "address": meta["address"], "city": meta["city"],
            "photoCount": len(photos),
            "brandFamilyCount": len(brand_families),
            "lastDate": photos[-1]["dt"].split(" ")[0] if photos else None,
            "byMonth": by_month,
        })

    accounts.sort(key=lambda a: -a["photoCount"])
    data = {
        "accounts": accounts,
        "totalPhotos": total_photos,
        "accountCount": len(accounts),
        "firstDate": min(all_dts).strftime("%b %-d, %Y") if all_dts else "",
        "lastDate": max(all_dts).strftime("%b %-d, %Y") if all_dts else "",
    }

    print(f"{len(accounts)} accounts, {total_photos} distinct photos "
          f"({len(rows)} SKU rows), {data['firstDate']} - {data['lastDate']}")
    for a in accounts:
        print(f"  {a['acct']} {a['name']}: {a['photoCount']} photos, {a['brandFamilyCount']} brand families")

    payload = json.dumps(data, indent=1)
    html = INDEX.read_text()
    start_marker = "/* PHOTO_DATA_START */"
    end_marker = "/* PHOTO_DATA_END */"
    start = html.index(start_marker) + len(start_marker)
    end = html.index(end_marker)
    html = html[:start] + f"\nconst DATA = {payload};\n" + html[end:]

    today = datetime.date.today().strftime("%b %-d, %Y")
    ds = html.index("<!-- DATA_REFRESHED_START -->") + len("<!-- DATA_REFRESHED_START -->")
    de = html.index("<!-- DATA_REFRESHED_END -->")
    html = html[:ds] + today + html[de:]

    INDEX.write_text(html)
    print(f"Wrote DATA into index.html, stamped Data Refreshed as {today}")


if __name__ == "__main__":
    main()
