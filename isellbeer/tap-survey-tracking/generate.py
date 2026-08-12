#!/usr/bin/env python3
"""
Rebuilds the embedded data in index.html from the iSellBeer US/THEM mediator
workbook (the tap-audit engine's own working file -- see the tap-audit skill).

Usage (from this folder):
    python3 generate.py
    (requires openpyxl: pip install openpyxl)

Input (keep this filename when re-exporting):
    iSellBeer_TAPS_US_THEM_Mediator.xlsx
    A multi-sheet workbook. Only two sheets are actually read; the rest
    (Master - US vs THEM, Brand Family Territory(ies), Whitelist, Brand
    Crosswalk, Brands (Enc), Customers Table (Enc), Master Matrix View) are
    the audit engine's own reference tables -- kept in the repo for
    provenance/future use, not parsed here yet:
      - The raw survey sheet: one row per surveyed tap (Account #, DBA,
        Distribution Area, Address, City, Date/Time, Photos, Route / Sales
        Rep, District Manager, Brand, Brand Family, Supplier, # of Taps,
        Distributor). This is also where real photo links live -- the
        "Photos" cell's hyperlink -- which a CSV export of the same report
        would flatten to plain display text. This sheet's own tab name has
        changed between exports (Sheet6, then Sheet9) as Kohler edits the
        workbook, so generate.py finds it by its header row (must contain
        "Account #", "Route / Sales Rep" and "Photos", but not "Corrected
        Distributor") rather than a hardcoded sheet name -- see
        find_raw_sheet() below.
      - "iSellBeer Import Template": the same rows plus the audit engine's
        helper/output columns, notably "Distributor" (the ORIGINAL iSellBeer
        app flag, pre-audit) and "Corrected Distributor" (the engine's final
        US/THEM ruling after checking the product catalog + territory
        tables). This is the authoritative status -- more current than the
        raw sheet's own Distributor column, which has lagged behind for a
        handful of rows in past exports (an incomplete write-back, not a
        judgment call).
    The two sheets are joined on "#" (row number) to get: raw fields + real
    photo (raw sheet) and corrected status + audit reasoning (Import
    Template).

index.html does all the rep -> account -> brand grouping client-side from
the flat row list emitted here, the same way it re-groups after search/
county/status filtering -- so there's exactly one grouping implementation
to keep correct, not two. This export is one snapshot in time (each account
has a single visit date here), so "most recent visit"/"most recent photo"
is just that snapshot's values.

Run: python3 generate.py
"""
import json, re, os, datetime
from collections import defaultdict
from openpyxl import load_workbook

# Policy override confirmed with the user 2026-08-12 (per the workbook's own
# "Unverified Brands" sheet): taps from these suppliers count as US even
# though the Import Template's Corrected Distributor formula hasn't been
# updated to match yet -- it still defaults them to THEM under the older
# "No Encompass Match" rule. The raw sheet's own Distributor column already
# carries the corrected value (manually ruled, green-highlighted at the
# source), so trust that instead of Corrected Distributor for just these
# rows. Don't re-derive this from scratch -- ask the user if a future export
# still disagrees.
SUPPLIER_STATUS_OVERRIDE_KEYWORDS = ('(IN-HOUSE)', 'OTHER HALF', 'INDUSTRIAL ARTS', 'PABST')


def resolve_status(raw_status, corrected, supplier):
    if raw_status == 'US' and corrected == 'THEM' and any(k in (supplier or '').upper() for k in SUPPLIER_STATUS_OVERRIDE_KEYWORDS):
        return raw_status
    return corrected


# Fill-down gap (found 2026-08-12, still present in the 8.12.26 exports):
# Import Template's Corrected Distributor (col Y) formula was never dragged
# down to cover the newest ~116 rows, so those cells are blank in the source
# file -- not miscalculated, just never written. Rather than wait on another
# export, replicate the exact same formula those cells would contain,
# ourselves, from columns that ARE already filled in every row:
#   =IF($B="","",IF(OR(Expected Distributor="US",Expected Distributor="THEM"),Expected Distributor,Distributor))
# i.e. trust Expected Distributor when it's a real US/THEM verdict, else fall
# back to the Import Template's own (pre-audit) Distributor column -- same
# fallback the formula itself uses for "Unable to Determine"/"No Territory
# Data" rows. Only kicks in when Corrected Distributor is genuinely blank;
# once a real export fills column Y down, this is a no-op.
def fill_corrected(corrected, expected, tmpl_distributor):
    if corrected in ('US', 'THEM'):
        return corrected
    if expected in ('US', 'THEM'):
        return expected
    return tmpl_distributor or 'UNVERIFIED'

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(HERE, 'iSellBeer_TAPS_US_THEM_Mediator.xlsx')
HTML = os.path.join(HERE, 'index.html')

RAW_COLUMNS = ['#', 'Account #', 'DBA', 'Distribution Area', 'Address', 'City', 'Date/Time', 'Photos',
               'Route / Sales Rep', 'District Manager', 'Brand', 'Brand Family', 'Supplier', '# of Taps', 'Distributor']


def find_raw_sheet(wb):
    """The raw survey sheet's tab name has changed between exports (Sheet6,
    then Sheet9) as Kohler edits the workbook -- find it by header shape
    instead of a hardcoded name, so the next rename doesn't break this."""
    for name in wb.sheetnames:
        header = [c.value for c in wb[name][1]]
        if 'Account #' in header and 'Route / Sales Rep' in header and 'Photos' in header and 'Corrected Distributor' not in header:
            return wb[name]
    raise SystemExit("Could not find the raw tap-survey sheet (expected a sheet with 'Account #', "
                      "'Route / Sales Rep' and 'Photos' columns, but no 'Corrected Distributor' column)")


def sheet_rows(ws, columns, with_photo_link=False):
    header = [c.value for c in ws[1]]
    col_idx = {name: header.index(name) for name in columns}
    photo_idx = col_idx.get('Photos') if with_photo_link else None
    out = {}
    for r in ws.iter_rows(min_row=2):
        num_cell = r[col_idx['#']]
        if num_cell.value in (None, ''):
            continue
        row = {c: ('' if r[col_idx[c]].value is None else str(r[col_idx[c]].value)) for c in columns}
        if with_photo_link:
            row['photo'] = r[photo_idx].hyperlink.target if r[photo_idx].hyperlink else None
        out[str(int(float(num_cell.value)))] = row
    return out


def parse_rep(raw):
    m = re.match(r'^\s*\d+\s*-\s*(.+)$', raw)
    name = m.group(1).strip() if m else raw.strip()
    return ' '.join(w[:1].upper() + w[1:] for w in name.split(' '))


def parse_datetime(raw):
    return datetime.datetime.strptime(raw.strip(), '%m/%d/%Y %I:%M %p')


def num(s):
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 1


def audit_reason(t_row):
    result = (t_row.get('Audit Result') or '').strip()
    canonical = (t_row.get('Canonical Brand Family') or '').strip()
    if result == 'Review' or canonical in ('', 'Not Mapped'):
        return 'No product-catalog match for this brand family — defaults to THEM'
    if result == 'MISMATCH':
        return 'iSellBeer’s own flag was wrong here — corrected against territory rules'
    return 'Confirmed against the product catalog and territory rules'


# ---------- segment (per Kohler, 2026-08-07: "Brand Segments (iSell)" and
# "Product Segments (Enc)", two new sheets in this same workbook, added
# specifically to answer "what segment is this brand" -- Encompass is the
# stated final source of truth when the two disagree) ----------
ACRONYMS = {'Ipa': 'IPA', 'Pos': 'POS', 'Fmb': 'FMB', 'Fab': 'FAB'}


def titleize(s):
    """iSell's Segment values are ALL CAPS in the source sheet; title-case
    for display but keep known acronyms (IPA, POS, ...) from being mangled
    into 'Ipa'/'Pos' by str.title()."""
    if not s:
        return s
    t = s.title()
    for wrong, right in ACRONYMS.items():
        t = re.sub(r'\b' + wrong + r'\b', right, t)
    return t


_KEG_SIZE_PATTERNS = [
    r'\s*\(\s*\d+(\.\d+)?\s*gal(lon)?\s*(keg)?\s*\)\s*$',
    r'\s+\d+(\.\d+)?\s*(l|liter|litre)s?\s*keg\s*$',
    r'\s+\d+/\d+\s*(bbl|barrel|keg)\s*$',
    r'\s*keg\s*\d+(\.\d+)?\s*(gal(lon)?\s*keg)?\s*$',
    r'\s+\d+(\.\d+)?\s*(gal(lon)?\s*keg|gal(lon)?|bbl|barrel)\s*$',
    r'\s+keg\s*$',
    r'\s+\d+(\.\d+)?\s*$',
]


def strip_keg_size(name):
    """'Coors Light 15.5 Gal Keg' -> 'Coors Light'. Product Segments (Enc)
    is a per-SKU catalog (one row per keg size); this collapses it to a
    base product name comparable against the tap survey's Brand text.
    Iterates since some names carry more than one trailing size/keg token
    (e.g. 'Sierra Nevada Big Little Thing Keg15.5 Gal Keg')."""
    n = name
    changed = True
    while changed:
        changed = False
        for pattern in _KEG_SIZE_PATTERNS:
            n2 = re.sub(pattern, '', n, flags=re.I).strip()
            if n2 != n:
                n, changed = n2, True
    return n


def build_segment_resolver(wb):
    """Returns (resolve_segment(brand, brand_family) -> str|None, stats dict).

    Two sources, per Kohler 2026-08-07:
      - 'Brand Segments (iSell)': one row per surveyed tap (same row count
        as the raw survey sheet, joined here by Brand/Brand Family text
        instead of by row number since its own "#" doesn't line up with
        the raw sheet's) giving that brand's Segment. Covers ~100% of taps
        by volume (it's effectively the same catalog, just re-exported)
        but its "Segment" column mixes two different ideas row to row --
        sometimes a beer style (Wheat Beer, Pilsner And Pale Lager),
        sometimes a price tier (Craft, Import, Domestic) -- because that's
        genuinely what's in the source column, not a bug here.
      - 'Product Segments (Enc)': Encompass's own product catalog (Sub-
        Segments: Beer - Craft/Import/Premium/Economy, Cider, ...), joined
        by stripping each SKU's keg-size suffix down to a base product
        name and matching that against the tap's Brand / Brand Family --
        directly, or via 'Brand Crosswalk' (the same sheet used for the
        executive-overview dashboard's velocity join) when the names
        don't line up as-is. Encompass only ever carries what WE sell, so
        this resolves for a minority of competitor brands by design.
    Final value = Encompass's Sub-Segment when resolvable (stated source of
    truth), else iSell's Segment, else None -- never a guess.
    """
    seg_by_brand, seg_by_bf = {}, {}
    isell = wb['Brand Segments (iSell)']
    isell_idx = {c.value: i for i, c in enumerate(isell[1])}
    for r in isell.iter_rows(min_row=2, values_only=True):
        seg = r[isell_idx['Segment']]
        if not seg:
            continue
        b = r[isell_idx['Brand']]
        bf = r[isell_idx['Brand Family']]
        if b:
            seg_by_brand.setdefault(b.strip().upper(), {}).setdefault(seg, 0)
            seg_by_brand[b.strip().upper()][seg] += 1
        if bf:
            seg_by_bf.setdefault(bf.strip().upper(), {}).setdefault(seg, 0)
            seg_by_bf[bf.strip().upper()][seg] += 1
    seg_by_brand = {k: max(v, key=v.get) for k, v in seg_by_brand.items()}
    seg_by_bf = {k: max(v, key=v.get) for k, v in seg_by_bf.items()}

    enc_sub_by_base = {}
    enc = wb['Product Segments (Enc)']
    enc_idx = {c.value: i for i, c in enumerate(enc[1])}
    for r in enc.iter_rows(min_row=2, values_only=True):
        name = r[enc_idx['Product Name']]
        sub = r[enc_idx['Sub-Segments']]
        if not name or not sub:
            continue
        base = strip_keg_size(name).strip().upper()
        enc_sub_by_base.setdefault(base, {}).setdefault(sub, 0)
        enc_sub_by_base[base][sub] += 1
    enc_sub_by_base = {k: max(v, key=v.get) for k, v in enc_sub_by_base.items()}
    enc_base_keys = list(enc_sub_by_base.keys())

    crosswalk = {}
    cw = wb['Brand Crosswalk']
    cw_idx = {c.value: i for i, c in enumerate(cw[1])}
    for r in cw.iter_rows(min_row=2, values_only=True):
        rb = r[cw_idx['Report Brand Family']]
        mapped = r[cw_idx['Mapped Encompass Brand Family']]
        if rb:
            crosswalk[rb.strip().upper()] = mapped.strip().upper() if mapped else None

    def enc_lookup(key):
        if key in enc_sub_by_base:
            return enc_sub_by_base[key]
        hits = {k for k in enc_base_keys if k.startswith(key + ' ') or key.startswith(k + ' ')}
        if len(hits) == 1:
            return enc_sub_by_base[hits.pop()]
        return None

    stats = {'enc': 0, 'isell': 0, 'none': 0}

    def resolve_segment(brand, brand_family):
        key_b = (brand or '').strip().upper()
        key_bf = (brand_family or '').strip().upper()
        enc_val = enc_lookup(key_b) if key_b else None
        if enc_val is None and key_bf:
            enc_val = enc_lookup(key_bf)
        if enc_val is None and key_bf in crosswalk and crosswalk[key_bf]:
            enc_val = enc_lookup(crosswalk[key_bf])
        if enc_val is not None:
            stats['enc'] += 1
            return enc_val
        isell_val = seg_by_brand.get(key_b) or seg_by_bf.get(key_bf)
        if isell_val is not None:
            stats['isell'] += 1
            return titleize(isell_val)
        stats['none'] += 1
        return None

    return resolve_segment, stats


# ---------- load ----------
wb = load_workbook(XLSX_PATH, data_only=True)
raw_by_num = sheet_rows(find_raw_sheet(wb), RAW_COLUMNS, with_photo_link=True)
tmpl_by_num = sheet_rows(wb['iSellBeer Import Template'], RAW_COLUMNS + ['Corrected Distributor', 'Expected Distributor', 'Audit Result', 'Canonical Brand Family'])
resolve_segment, segment_stats = build_segment_resolver(wb)

records = []
for n, raw in raw_by_num.items():
    t = tmpl_by_num.get(n)
    if t is None:
        continue
    if not raw['Route / Sales Rep'].strip():
        # No rep to attribute this row to (seen once so far: a stray 0-tap
        # placeholder row with no brand either) -- would otherwise surface
        # as a blank-named rep card with nothing real in it.
        continue
    visited = parse_datetime(raw['Date/Time'])
    raw_status = (t['Distributor'] or '').strip().upper() or 'UNVERIFIED'
    corrected_cell = (t['Corrected Distributor'] or '').strip().upper()
    expected_cell = (t['Expected Distributor'] or '').strip().upper()
    corrected = fill_corrected(corrected_cell, expected_cell, raw_status)
    # Sheet9's own Distributor (not Import Template's copy of it) is where the
    # manual "Unverified Brands" green-highlight corrections actually live.
    sheet_status = (raw['Distributor'] or '').strip().upper() or 'UNVERIFIED'
    final_status = resolve_status(sheet_status, corrected, raw['Supplier'])
    records.append({
        'account': raw['Account #'].strip(),
        'dba': raw['DBA'].strip(),
        'county': raw['Distribution Area'].strip(),
        'address': raw['Address'].strip(),
        'city': raw['City'].strip(),
        'visited': visited.isoformat(),
        'visitedDisplay': visited.strftime('%b %-d, %Y'),
        'rep': parse_rep(raw['Route / Sales Rep']),
        'districtManager': raw['District Manager'].strip(),
        'brand': raw['Brand'].strip(),
        'brandFamily': raw['Brand Family'].strip(),
        'supplier': raw['Supplier'].strip(),
        'segment': resolve_segment(raw['Brand'], raw['Brand Family']) or 'Unclassified',
        'taps': num(raw['# of Taps']),
        'status': final_status,
        'flipped': final_status != raw_status,
        'rawStatus': raw_status,
        'reason': audit_reason(t),
        'photo': raw['photo'],
    })

counties = sorted(set(r['county'] for r in records))
reps = sorted(set(r['rep'] for r in records))
district_managers = sorted(set(r['districtManager'] for r in records if r['districtManager']))
segments = sorted(set(r['segment'] for r in records), key=lambda s: (s == 'Unclassified', s))
accounts = set((r['account'], r['dba']) for r in records)

total_taps = sum(r['taps'] for r in records)
total_us = sum(r['taps'] for r in records if r['status'] == 'US')
total_them = sum(r['taps'] for r in records if r['status'] == 'THEM')
total_unv = total_taps - total_us - total_them
total_flipped = sum(1 for r in records if r['flipped'])
photos_available = len(set(r['account'] for r in records if r['photo']))

payload = {
    'generatedAt': datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M UTC'),
    'photosSource': 'xlsx',
    'counties': counties,
    'reps': reps,
    'districtManagers': district_managers,
    'segments': segments,
    'summary': {
        'taps': total_taps, 'us': total_us, 'them': total_them, 'unv': total_unv,
        'usPct': round(total_us / total_taps * 100, 1) if total_taps else 0,
        'themPct': round(total_them / total_taps * 100, 1) if total_taps else 0,
        'repCount': len(reps), 'accountCount': len(accounts),
        'photosAvailable': photos_available,
        'flipped': total_flipped,
    },
    'records': records,
}

data_json = json.dumps(payload, separators=(',', ':'))

html = open(HTML, encoding='utf-8').read()
new_html, n = re.subn(
    r'(<script id="tap-data" type="application/json">).*?(</script>)',
    lambda m: m.group(1) + data_json + m.group(2),
    html, count=1, flags=re.S,
)
assert n == 1, 'tap-data script tag not found in index.html'
open(HTML, 'w', encoding='utf-8').write(new_html)

print(f"{len(reps)} reps, {len(accounts)} accounts, {total_taps} taps "
      f"({total_us} ours / {total_them} competitor / {total_unv} unverified)")
print(f"Corrected vs. iSellBeer's own raw flag: {total_flipped} of {len(records)} rows")
print(f"Accounts with a photo on file: {photos_available} of {len(accounts)}")
seg_total = sum(segment_stats.values())
print(f"Segment resolution ({len(segments)} distinct values): "
      f"{segment_stats['enc']} rows from Product Segments (Enc), "
      f"{segment_stats['isell']} rows from Brand Segments (iSell) only, "
      f"{segment_stats['none']} unclassified"
      + (f" (of {seg_total})" if seg_total else ""))
