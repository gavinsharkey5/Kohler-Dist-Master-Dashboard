"""Turn the delivered "Audit Matrix vF1" workbook into the mediator
workbook the tap tracker reads. Two defects in the delivered file make it
unreadable by generate.py as-is:

  1. Sheet9's "#" restarts at 1 for the 239 appended rows -> 196 duplicate
     join keys, which collide in generate.py's dict (last row wins).
  2. "iSellBeer Import Template" holds only 82 of the 5,581 surveyed taps,
     so generate.py's `if t is None: continue` would drop 5,499 of them.

Fix 1 renumbers "#" 1..N in the sheet's delivered order. Fix 2 is the
tap-audit skill's step 2 -- repaste every raw row into the template and
re-evaluate the audit columns P:Y. LibreOffice cannot open this workbook
in this environment, so P:Y are computed by engine.py, a line-by-line
Python replica of those same formulas, verified to reproduce Excel's own
S/T/U/V/W/X/Y on all 5,342 rows of the previous mediator.

openpyxl cannot hold a formula and its cached value at once, so the
sheets generate.py reads must carry values: the Import Template (A:Y) and
"Master - US vs THEM" (A/B/F/G, which drive the brand-rights payload) are
written as values from the delivered file's own cache. Every other sheet
keeps its formulas and recalculates when Excel opens it.
"""
import shutil, sys
from pathlib import Path
import openpyxl
sys.path.insert(0, str(Path(__file__).resolve().parent))
import audit_engine as engine

SRC, DST = Path(sys.argv[1]), Path(sys.argv[2])
shutil.copy(SRC, DST)

vals = openpyxl.load_workbook(SRC, data_only=True)      # cached values
wb = openpyxl.load_workbook(DST, data_only=False)       # structure/formulas
refs = engine.build(vals)

raw = wb["Sheet9"]
rows = [[c.value for c in r] for r in raw.iter_rows(min_row=2)]
rows = [r for r in rows if any(v is not None for v in r)]
for i, r in enumerate(rows, start=1):
    raw.cell(row=i + 1, column=1).value = i
    r[0] = i
print(f"Sheet9: {len(rows)} rows, '#' renumbered 1..{len(rows)}")

tmpl = wb["iSellBeer Import Template"]
audit_stats = {}
for i, r in enumerate(rows, start=2):
    for c in range(1, 16):
        tmpl.cell(row=i, column=c).value = r[c - 1]
    out = engine.audit({"B": r[1], "D": r[3], "K": r[10], "L": r[11], "O": r[14]}, refs)
    for j, v in enumerate(out):
        tmpl.cell(row=i, column=16 + j).value = v
    audit_stats[out[8]] = audit_stats.get(out[8], 0) + 1     # X = Audit Result
last = len(rows) + 1
for i in range(last + 1, tmpl.max_row + 1):
    for c in range(1, 26):
        tmpl.cell(row=i, column=c).value = None
print(f"Import Template: A2:Y{last} written as values; audit results {audit_stats}")

mws = wb["Master - US vs THEM"]
for i, r in enumerate(vals["Master - US vs THEM"].iter_rows(min_row=2, values_only=True), start=2):
    for c in (1, 2, 6, 7):
        mws.cell(row=i, column=c).value = r[c - 1]
print(f"Master - US vs THEM: A/B/F/G materialised to values ({mws.max_row - 1} rows)")

wb.save(DST)
print("wrote", DST)
