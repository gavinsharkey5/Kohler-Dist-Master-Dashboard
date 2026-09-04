#!/usr/bin/env python3
"""
Rebuilds the embedded data in index.html: a top-line, phone-friendly summary
of our tap-handle share vs. competitors', for the head of the company (not
the reps -- see ../tap-survey-tracking/ for the rep-facing drill-down that
this reuses raw data from, but does not modify).

Inputs (keep these filenames when refreshing):
  iSellBeer_TAPS_US_THEM_Audit_Matrix.xlsx
      Same shape as ../tap-survey-tracking/'s mediator workbook (see that
      folder's README for the full sheet-by-sheet explanation) -- one row per
      surveyed tap handle, joined against the audit engine's Corrected
      Distributor (US/THEM) ruling. Also reads its "Brand Crosswalk" sheet
      (Report Brand Family -> Mapped Encompass Brand Family) to connect tap
      brand families to the Encompass units-sold export below.
  encompass_units_sold.csv
      RDE "iSellBeer TAPS Exec Overview" export: Customer Num, Customer Name,
      Area, Shipping Address, City, Date, Sales Rep Name, District Manager
      Name, Brand, Brand Family, Supplier, Units <year>. Used only for the
      velocity section -- units sold, at the same accounts we have a tap
      survey for, per our own brand families (Encompass only carries OUR
      sales, so this can never cover competitor brands).

"# of Taps" (not row count) is what's summed for every tally here, same
convention as ../tap-survey-tracking/ -- an account with 3 handles of one
brand counts as 3 taps, not 1 row.

Distribution-area data-quality fix (per Kohler, 2026-07-30): three area
labels in the source aren't real distribution areas --
  - "Passaic-FF" is folded into "Passaic" (same area, different label).
  - "Sales" is a placeholder for rows that were never assigned a real area --
    every one is re-assigned here from its City (majority-vote against every
    correctly-labeled row sharing that city), falling back to Address for a
    handful of cities that don't appear anywhere else in the export (Old
    Tappan and Ridgefield -> Bergen, Passaic city -> Passaic area -- all
    well-established NJ municipalities, not a judgment call).
  - "Morris 2" is left alone -- it's a real area, just not one of Kohler's
    named core-market or non-focus areas, so it surfaces in its own small
    "Other Areas" bucket rather than being folded into Morris 1/3.

Run: python3 generate.py
"""
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

# Policy override confirmed with the user 2026-08-12 (per the workbook's own
# "Unverified Brands" sheet): taps from these suppliers count as US even
# though the Import Template's Corrected Distributor formula hasn't been
# updated to match yet -- it still defaults them to THEM under the older
# "No Encompass Match" rule. The raw sheet's own Distributor column already
# carries the corrected value (manually ruled, green-highlighted at the
# source), so trust that instead of Corrected Distributor for just these
# rows. Same rule as ../tap-survey-tracking/generate.py -- don't re-derive
# from scratch, ask the user if a future export still disagrees.
SUPPLIER_STATUS_OVERRIDE_KEYWORDS = ('(IN-HOUSE)', 'OTHER HALF', 'INDUSTRIAL ARTS', 'PABST')


def resolve_status(raw_status, corrected, supplier):
    if raw_status == 'US' and corrected == 'THEM' and any(k in (supplier or '').upper() for k in SUPPLIER_STATUS_OVERRIDE_KEYWORDS):
        return raw_status
    return corrected


# Fill-down gap (found 2026-08-12, still present in the 8.12.26 exports):
# Import Template's Corrected Distributor (col Y) formula was never dragged
# down to cover the newest ~116 rows, so those cells are blank in the source
# file -- not miscalculated, just never written. Rather than wait on another
# export, replicate the exact same formula those cells would contain,
# ourselves, from columns that ARE already filled in every row: trust
# Expected Distributor when it's a real US/THEM verdict, else fall back to
# the Import Template's own (pre-audit) Distributor column -- same fallback
# the formula itself uses for "Unable to Determine"/"No Territory Data"
# rows. Same rule as ../tap-survey-tracking/generate.py. Only kicks in when
# Corrected Distributor is genuinely blank; once a real export fills column
# Y down, this is a no-op.
def fill_corrected(corrected, expected, tmpl_distributor):
    if corrected in ('US', 'THEM'):
        return corrected
    if expected in ('US', 'THEM'):
        return expected
    return tmpl_distributor or 'UNVERIFIED'

HERE = Path(__file__).parent
XLSX_PATH = HERE / "iSellBeer_TAPS_US_THEM_Audit_Matrix.xlsx"
UNITS_CSV = HERE / "encompass_units_sold.csv"
HTML = HERE / "index.html"

RAW_COLUMNS = ['#', 'Account #', 'DBA', 'Distribution Area', 'Address', 'City', 'Date/Time',
               'Route / Sales Rep', 'District Manager', 'Brand', 'Brand Family', 'Supplier',
               '# of Taps', 'Distributor']

# Per Kohler, 2026-07-30.
CORE_AREAS = ['BERGEN', 'SUSSEX', 'PASSAIC', 'MORRIS 1', 'MORRIS 3']
NON_FOCUS_AREAS = ['ESSEX', 'HUDSON', 'UNION']
# Manually-known NJ municipalities for the handful of "Sales"-labeled rows
# whose city doesn't appear anywhere else in the export with a real area.
CITY_AREA_FALLBACK = {'OLD TAPPAN': 'BERGEN', 'RIDGEFIELD': 'BERGEN', 'PASSAIC': 'PASSAIC'}
TOP_N_BRANDS = 8
TOP_N_AREA_BRANDS = 5


def find_raw_sheet(wb):
    for name in wb.sheetnames:
        header = [c.value for c in wb[name][1]]
        if 'Account #' in header and 'Route / Sales Rep' in header and 'Corrected Distributor' not in header:
            return wb[name]
    raise SystemExit("Could not find the raw tap-survey sheet")


def sheet_rows(ws, columns):
    header = [c.value for c in ws[1]]
    col_idx = {name: header.index(name) for name in columns}
    out = {}
    for r in ws.iter_rows(min_row=2):
        num_cell = r[col_idx['#']]
        if num_cell.value in (None, ''):
            continue
        row = {c: ('' if r[col_idx[c]].value is None else str(r[col_idx[c]].value)) for c in columns}
        out[str(int(float(num_cell.value)))] = row
    return out


def num(s):
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 1


def pct(part, whole):
    return round(part / whole * 100, 1) if whole else 0.0


# ---------- segment resolution (copied from ../tap-survey-tracking/generate.py,
# build_segment_resolver & helpers -- keep the two in sync if the workbook's
# segment sheets change shape). Two sources, per Kohler 2026-08-07: Encompass's
# "Product Segments (Enc)" Sub-Segment is the stated source of truth when
# resolvable, else iSell's "Brand Segments (iSell)" Segment, else None --
# never a guess. ----------
ACRONYMS = {'Ipa': 'IPA', 'Pos': 'POS', 'Fmb': 'FMB', 'Fab': 'FAB'}


def titleize(s):
    if not s:
        return s
    t = s.title()
    for wrong, right in ACRONYMS.items():
        t = re.sub(r'\b' + wrong + r'\b', right, t)
    return t


_KEG_SIZE_PATTERNS = [
    r'\s*\(\s*\d+(\.\d+)?\s*gal(lon)?\s*(keg)?\s*\)\s*$',
    r'\s+\d+(\.\d+)?\s*(l|liter|litre)s?\s*keg\s*$',
    r'\s+\d+/\d+\s*(bbl|barrel|keg)\s*$',
    r'\s*keg\s*\d+(\.\d+)?\s*(gal(lon)?\s*keg)?\s*$',
    r'\s+\d+(\.\d+)?\s*(gal(lon)?\s*keg|gal(lon)?|bbl|barrel)\s*$',
    r'\s+keg\s*$',
    r'\s+\d+(\.\d+)?\s*$',
]


def strip_keg_size(name):
    n = name
    changed = True
    while changed:
        changed = False
        for pattern in _KEG_SIZE_PATTERNS:
            n2 = re.sub(pattern, '', n, flags=re.I).strip()
            if n2 != n:
                n, changed = n2, True
    return n


def build_segment_resolver(wb):
    seg_by_brand, seg_by_bf = {}, {}
    isell = wb['Brand Segments (iSell)']
    isell_idx = {c.value: i for i, c in enumerate(isell[1])}
    for r in isell.iter_rows(min_row=2, values_only=True):
        seg = r[isell_idx['Segment']]
        if not seg:
            continue
        b = r[isell_idx['Brand']]
        bf = r[isell_idx['Brand Family']]
        if b:
            seg_by_brand.setdefault(b.strip().upper(), {}).setdefault(seg, 0)
            seg_by_brand[b.strip().upper()][seg] += 1
        if bf:
            seg_by_bf.setdefault(bf.strip().upper(), {}).setdefault(seg, 0)
            seg_by_bf[bf.strip().upper()][seg] += 1
    seg_by_brand = {k: max(v, key=v.get) for k, v in seg_by_brand.items()}
    seg_by_bf = {k: max(v, key=v.get) for k, v in seg_by_bf.items()}

    enc_sub_by_base = {}
    enc = wb['Product Segments (Enc)']
    enc_idx = {c.value: i for i, c in enumerate(enc[1])}
    for r in enc.iter_rows(min_row=2, values_only=True):
        name = r[enc_idx['Product Name']]
        sub = r[enc_idx['Sub-Segments']]
        if not name or not sub:
            continue
        base = strip_keg_size(name).strip().upper()
        enc_sub_by_base.setdefault(base, {}).setdefault(sub, 0)
        enc_sub_by_base[base][sub] += 1
    enc_sub_by_base = {k: max(v, key=v.get) for k, v in enc_sub_by_base.items()}
    enc_base_keys = list(enc_sub_by_base.keys())

    seg_crosswalk = {}
    cw = wb['Brand Crosswalk']
    cw_idx = {c.value: i for i, c in enumerate(cw[1])}
    for r in cw.iter_rows(min_row=2, values_only=True):
        rb = r[cw_idx['Report Brand Family']]
        mapped = r[cw_idx['Mapped Encompass Brand Family']]
        if rb:
            seg_crosswalk[rb.strip().upper()] = mapped.strip().upper() if mapped else None

    def enc_lookup(key):
        if key in enc_sub_by_base:
            return enc_sub_by_base[key]
        hits = {k for k in enc_base_keys if k.startswith(key + ' ') or key.startswith(k + ' ')}
        if len(hits) == 1:
            return enc_sub_by_base[hits.pop()]
        return None

    def resolve_segment(brand, brand_family):
        key_b = (brand or '').strip().upper()
        key_bf = (brand_family or '').strip().upper()
        enc_val = enc_lookup(key_b) if key_b else None
        if enc_val is None and key_bf:
            enc_val = enc_lookup(key_bf)
        if enc_val is None and key_bf in seg_crosswalk and seg_crosswalk[key_bf]:
            enc_val = enc_lookup(seg_crosswalk[key_bf])
        if enc_val is not None:
            return enc_val
        isell_val = seg_by_brand.get(key_b) or seg_by_bf.get(key_bf)
        if isell_val is not None:
            return titleize(isell_val)
        return None

    return resolve_segment


# The workbook's segment values mix price tiers ("Craft", "Beer - Import")
# with beer styles ("Wheat Beer", "Pilsner And Pale Lager") row to row --
# that's genuinely what's in the source, not a bug. Only the tier-shaped
# values map cleanly onto this page's executive buckets; style-shaped values
# are deliberately NOT mapped (a style says nothing reliable about tier --
# "Pilsner And Pale Lager" covers both Stella Artois and Bud Light), so
# brands whose rows only ever carry style labels keep using index.html's
# existing brand-name heuristic instead (classifySegment() falls back
# whenever no `segment` field was emitted for a brand).
SEGMENT_NORMALIZE = {
    'Beer - Craft': 'craft', 'Craft': 'craft', 'Craft - Seasonal': 'craft',
    'Beer - Import': 'import', 'Import': 'import', 'Imports': 'import',
    'Beer - Premium': 'domestic', 'Beer - Economy': 'domestic', 'Domestic': 'domestic',
    'Above Premium': 'domestic',
    'Cider': 'cider', 'Hard Cider': 'cider', 'FAB': 'cider', 'FMB': 'cider',
    'Seltzer': 'cider', 'Hard Iced Tea': 'cider', 'Vodka': 'cider',
}

# Catch-all/generic brand labels -- never attributed a segment, never listed
# as a "dual-sourced brand" (same spirit as index.html's FACT_EXCLUDE_NAMES:
# a generic bucket isn't a brand).
GENERIC_BRAND_NAMES = {'OTHER SUPPLIER', 'OTHER BRAND FAMILY', 'OTHER', 'OTHER SUPPLIER (IN-HOUSE)', ''}

# ---------- load tap survey ----------
wb = load_workbook(XLSX_PATH, data_only=True)
raw_by_num = sheet_rows(find_raw_sheet(wb), RAW_COLUMNS)
tmpl_by_num = sheet_rows(wb['iSellBeer Import Template'], RAW_COLUMNS + ['Corrected Distributor', 'Expected Distributor'])

records = []
for n, raw in raw_by_num.items():
    t = tmpl_by_num.get(n)
    if t is None or not raw['Route / Sales Rep'].strip():
        continue
    area = raw['Distribution Area'].strip().upper()
    if area == 'PASSAIC-FF':
        area = 'PASSAIC'
    tmpl_distributor = (t['Distributor'] or '').strip().upper() or 'UNVERIFIED'
    corrected_cell = (t['Corrected Distributor'] or '').strip().upper()
    expected_cell = (t['Expected Distributor'] or '').strip().upper()
    corrected = fill_corrected(corrected_cell, expected_cell, tmpl_distributor)
    # Sheet9's own Distributor (not Import Template's copy of it) is where the
    # manual "Unverified Brands" green-highlight corrections actually live.
    sheet_status = (raw['Distributor'] or '').strip().upper() or 'UNVERIFIED'
    records.append({
        'account': raw['Account #'].strip(),
        'dba': raw['DBA'].strip(),
        'area': area,
        'city': raw['City'].strip().upper(),
        'address': raw['Address'].strip().upper(),
        'brand': raw['Brand'].strip(),
        'brandFamily': raw['Brand Family'].strip().upper(),
        'supplier': raw['Supplier'].strip(),
        'taps': num(raw['# of Taps']),
        'status': resolve_status(sheet_status, corrected, raw['Supplier']),
        'visited': datetime.strptime(raw['Date/Time'].strip(), '%m/%d/%Y %I:%M %p').isoformat(),
    })

# ---------- current lineup only (added 2026-09-04) ----------
# Same defect and same fix as ../tap-survey-tracking/generate.py -- read that
# file's "current lineup vs. survey history" block for the full reasoning;
# both pages join the same two sheets of the same workbook, so both had it.
# In short: the 9.4.26 export is the first where an account can be surveyed
# more than once (a survey submission = one Account # + Date/Time pair), and
# summing every row counted a brand once per pass and kept pouring brands
# that had since come off the wall. Only the newest pass per account counts
# toward what is on tap TODAY, which is the only question this page answers
# -- so the filter goes here, above every total, split, area card, velocity
# join and flip-target list, rather than in each of them.
#
# This page keeps no history section (it is the one-scroll executive view,
# not the rep drill-down); the tap tracker holds the superseded passes.
survey_dates = defaultdict(set)
for r in records:
    survey_dates[r['account']].add(r['visited'])
current_visit = {acct: max(dates) for acct, dates in survey_dates.items()}
superseded_taps = sum(r['taps'] for r in records if r['visited'] != current_visit[r['account']])
resurveyed_accounts = sum(1 for dates in survey_dates.values() if len(dates) > 1)
records = [r for r in records if r['visited'] == current_visit[r['account']]]
assert all(len({r['visited'] for r in records if r['account'] == a}) == 1 for a in current_visit), \
    'an account is still showing more than one survey pass'

# ---------- fix the "Sales" placeholder area via city majority-vote ----------
city_area_votes = defaultdict(Counter)
for r in records:
    if r['area'] and r['area'] != 'SALES':
        city_area_votes[r['city']][r['area']] += 1

unresolved_cities = set()
for r in records:
    if r['area'] == 'SALES':
        votes = city_area_votes.get(r['city'])
        if votes:
            r['area'] = votes.most_common(1)[0][0]
        elif r['city'] in CITY_AREA_FALLBACK:
            r['area'] = CITY_AREA_FALLBACK[r['city']]
        else:
            unresolved_cities.add(r['city'])
            r['area'] = 'UNASSIGNED'
if unresolved_cities:
    print(f"WARNING: could not resolve area for city/cities (left as UNASSIGNED): {sorted(unresolved_cities)}")

area_group = {}
for a in CORE_AREAS:
    area_group[a] = 'core'
for a in NON_FOCUS_AREAS:
    area_group[a] = 'nonFocus'


def group_of(area):
    return area_group.get(area, 'other')


# ---------- top-line summary ----------
# Per Kohler, 2026-07-30: the headline numbers (hero + brand breakdown) are
# scoped to the core market only, since that's what the manager actually
# cares about day to day -- company-wide totals are kept as a smaller
# reference figure (see companyWide below) rather than dropped, so nothing's
# hidden, just de-emphasized.
def summarize(rows):
    taps = sum(r['taps'] for r in rows)
    us = sum(r['taps'] for r in rows if r['status'] == 'US')
    them = sum(r['taps'] for r in rows if r['status'] == 'THEM')
    return {
        'totalTaps': taps, 'usTaps': us, 'themTaps': them, 'unverifiedTaps': taps - us - them,
        'usPct': pct(us, taps), 'themPct': pct(them, taps),
        'accountsSurveyed': len(set(r['account'] for r in rows)),
    }


core_records = [r for r in records if group_of(r['area']) == 'core']
summary = summarize(core_records)
company_wide = summarize(records)


# ---------- brand mix (top N + Other), each side -- core market only ----------
def brand_breakdown(status, top_n, rows):
    totals = defaultdict(int)
    for r in rows:
        if r['status'] == status:
            totals[r['brandFamily']] += r['taps']
    side_total = sum(totals.values())
    ranked = sorted(totals.items(), key=lambda kv: -kv[1])
    top = ranked[:top_n]
    other_taps = sum(t for _, t in ranked[top_n:])
    other_count = len(ranked) - len(top)
    return {
        'total': side_total,
        'top': [{'brand': b.title(), 'taps': t, 'pct': pct(t, side_total)} for b, t in top],
        'other': {'taps': other_taps, 'pct': pct(other_taps, side_total), 'brandCount': other_count},
    }


brands_us = brand_breakdown('US', TOP_N_BRANDS, core_records)
brands_them = brand_breakdown('THEM', TOP_N_BRANDS, core_records)


# ---------- full brand totals, core market (for the pie-chart brand picker) ----------
def all_brand_totals(status, rows):
    totals = defaultdict(int)
    for r in rows:
        if r['status'] == status:
            totals[r['brandFamily']] += r['taps']
    side_total = sum(totals.values())
    ranked = sorted(totals.items(), key=lambda kv: -kv[1])
    return [{'brand': b.title(), 'taps': t, 'pct': pct(t, side_total)} for b, t in ranked]


all_brands_us = all_brand_totals('US', core_records)
all_brands_them = all_brand_totals('THEM', core_records)


# ---------- by area ----------
def area_summary(area_name):
    rows = [r for r in records if r['area'] == area_name]
    taps = sum(r['taps'] for r in rows)
    us = sum(r['taps'] for r in rows if r['status'] == 'US')
    them = sum(r['taps'] for r in rows if r['status'] == 'THEM')
    return rows, {
        'area': area_name.title(), 'totalTaps': taps, 'usTaps': us, 'themTaps': them,
        'usPct': pct(us, taps), 'themPct': pct(them, taps),
    }


def brands_in(rows, status):
    """All brand families with any taps in these rows, ranked -- not just the
    top N -- so the frontend's brand customizer can look up ANY brand's
    count in this area on demand, not just the precomputed default set."""
    totals = defaultdict(int)
    for r in rows:
        if r['status'] == status:
            totals[r['brandFamily']] += r['taps']
    ranked = sorted(totals.items(), key=lambda kv: -kv[1])
    return [{'brand': b.title(), 'taps': t} for b, t in ranked]


core_areas_out = []
for a in CORE_AREAS:
    rows, s = area_summary(a)
    all_us_here = brands_in(rows, 'US')
    all_them_here = brands_in(rows, 'THEM')
    s['topUs'] = all_us_here[:TOP_N_AREA_BRANDS]
    s['topThem'] = all_them_here[:TOP_N_AREA_BRANDS]
    s['allUs'] = all_us_here
    s['allThem'] = all_them_here
    core_areas_out.append(s)

non_focus_areas_out = [area_summary(a)[1] for a in NON_FOCUS_AREAS]

other_area_names = sorted({r['area'] for r in records if group_of(r['area']) == 'other' and r['area'] != 'UNASSIGNED'})
other_areas_out = [area_summary(a)[1] for a in other_area_names]
unassigned_rows, unassigned_summary = area_summary('UNASSIGNED')
if unassigned_summary['totalTaps']:
    other_areas_out.append(unassigned_summary)


# ---------- brand lookup: any brand family's handle share, in every area ----------
# Combines US + THEM taps for the same brand-family text (a handful of brands,
# e.g. Blue Moon, show up on both sides at different accounts -- this answers
# "how much of this brand is out there in area X", not "how much do we
# specifically get credit for"). Covers every area, not just the core market,
# since this is a lookup tool, not the headline framing above.
area_total_taps = {}
area_group_label = {}
for grp, out_list in (('core', core_areas_out), ('nonFocus', non_focus_areas_out), ('other', other_areas_out)):
    for a in out_list:
        area_total_taps[a['area']] = a['totalTaps']
        area_group_label[a['area']] = grp

brand_area_taps = defaultdict(lambda: defaultdict(int))
brand_side_taps = defaultdict(lambda: {'US': 0, 'THEM': 0})
for r in records:
    if r['area'] == 'UNASSIGNED':
        continue
    area_title = r['area'].title()
    brand_area_taps[r['brandFamily']][area_title] += r['taps']
    if r['status'] in ('US', 'THEM'):
        brand_side_taps[r['brandFamily']][r['status']] += r['taps']

brand_lookup = []
for bf, area_map in brand_area_taps.items():
    sides = brand_side_taps[bf]
    side = 'US' if sides['US'] >= sides['THEM'] else 'THEM'
    total = sum(area_map.values())
    by_area = [
        {'area': area_title, 'group': area_group_label.get(area_title, 'other'), 'taps': taps,
         'areaTotal': area_total_taps.get(area_title, 0), 'pct': pct(taps, area_total_taps.get(area_title, 0))}
        for area_title, taps in sorted(area_map.items(), key=lambda kv: -kv[1])
    ]
    brand_lookup.append({'brand': bf.title(), 'side': side, 'totalTaps': total, 'byArea': by_area})
brand_lookup.sort(key=lambda b: -b['totalTaps'])


# ---------- velocity: join tap brands (US only) to Encompass units sold ----------
# Encompass only ever carries OUR sales, so velocity is only ever computable
# for our own (US) brand families -- there is no Encompass record of what a
# competitor sold through a handle we don't own.
crosswalk = {}
cw = wb['Brand Crosswalk']
cw_header = [c.value for c in cw[1]]
cw_idx = {h: i for i, h in enumerate(cw_header)}
for r in cw.iter_rows(min_row=2, values_only=True):
    rb = (r[cw_idx['Report Brand Family']] or '').strip().upper()
    mapped = r[cw_idx['Mapped Encompass Brand Family']]
    if rb:
        crosswalk[rb] = mapped.strip().upper() if mapped else None

units_rows = []
with open(UNITS_CSV, newline='', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    units_col = next(c for c in reader.fieldnames if c.startswith('Units'))
    for row in reader:
        units_rows.append({
            'account': row['Customer Num'].strip(),
            'brandFamily': row['Brand Family'].strip().upper(),
            'supplier': row['Supplier'].strip().upper(),
            'units': float(row[units_col] or 0),
        })

units_bf_set = {u['brandFamily'] for u in units_rows}
units_by_account_bf = defaultdict(float)
units_by_account_supplier = defaultdict(float)
for u in units_rows:
    units_by_account_bf[(u['account'], u['brandFamily'])] += u['units']
    units_by_account_supplier[(u['account'], u['supplier'])] += u['units']

tap_accounts = {r['account'] for r in records}
units_accounts = {u['account'] for u in units_rows}
matched_accounts = tap_accounts & units_accounts


def resolve_encompass_key(brand_family):
    """Returns ('bf', name) or ('supplier', name) or None."""
    if brand_family in units_bf_set:
        return ('bf', brand_family)
    mapped = crosswalk.get(brand_family)
    if mapped:
        if mapped in units_bf_set:
            return ('bf', mapped)
        mapped_upper = mapped.upper()
        if any(u['supplier'] == mapped_upper for u in units_rows):
            return ('supplier', mapped_upper)
    return None


# Company-wide (not core-market-only, same scope as the rest of this
# section) -- every US brand family that appears anywhere, so the velocity
# table's brand customizer can pick from more than just the default top 8.
company_wide_us_totals = defaultdict(int)
for r in records:
    if r['status'] == 'US':
        company_wide_us_totals[r['brandFamily']] += r['taps']
candidate_us_brands = sorted(company_wide_us_totals, key=lambda b: -company_wide_us_totals[b])

velocity_brands = []
velocity_unmatched = []
for brand_family in candidate_us_brands:
    key = resolve_encompass_key(brand_family)
    if key is None:
        if brand_family.title() in [e['brand'] for e in brands_us['top']]:
            velocity_unmatched.append(brand_family.title())
        continue
    kind, name = key
    matched_taps = sum(r['taps'] for r in records
                        if r['brandFamily'] == brand_family and r['status'] == 'US' and r['account'] in matched_accounts)
    if kind == 'bf':
        matched_units = sum(units_by_account_bf[(a, name)] for a in matched_accounts)
    else:
        matched_units = sum(units_by_account_supplier[(a, name)] for a in matched_accounts)
    velocity_brands.append({
        'brand': brand_family.title(),
        'matchedTaps': matched_taps,
        'unitsSold': round(matched_units, 1),
        'unitsPerTap': round(matched_units / matched_taps, 2) if matched_taps else None,
        # Which Encompass pool the units came from. Several tap-survey brand
        # families can resolve to the SAME pool ("Yuengling" and "Yuengling
        # Brewery" both -> bf:YUENGLING), in which case each alias is
        # credited the full pool's units against only its own taps -- the
        # small alias then shows an absurd units-per-handle. Consumers that
        # rank or compare velocity (the Full Detail quadrant cards) keep
        # only the most-surveyed alias per key; the raw table keeps every
        # row since the note already flags figures as directional.
        'encKey': f'{kind}:{name}',
    })
velocity_brands.sort(key=lambda b: -(b['unitsPerTap'] or 0))

resolved_brand_names = {b['brand'] for b in velocity_brands}
default_velocity_brands = [e['brand'] for e in brands_us['top'] if e['brand'] in resolved_brand_names]

velocity = {
    'accountsSurveyed': len(tap_accounts),
    'accountsMatched': len(matched_accounts),
    'brands': velocity_brands,
    'defaultBrands': default_velocity_brands,
    'unmatchedBrands': velocity_unmatched,
}

# ---------- per-brand segment attachment (Full Detail "Segment Battleground",
# and the Snapshot segment donut picks these up automatically -- index.html's
# classifySegment() was built to prefer a real `segment` field over its own
# heuristic from day one) ----------
resolve_segment = build_segment_resolver(wb)
bf_seg_votes = defaultdict(Counter)
for r in records:
    bucket = SEGMENT_NORMALIZE.get(resolve_segment(r['brand'], r['brandFamily']))
    if bucket:
        bf_seg_votes[r['brandFamily']][bucket] += r['taps']

brand_family_segment = {}
for bf, votes in bf_seg_votes.items():
    if bf in GENERIC_BRAND_NAMES:
        continue
    bucket, top_taps = votes.most_common(1)[0]
    # Require a clear tap-weighted majority before attributing a whole brand
    # family to one bucket; a family whose rows genuinely split stays with
    # the client-side heuristic rather than getting a coin-flip label.
    if top_taps / sum(votes.values()) >= 0.6:
        brand_family_segment[bf] = bucket

for lst in (brands_us['top'], brands_them['top'], all_brands_us, all_brands_them):
    for e in lst:
        seg = brand_family_segment.get(e['brand'].upper())
        if seg:
            e['segment'] = seg

# ---------- account-level control analysis (core market only) ----------
acct_map = {}
for r in core_records:
    a = acct_map.setdefault(r['account'], {
        'dba': r['dba'], 'area': r['area'].title(), 'city': r['city'].title(),
        'us': 0, 'them': 0, 'usBrands': Counter(), 'themBrands': Counter(),
    })
    if r['status'] == 'US':
        a['us'] += r['taps']
        a['usBrands'][r['brandFamily'].title()] += r['taps']
    elif r['status'] == 'THEM':
        a['them'] += r['taps']
        a['themBrands'][r['brandFamily'].title()] += r['taps']

accounts_list = [a for a in acct_map.values() if a['us'] + a['them'] > 0]

CONTROL_BANDS = [
    ('all_us', 'Fully ours (100%)', lambda p: p >= 100),
    ('lead', 'We lead (60–99%)', lambda p: 60 <= p < 100),
    ('contested', 'Contested (40–59%)', lambda p: 40 <= p < 60),
    ('trail', 'They lead (1–39%)', lambda p: 0 < p < 40),
    ('all_them', 'Fully theirs (0%)', lambda p: p <= 0),
]
bands_out = []
for key, label, test in CONTROL_BANDS:
    hit = [a for a in accounts_list if test(a['us'] / (a['us'] + a['them']) * 100)]
    bands_out.append({'key': key, 'label': label, 'accounts': len(hit),
                      'taps': sum(x['us'] + x['them'] for x in hit)})

sorted_by_taps = sorted(accounts_list, key=lambda a: -(a['us'] + a['them']))
core_total_taps = sum(a['us'] + a['them'] for a in accounts_list)
concentration = {
    'accounts': len(accounts_list),
    'top10': pct(sum(a['us'] + a['them'] for a in sorted_by_taps[:10]), core_total_taps),
    'top25': pct(sum(a['us'] + a['them'] for a in sorted_by_taps[:25]), core_total_taps),
    'top50': pct(sum(a['us'] + a['them'] for a in sorted_by_taps[:50]), core_total_taps),
}


def acct_row(a):
    total = a['us'] + a['them']
    return {
        'dba': a['dba'], 'area': a['area'], 'city': a['city'],
        'taps': total, 'usTaps': a['us'], 'themTaps': a['them'], 'usPct': pct(a['us'], total),
        # topThem deliberately KEEPS generic labels ("Other Supplier" = real
        # unidentified competitor handles, useful signal at a flip target);
        # topUs drops them -- "our biggest brands there" citing "Other Brand
        # Family" tells the reader nothing.
        'topThem': [{'brand': b, 'taps': t} for b, t in a['themBrands'].most_common(3)],
        'topUs': [{'brand': b, 'taps': t} for b, t in a['usBrands'].most_common(6)
                  if b.upper() not in GENERIC_BRAND_NAMES][:2],
    }


flip_targets = [acct_row(a) for a in sorted(accounts_list, key=lambda a: (-a['them'], -(a['us'] + a['them'])))[:20]]
# Anchors are accounts we CONTROL (>=60% ours), ranked by how many of our
# handles are at stake there -- not just our biggest raw presence, which
# would re-surface the same giant contested accounts as the flip-target
# list and blur the two ideas together.
anchors = [acct_row(a) for a in sorted(
    (a for a in accounts_list if a['us'] / (a['us'] + a['them']) >= 0.6),
    key=lambda a: (-a['us'], -(a['us'] + a['them'])))[:10]]

account_analysis = {'bands': bands_out, 'concentration': concentration,
                    'flipTargets': flip_targets, 'anchors': anchors}

# ---------- dual-sourced brands: brand families pouring on BOTH sides in the
# core market (e.g. Blue Moon at one account through us, at another through a
# competing distributor) -- the THEM handles here are volume of brands we
# demonstrably carry, going through someone else ----------
shared_stats = defaultdict(lambda: {'us': 0, 'them': 0, 'usA': set(), 'themA': set(), 'themAreas': Counter()})
for r in core_records:
    if r['brandFamily'] in GENERIC_BRAND_NAMES:
        continue
    s = shared_stats[r['brandFamily']]
    if r['status'] == 'US':
        s['us'] += r['taps']
        s['usA'].add(r['account'])
    elif r['status'] == 'THEM':
        s['them'] += r['taps']
        s['themA'].add(r['account'])
        s['themAreas'][r['area'].title()] += r['taps']

shared_brands = sorted(
    ({'brand': bf.title(), 'usTaps': s['us'], 'themTaps': s['them'],
      'usAccounts': len(s['usA']), 'themAccounts': len(s['themA']),
      'topThemAreas': [a for a, _ in s['themAreas'].most_common(2)]}
     for bf, s in shared_stats.items() if s['us'] >= 2 and s['them'] >= 2),
    key=lambda x: -x['themTaps'])[:15]

payload = {
    'generatedAt': datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'),
    'summary': summary,
    'companyWide': company_wide,
    'brandsUs': brands_us,
    'brandsThem': brands_them,
    'allBrandsUs': all_brands_us,
    'allBrandsThem': all_brands_them,
    'brandLookup': brand_lookup,
    'areas': {'core': core_areas_out, 'nonFocus': non_focus_areas_out, 'other': other_areas_out},
    'velocity': velocity,
    'accountAnalysis': account_analysis,
    'sharedBrands': shared_brands,
}

data_json = json.dumps(payload, separators=(',', ':'))
html = HTML.read_text(encoding='utf-8')
new_html, n = re.subn(
    r'(<script id="exec-data" type="application/json">).*?(</script>)',
    lambda m: m.group(1) + data_json + m.group(2),
    html, count=1, flags=re.S,
)
assert n == 1, 'exec-data script tag not found in index.html'
HTML.write_text(new_html, encoding='utf-8')

print(f"CORE MARKET: {summary['totalTaps']} taps at {summary['accountsSurveyed']} accounts: "
      f"{summary['usTaps']} ours ({summary['usPct']}%) / {summary['themTaps']} competitor ({summary['themPct']}%)")
print(f"  by area: {', '.join(a['area'] + ' ' + str(a['totalTaps']) for a in core_areas_out)}")
print(f"Current lineup only: {resurveyed_accounts} account(s) re-surveyed, "
      f"{superseded_taps} superseded taps excluded company-wide")
print(f"Company-wide (reference only): {company_wide['totalTaps']} taps, "
      f"{company_wide['usPct']}% ours / {company_wide['themPct']}% competitor")
print(f"Velocity: {len(velocity_brands)} US brand(s) matched to Encompass out of {len(candidate_us_brands)} candidates "
      f"({len(default_velocity_brands)}/{TOP_N_BRANDS} of the default top brands resolved); "
      f"{len(matched_accounts)}/{len(tap_accounts)} accounts have a units-sold record")
if velocity_unmatched:
    print(f"  No Encompass mapping found for: {velocity_unmatched}")
seg_covered = sum(e['taps'] for e in all_brands_us + all_brands_them if e.get('segment'))
seg_total = sum(e['taps'] for e in all_brands_us + all_brands_them)
print(f"Deep dive: workbook segments attached to {len(brand_family_segment)} brand families "
      f"({pct(seg_covered, seg_total)}% of core handles; the rest use index.html's heuristic); "
      f"{len(shared_brands)} dual-sourced brands; "
      f"top flip target: {flip_targets[0]['dba']} ({flip_targets[0]['themTaps']} competitor handles)")
