#!/usr/bin/env python3
"""
Rebuilds DisplayPhotoReport.csv and the embedded <script id="da-data">
JSON in index.html from a fresh iSellBeer "Report" export (Report_NN.xlsx).

Reconstructed methodology (the source export doesn't document this — it
was reverse-engineered from the previously-committed data and verified
to reproduce identical aggregate totals):

  - One "display" = one photo submission = rows sharing the same
    (Photo Taker, Account #, Date/Time). If a photo logs multiple SKUs,
    their Quantity values are summed into that display's case count.
  - A row's canonical brand is its "Brand Family" column when non-blank,
    else its raw "Brand" column (a handful of brands — Carbliss, Monaco,
    Monaco Cocktails, Sinless RTD, Sun Cruisers RTD — have no Brand
    Family populated in the source export).
  - Each display's canonical brand(s) are usually all Priority or all All
    Other. When a display mixes both (first seen 2026-08-06, Report_34),
    it's classified by whichever side has more cases, then scored on the
    display's TOTAL cases against that classification's tier table --
    confirmed with the user (see the cases_by_class block in main()).
  - Tier is by total cases: 0 (<10, non-qualifying), 1 (10-19),
    2 (20-39), 3 (40-69), 4 (70+).
  - Points = TIER_POINTS[classification][tier]. allother/tier1 first
    appeared 2026-08-05 (Report_33) — confirmed with the user as 100 pts
    (half of allother/tier2), so it's no longer a gap in TIER_POINTS.
  - Sales Reps and Sales Associates earn on the same point scale.
  - Within a person's display list: sorted by points desc, then cases desc
    (not by date — a date-sort was tried once and explicitly reverted).

Run: python3 generate.py Report_NN.xlsx            (full-period export)
     python3 generate.py Report_NN.xlsx --merge    (partial export: rebuild
       only from the export's earliest row onward, carrying everything
       before that over from what's already published -- see main())
"""
import csv
import datetime
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
OUT_CSV = HERE / "DisplayPhotoReport.csv"
OUT_HTML = HERE / "index.html"

PRIORITY_BRANDS = {
    'BLUE MOON', 'CARBLISS', 'CAYMAN JACK', 'COORS', 'COORS LIGHT', 'CORONA', 'CORONA EXTRA',
    'CORONA NON-ALCOHOLIC', 'DOS EQUIS', 'FEVER TREE', 'GARAGE BEER - CONTRACT BREWING',
    'DOGFISH HEAD BREWERY', 'HEINEKEN', "LEINENKUGEL'S", 'MILLER LITE', 'MODELO CERVEZA', 'MODELO CHELADA', 'MODELO ORO', 'MONACO',
    'MONACO COCKTAILS', 'PACIFICO', 'PERONI NASTRO AZZURRO', 'RED STRIPE', 'SAM ADAMS SEASONAL',
    'SAMUEL ADAMS', 'SAPPORO', 'SINLESS RTD', 'SINLESS SPIRITS', 'SUN CRUISER ICED TEA VODKA',
    'SUN CRUISER LEMONADE', 'SUN CRUISERS RTD', 'TRULY', 'TRULY HARD SELTZER',
    'TRULY SEASONAL HARD SELTZER', 'TWISTED TEA', 'WHITE CLAW CLAWTAILS',
    'WHITE CLAW ORIGINAL HARD SELTZER', 'WHITE CLAW SURF HARD SELTZER',
    'WHITE CLAW SURGE HARD SELTZER', 'WHITE CLAW VODKA + SODA', 'YUENGLING BREWERY',
    # Confirmed with the user 2026-07-30.
    'WHITE CLAW ZERO', 'SIMPLY SPIKED LEMONADE',
    # Confirmed with the user 2026-08-06.
    'LYTT',
    # Confirmed with the user 2026-08-11: CERVEZA VICTORIA is a Constellation
    # Brands portfolio-mate of CORONA/MODELO/PACIFICO (same supplier), all
    # already Priority. MONTAUK BREWING COMPANY confirmed Priority same day
    # (despite being a craft brewery like several All Other entries below --
    # explicitly asked, don't re-derive from the craft-brewery pattern).
    'CERVEZA VICTORIA', 'MONTAUK BREWING COMPANY',
    # Confirmed with the user 2026-08-17 (Report_39): ANGRY ORCHARD joins its
    # Boston Beer portfolio-mates (SAMUEL ADAMS/TRULY/TWISTED TEA/SUN CRUISER/
    # DOGFISH HEAD/SINLESS/LYTT), all already Priority. KEYSTONE BEER was
    # explicitly asked rather than inferred from its supplier -- Molson Coors
    # covers both Priority brands (COORS/MILLER LITE/BLUE MOON) and an All
    # Other one (REDD'S), so supplier alone doesn't settle a Molson Coors
    # brand's tier.
    'ANGRY ORCHARD', 'KEYSTONE BEER',
}
ALLOTHER_BRANDS = {
    'ATHLETIC BREWING COMPANY', 'CARIB BREWERY', 'DELTA THC SELTZER', 'FAMOSA', 'KIRIN',
    'KIRIN ICHIBAN', 'POPSICLE FMB', 'SIERRA NEVADA BREWING COMPANY',
    # Confirmed with the user 2026-07-30.
    'TALKHOUSE ENCORE TEQUILA SODA', 'TALKHOUSE ENCORE VARIETY PACK',
    # Confirmed with the user 2026-08-04. NOCA appears under both its Brand
    # Family value ("NOCA BEVERAGES") and, when Brand Family is blank, its
    # raw Brand value ("NOCA") -- both need to be listed.
    'NOCA', 'NOCA BEVERAGES', 'SARATOGA WATER',
    # Confirmed with the user 2026-08-05.
    "REDD'S", 'VICTORY BREWING COMPANY',
    # Confirmed with the user 2026-08-06.
    'HACKER-PSCHORR', 'HOFBRAU', 'PAULANER',
    # Confirmed with the user 2026-08-11.
    'SOUTHERN TIER BREWING COMPANY',
    # Confirmed with the user 2026-08-17 (Report_39). First brand from supplier
    # VIVA BEVERAGES, so there was no portfolio precedent either way.
    'VIVA TEQUILA SELTZER',
    # Confirmed with the user 2026-08-18 (Report_3 export). Oktoberfest/fall
    # seasonals from one Bottle King Wayne submission; CAPE MAY appears under
    # two Brand Family values ("CAPE MAY" and "CAPE MAY SEASONALS").
    'CAPE MAY', 'CAPE MAY SEASONALS', 'SHINER', 'STEVENS POINT BREWERY',
    'WEIHENSTEPHANER',
    # Confirmed with the user 2026-08-19 (Report_40). Fall seasonals from one
    # John O'Donoghue submission, same pattern as the 8-18 Oktoberfest batch
    # above (WHOLE HOG is a Stevens Point Brewery label).
    'FLYING DOG SEASONAL RELEASE', 'LONG TRAIL BREWING CO', 'SARANAC BREWERY',
    'WHOLE HOG',
}
# Brand Family aliases -- iSellBeer sometimes tags the same product with an
# inconsistent Brand Family value (e.g. the contract brewer's name instead of
# the beer's own brand family). Confirmed with the user 2026-07-20: BRAXTON
# rows are Garage Beer (Supplier "GARAGE BEER CO.", Brand "GARAGE BEER
# LAGER") mislabeled with the contract brewer's name as Brand Family.
# Confirmed with the user 2026-07-22: "GARAGE BEER" and "SAM ADAMS" are
# shorter Brand Family labels for the same already-classified brands.
# Confirmed with the user 2026-07-23: "YUENGLING" is the same shorter-label
# pattern for "YUENGLING BREWERY".
# Confirmed with the user 2026-07-30: "ATHLETIC BREWING CO" and "CARBLISS
# COCKTAILS" are the same shorter-label pattern for "ATHLETIC BREWING
# COMPANY" and "CARBLISS" respectively.
# Confirmed with the user 2026-08-11: "HACKER-PSCHORR BREWERY" is the same
# longer-label pattern for "HACKER-PSCHORR" (already All Other).
BRAND_FAMILY_ALIASES = {
    'BRAXTON': 'GARAGE BEER - CONTRACT BREWING',
    'GARAGE BEER': 'GARAGE BEER - CONTRACT BREWING',
    'SAM ADAMS': 'SAMUEL ADAMS',
    'YUENGLING': 'YUENGLING BREWERY',
    'ATHLETIC BREWING CO': 'ATHLETIC BREWING COMPANY',
    'CARBLISS COCKTAILS': 'CARBLISS',
    'HACKER-PSCHORR BREWERY': 'HACKER-PSCHORR',
}
TIER_POINTS = {
    'priority': {1: 200, 2: 300, 3: 500, 4: 1000},
    # allother/tier1 confirmed with the user 2026-08-05 (first time this
    # combination has appeared in any export) -- half of allother/tier2,
    # matching the roughly 2x step-up pattern between the other tiers.
    'allother': {1: 100, 2: 200, 3: 300, 4: 600},
}
COLS = ['num', 'taker', 'role', 'dm', 'acct', 'dba', 'address', 'city',
        'supplier', 'brand_family', 'brand', 'sku', 'qty', 'dt', 'photo']


def tier_for(cases):
    if cases >= 70:
        return 4
    if cases >= 40:
        return 3
    if cases >= 20:
        return 2
    if cases >= 10:
        return 1
    return 0


def canonical_brand(row):
    bf = (row['brand_family'] or '').strip()
    brand = bf if bf else (row['brand'] or '').strip()
    return BRAND_FAMILY_ALIASES.get(brand, brand)


def classify(brand):
    if brand in PRIORITY_BRANDS:
        return 'priority'
    if brand in ALLOTHER_BRANDS:
        return 'allother'
    return None


def parse_dt(s):
    """'MM/DD/YYYY HH:MM AM/PM' -- the only Date/Time format iSellBeer emits."""
    return datetime.datetime.strptime(str(s).strip(), "%m/%d/%Y %I:%M %p")


def read_rows(src):
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Report"]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {col: ws.cell(row=r, column=i + 1).value for i, col in enumerate(COLS)}
        if not row['taker']:
            continue
        photo_cell = ws.cell(row=r, column=15)
        row['photo_url'] = photo_cell.hyperlink.target if photo_cell.hyperlink else None
        rows.append(row)
    return rows


def build_displays(rows):
    """Source rows -> one display dict per photo submission."""
    groups = defaultdict(list)
    for row in rows:
        groups[(row['taker'], row['acct'], row['dt'])].append(row)

    displays = []
    for (taker, acct, dt), grp in groups.items():
        cases = sum(g['qty'] for g in grp)
        brands = sorted({canonical_brand(g) for g in grp})
        classes = {classify(b) for b in brands}
        if len(classes) > 1:
            # First seen 2026-08-06 (Report_34): a display mixing Priority and
            # All Other brands. Confirmed with the user: classify the whole
            # display by whichever side has more cases, then score the
            # display's TOTAL cases against that classification's tier table.
            cases_by_class = defaultdict(int)
            for g in grp:
                cases_by_class[classify(canonical_brand(g))] += g['qty']
            classification = max(cases_by_class, key=cases_by_class.get)
        else:
            classification = classes.pop()
        t = tier_for(cases)
        points = TIER_POINTS.get(classification, {}).get(t, 0) if t else 0
        photos = sorted({g['photo_url'] for g in grp if g['photo_url']})
        first = grp[0]
        displays.append({
            "taker": taker, "role": first['role'], "acct": acct, "dba": first['dba'],
            "city": first['city'], "dt": dt, "cases": cases, "classification": classification,
            "tier": t, "points": points, "brands": brands, "photos": photos,
        })
    return displays


def assemble(displays):
    """Flat display list -> the {meta, people} payload index.html embeds."""
    by_person = defaultdict(list)
    for d in displays:
        by_person[d['taker']].append(d)

    people_out = []
    for name, ds in by_person.items():
        # Sorted by points desc, then cases desc, then date/time -- NOT
        # chronologically (a date-sort was tried once and explicitly reverted).
        ds = sorted(ds, key=lambda d: (-d['points'], -d['cases'], d['dt']))
        qualifying = [d for d in ds if d['tier'] >= 1]
        people_out.append({
            "name": name,
            # A person's role is constant in practice; take it from their most
            # recent display so a merge can't resurrect a stale one.
            "role": max(ds, key=lambda d: parse_dt(d['dt']))['role'],
            "points": sum(d['points'] for d in ds),
            "qualifying": len(qualifying),
            "total": len(ds),
            "priorityQualifying": len([d for d in qualifying if d['classification'] == 'priority']),
            "otherQualifying": len([d for d in qualifying if d['classification'] == 'allother']),
            "displays": ds,
        })
    people_out.sort(key=lambda p: -p['points'])

    dates = [parse_dt(d['dt']) for d in displays]
    return {
        "meta": {
            "startDate": min(dates).strftime("%m/%d/%Y"),
            "endDate": max(dates).strftime("%m/%d/%Y"),
            "totalDisplays": len(displays),
            "totalPoints": sum(d['points'] for d in displays),
        },
        "people": people_out,
    }


def embedded_data(html):
    m = re.search(r'<script id="da-data" type="application/json">(.*?)</script>', html, re.DOTALL)
    if not m:
        raise SystemExit("Could not find da-data script tag in index.html")
    return json.loads(m.group(1))


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('-')]
    flags = [a for a in sys.argv[1:] if a.startswith('-')]
    if len(args) != 1 or set(flags) - {'--merge'}:
        raise SystemExit("Usage: python3 generate.py Report_NN.xlsx [--merge]")
    src = Path(args[0])
    merge = '--merge' in flags

    rows = read_rows(src)
    if not rows:
        raise SystemExit(f"No data rows found in {src.name}")

    unknown = {canonical_brand(r) for r in rows} - PRIORITY_BRANDS - ALLOTHER_BRANDS
    if unknown:
        raise SystemExit(
            f"Unclassified brand(s) found — add to PRIORITY_BRANDS or ALLOTHER_BRANDS "
            f"in this script after confirming with the user: {sorted(unknown)}"
        )

    # --merge (added 2026-08-20): the export covers only part of the tracked
    # period -- everything from its EARLIEST row onward is rebuilt from it, and
    # anything before that is carried over verbatim from what's already
    # published (displays out of index.html's embedded JSON, which keeps their
    # photo links; source rows out of DisplayPhotoReport.csv). Use it when the
    # user pulls a partial export; a full-period export needs no flag.
    # Carried-over displays keep the points they were BUILT with -- if a brand
    # has been reclassified since, only the rebuilt window reflects that.
    preserved_displays, preserved_csv = [], []
    if merge:
        cutoff = min(parse_dt(r['dt']) for r in rows).replace(hour=0, minute=0)
        old = embedded_data(OUT_HTML.read_text())
        old_displays = [d for p in old['people'] for d in p['displays']]
        preserved_displays = [d for d in old_displays if parse_dt(d['dt']) < cutoff]
        replaced = len(old_displays) - len(preserved_displays)
        with open(OUT_CSV, newline="") as f:
            preserved_csv = [r for r in list(csv.reader(f))[1:] if parse_dt(r[13]) < cutoff]
        print(f"Merge: export starts {cutoff:%m/%d/%Y} -- carrying over "
              f"{len(preserved_displays)} displays / {len(preserved_csv)} source rows "
              f"before it, rebuilding the {replaced} display(s) from then on.")

    # write the plain CSV export (human-diffable, no hyperlinks)
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(['#', 'Photo Taker', "Photo Taker's Role", 'District Manager', 'Account #',
                    'DBA', 'Address', 'City', 'Supplier', 'Brand Family', 'Brand', 'SKU',
                    'Quantity', 'Date/Time', 'Photo'])
        w.writerows(preserved_csv)
        for row in rows:
            w.writerow([row['num'], row['taker'], row['role'], row['dm'], row['acct'],
                        row['dba'], row['address'], row['city'], row['supplier'],
                        row['brand_family'] or '', row['brand'], row['sku'], row['qty'],
                        row['dt'], 'Photo'])

    data = assemble(preserved_displays + build_displays(rows))

    html = OUT_HTML.read_text()
    new_html = re.sub(
        r'(<script id="da-data" type="application/json">).*?(</script>)',
        lambda m: m.group(1) + json.dumps(data) + m.group(2),
        html,
        flags=re.DOTALL,
    )
    OUT_HTML.write_text(new_html)
    qualifying = sum(1 for p in data['people'] for d in p['displays'] if d['tier'] >= 1)
    print(f"Wrote {data['meta']['totalDisplays']} displays ({qualifying} qualifying), "
          f"{data['meta']['totalPoints']} total points across {len(data['people'])} people.")
    print(f"Date range: {data['meta']['startDate']} to {data['meta']['endDate']}")


if __name__ == "__main__":
    main()
