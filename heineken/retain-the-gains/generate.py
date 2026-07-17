#!/usr/bin/env python3
"""
Rebuilds the embedded DATA in index.html from HUSA's official retention
exports (source of truth) plus Kohler's own customer roster (for rep
attribution, since HUSA's "Sales Person" column is not always current).

Inputs:
  HUSA_Retained_PODS_ON.xlsx   'ON  PREMISE' sheet, 5 SKU-family columns
  HUSA_Retained_PODS_OFF.xlsx  'OFF PREMISE' sheet, 16 individual-SKU columns
  Kohler_Customer_Data.csv     Kohler's Encompass customer/rep roster

Cell code convention in the SKU columns — confirmed directly with a
HUSA rep on 2026-07-17 (do not re-derive this from the PDF chart
numbers alone; an earlier guess based purely on numeric pattern-matching
had codes 2 and 3 backwards in spirit, even though the raw counts
matched):
  1 = Unretained        had this SKU at some point, doesn't currently.
                         A generic, longstanding gap.
  2 = Open Opportunity  never filled in Mar-Apr AND still not filled in
                         Jun-Aug. NOT a loss — just untouched whitespace
                         that was never won in the first place.
  3 = Lost New Gain     WAS gained as a new placement in Mar-Apr, but
                         has not been (re-)gained in Jun-Aug — i.e. also
                         a loss, just a higher-priority one since it's a
                         recently-won POD that slipped, not a
                         longstanding gap.
  blank = not applicable (outlet was never eligible/distributed for that SKU)

There is NO code in this file representing "currently, successfully
held" — every non-blank cell is some flavor of gap or open whitespace.
The column is literally headered "UnRetained" for exactly this reason:
this file is a problem list, not a status report. Do not build a
"retention rate" % metric from it — there's no positive state to rate
against. "Total gaps to close" = count of codes 1 + 3. Code 2 is a
separate "open opportunity" list, not part of the gap count.

Rep attribution: HUSA identifies outlets by VIP ID / TDLinx / Dist SAP;
Kohler's own system (Encompass) uses its own Customer Num. There's no
shared ID, so outlets are matched to Kohler's roster by normalized
address + city, falling back to a fuzzy address match within the same
city. Outlets that don't match anything in the roster are kept (grouped
under "Unmatched — Needs Review") rather than silently dropped or
mis-attributed to HUSA's possibly-stale Sales Person field.
"""
import csv
import json
import re
import difflib
from collections import defaultdict, Counter
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
ON_XLSX = HERE / "HUSA_Retained_PODS_ON.xlsx"
OFF_XLSX = HERE / "HUSA_Retained_PODS_OFF.xlsx"
CUSTOMER_CSV = HERE / "Kohler_Customer_Data.csv"
CURRENT_ON_XLSX = HERE / "HUSA_CurrentPOD_ON.xlsx"
CURRENT_OFF_XLSX = HERE / "HUSA_CurrentPOD_OFF.xlsx"
OUT_HTML = HERE / "index.html"

# HUSA's own goal figures from the T2 "Retain the Gains" slide. Not
# derivable from any export we have — these are fixed distributor-level
# targets Heineken set independently, so they're hardcoded here the same
# way SUPPLIER_BUDGETS is hardcoded in supplier-budget/generate.py.
EP_DB_GOAL = {"Off Premise": 794, "On Premise": 873}

LOST_NEW_GAIN, OPEN_OPPORTUNITY, UNRETAINED = 3, 2, 1

OUTLET_COLS = {
    "vip_id": 1, "tdlinx": 2, "dist_sap": 3, "outlet_name": 4,
    "address": 5, "city": 6, "state": 7, "abc_outlet": 8,
    "county": 9, "husa_sales_person": 10, "route": 11, "corp_eligible": 12,
}


def norm_addr(s):
    s = (s or "").upper().strip()
    s = re.sub(r"[.,#]", "", s)
    repl = {
        r"\bSTREET\b": "ST", r"\bAVENUE\b": "AVE", r"\bROAD\b": "RD",
        r"\bBOULEVARD\b": "BLVD", r"\bDRIVE\b": "DR", r"\bLANE\b": "LN",
        r"\bPLACE\b": "PL", r"\bCOURT\b": "CT", r"\bHIGHWAY\b": "HWY",
        r"\bSUITE\b": "STE", r"\bNORTH\b": "N", r"\bSOUTH\b": "S",
        r"\bEAST\b": "E", r"\bWEST\b": "W",
    }
    for pat, rep in repl.items():
        s = re.sub(pat, rep, s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def norm_city(s):
    return re.sub(r"\s+", " ", (s or "").upper().strip())


def load_customer_roster(path):
    """Returns list of dicts + lookup indexes for matching."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append({
                "rep": r["Sales Rep Assigned"].strip(),
                "channel": r["On-Off Premise"].strip(),
                "customer_num": r["Customer Num"].strip(),
                "name": r["Customer Name"].strip(),
                "address": r["Shipping Address"].strip(),
                "city": r["City"].strip(),
                "dist_area": r["Distribution Area"].strip(),
                "norm_addr": norm_addr(r["Shipping Address"]),
                "norm_city": norm_city(r["City"]),
            })
    by_addr_city = {}
    by_city = defaultdict(list)
    for r in rows:
        by_addr_city.setdefault((r["norm_addr"], r["norm_city"]), r)
        by_city[r["norm_city"]].append(r)
    return rows, by_addr_city, by_city


def match_rep(outlet_addr, outlet_city, by_addr_city, by_city):
    na, nc = norm_addr(outlet_addr), norm_city(outlet_city)
    hit = by_addr_city.get((na, nc))
    if hit:
        return hit["rep"], "exact"
    candidates = by_city.get(nc, [])
    if candidates:
        best, best_ratio = None, 0.0
        for c in candidates:
            ratio = difflib.SequenceMatcher(None, na, c["norm_addr"]).ratio()
            if ratio > best_ratio:
                best, best_ratio = c, ratio
        if best_ratio >= 0.82:
            return best["rep"], "fuzzy"
    return None, "unmatched"


def parse_sheet(xlsx_path, sheet_name, sku_cols, channel):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    outlets = []
    for row in range(5, ws.max_row + 1):
        outlet_name = ws.cell(row=row, column=OUTLET_COLS["outlet_name"]).value
        if not outlet_name or not str(outlet_name).strip():
            continue
        rec = {k: ws.cell(row=row, column=c).value for k, c in OUTLET_COLS.items()}
        rec["channel"] = channel
        skus = {}
        for col, sku_name in sku_cols.items():
            v = ws.cell(row=row, column=col).value
            if v in (LOST_NEW_GAIN, OPEN_OPPORTUNITY, UNRETAINED):
                skus[sku_name] = v
        if skus:
            rec["skus"] = skus
            outlets.append(rec)
    return outlets


def parse_current_pod_on(path):
    """HUSA_CurrentPOD_ON.xlsx: simple current-state POD flags (0/1) per
    SKU family, with a per-outlet Total column — this is what sums to
    HUSA's "DB's" (On Premise) figure."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Export"]
    total_col = next(c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value == "Total")
    outlets = []
    for row in range(3, ws.max_row + 1):
        outlet_name = ws.cell(row=row, column=6).value
        if not outlet_name or not str(outlet_name).strip():
            continue
        total = ws.cell(row=row, column=total_col).value
        if total is None:
            continue
        rep = (ws.cell(row=row, column=11).value or "").strip().title()
        outlets.append({"outlet": outlet_name, "rep": rep or "Unmatched — Needs Review", "total": total})
    return outlets


def parse_current_pod_off(path):
    """HUSA_CurrentPOD_OFF.xlsx: same idea as the ON file, but split by
    outlet Format (Large/Small) with a separate Total column per format.
    Confirmed against HUSA's own PDF that "EP's" (Off Premise) sums only
    the Large-format Total column (650 vs. HUSA's reported 645) — Large +
    Small combined overshoots (657)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Export"]
    fmt_by_col = {}
    current_fmt = None
    for c in range(1, ws.max_column + 1):
        label = ws.cell(row=1, column=c).value
        if label:
            current_fmt = label
        fmt_by_col[c] = current_fmt
    large_total_col = next(
        c for c in range(1, ws.max_column + 1)
        if ws.cell(row=2, column=c).value == "Total" and fmt_by_col[c] == "Large"
    )
    outlets = []
    for row in range(4, ws.max_row + 1):
        outlet_name = ws.cell(row=row, column=5).value
        if not outlet_name or not str(outlet_name).strip():
            continue
        total = ws.cell(row=row, column=large_total_col).value
        if total is None:
            continue
        rep = (ws.cell(row=row, column=10).value or "").strip().title()
        outlets.append({"outlet": outlet_name, "rep": rep or "Unmatched — Needs Review", "total": total})
    return outlets


def current_pod_summary(outlets, channel):
    goal = EP_DB_GOAL[channel]
    current = sum(o["total"] for o in outlets)
    diff = goal - current
    pct = round(100 * current / goal, 1) if goal else None
    by_rep = defaultdict(int)
    for o in outlets:
        by_rep[o["rep"]] += o["total"]
    rep_rows = [{"rep": rep, "total": total} for rep, total in by_rep.items()]
    rep_rows.sort(key=lambda r: (r["rep"] == "Unmatched — Needs Review", -r["total"]))
    return {"goal": goal, "current": current, "diff": diff, "pct": pct, "byRep": rep_rows}


def main():
    wb_on = openpyxl.load_workbook(ON_XLSX, data_only=True)
    ws_on = wb_on["ON  PREMISE"]
    on_sku_cols = {c: ws_on.cell(row=3, column=c).value for c in range(13, 18)}

    wb_off = openpyxl.load_workbook(OFF_XLSX, data_only=True)
    ws_off = wb_off["OFF PREMISE"]
    off_sku_cols = {c: ws_off.cell(row=3, column=c).value for c in range(13, 29)}

    on_outlets = parse_sheet(ON_XLSX, "ON  PREMISE", on_sku_cols, "On Premise")
    off_outlets = parse_sheet(OFF_XLSX, "OFF PREMISE", off_sku_cols, "Off Premise")
    all_outlets = on_outlets + off_outlets

    _, by_addr_city, by_city = load_customer_roster(CUSTOMER_CSV)

    match_stats = Counter()
    rep_discrepancies = []
    for o in all_outlets:
        roster_rep, method = match_rep(o["address"], o["city"], by_addr_city, by_city)
        husa_rep = (o["husa_sales_person"] or "").strip().title()
        if roster_rep:
            o["rep"] = roster_rep
            o["match_method"] = method
            if husa_rep and husa_rep.upper() != roster_rep.strip().upper():
                rep_discrepancies.append({
                    "outlet": o["outlet_name"], "channel": o["channel"],
                    "city": o["city"], "husaRep": husa_rep, "rosterRep": roster_rep,
                })
        elif husa_rep:
            o["rep"] = husa_rep
            o["match_method"] = "husa_fallback"
        else:
            o["rep"] = "Unmatched — Needs Review"
            o["match_method"] = "unmatched"
        match_stats[o["match_method"]] += 1

    print(f"Rep match results: {dict(match_stats)} (of {len(all_outlets)} outlets)")
    print(f"Rep assignment discrepancies (HUSA vs. Kohler roster): {len(rep_discrepancies)}")

    # ---- rollups ----
    # by SKU (within channel)
    sku_totals = defaultdict(lambda: Counter())  # (channel, sku) -> Counter
    rep_totals = defaultdict(lambda: Counter())  # rep -> Counter over unretained/lostNewGain/openOpportunity
    rep_sku = defaultdict(lambda: defaultdict(lambda: Counter()))  # rep -> sku -> Counter
    gap_list = []  # outlet-level rows still a gap (Unretained or Lost New Gain), for the close-the-gap view
    opportunity_list = []  # outlet-level rows that are Open Opportunities (never filled), separate from gaps

    for o in all_outlets:
        for sku, code in o["skus"].items():
            key = (o["channel"], sku)
            sku_totals[key][code] += 1
            rep_totals[o["rep"]][code] += 1
            rep_sku[o["rep"]][(o["channel"], sku)][code] += 1
            row = {
                "rep": o["rep"],
                "channel": o["channel"],
                "outlet": o["outlet_name"],
                "address": o["address"],
                "city": o["city"],
                "county": o["county"],
                "sku": sku,
                "husaRep": o["husa_sales_person"],
                "matchMethod": o["match_method"],
            }
            if code in (UNRETAINED, LOST_NEW_GAIN):
                row["gapType"] = "unretained" if code == UNRETAINED else "lostNewGain"
                gap_list.append(row)
            elif code == OPEN_OPPORTUNITY:
                opportunity_list.append(row)

    def code_summary(counter):
        u = counter.get(UNRETAINED, 0)
        l = counter.get(LOST_NEW_GAIN, 0)
        o = counter.get(OPEN_OPPORTUNITY, 0)
        return {"unretained": u, "lostNewGain": l, "totalGaps": u + l, "openOpportunity": o}

    sku_summary = []
    for (channel, sku), counter in sku_totals.items():
        s = code_summary(counter)
        s.update({"channel": channel, "sku": sku})
        sku_summary.append(s)
    sku_summary.sort(key=lambda s: (s["channel"], -s["totalGaps"]))

    rep_summary = []
    for rep, counter in rep_totals.items():
        s = code_summary(counter)
        s["rep"] = rep
        rep_summary.append(s)
    rep_summary.sort(key=lambda s: (s["rep"] == "Unmatched — Needs Review", -s["totalGaps"]))

    rep_sku_summary = defaultdict(list)
    for rep, skus in rep_sku.items():
        for (channel, sku), counter in skus.items():
            s = code_summary(counter)
            s.update({"channel": channel, "sku": sku})
            rep_sku_summary[rep].append(s)
        rep_sku_summary[rep].sort(key=lambda s: (s["channel"], -s["totalGaps"]))

    def rollup(rows):
        return code_summary(Counter({
            UNRETAINED: sum(s["unretained"] for s in rows),
            LOST_NEW_GAIN: sum(s["lostNewGain"] for s in rows),
            OPEN_OPPORTUNITY: sum(s["openOpportunity"] for s in rows),
        }))

    overall = rollup(sku_summary)
    overall_by_channel = {
        channel: rollup([s for s in sku_summary if s["channel"] == channel])
        for channel in ("On Premise", "Off Premise")
    }

    gap_list.sort(key=lambda g: (g["rep"] == "Unmatched — Needs Review", g["rep"], g["channel"], g["sku"]))
    opportunity_list.sort(key=lambda g: (g["rep"] == "Unmatched — Needs Review", g["rep"], g["channel"], g["sku"]))

    current_on = parse_current_pod_on(CURRENT_ON_XLSX)
    current_off = parse_current_pod_off(CURRENT_OFF_XLSX)
    ep_db_summary = {
        "Off Premise": current_pod_summary(current_off, "Off Premise"),
        "On Premise": current_pod_summary(current_on, "On Premise"),
    }
    print(f"EP's (Off Premise, Large format): {ep_db_summary['Off Premise']['current']} vs. goal {EP_DB_GOAL['Off Premise']}")
    print(f"DB's (On Premise): {ep_db_summary['On Premise']['current']} vs. goal {EP_DB_GOAL['On Premise']}")

    data = {
        "overall": overall,
        "overallByChannel": overall_by_channel,
        "skuSummary": sku_summary,
        "repSummary": rep_summary,
        "repSkuSummary": rep_sku_summary,
        "gapList": gap_list,
        "opportunityList": opportunity_list,
        "matchStats": dict(match_stats),
        "outletCount": len(all_outlets),
        "repDiscrepancies": rep_discrepancies,
        "epDbSummary": ep_db_summary,
    }

    html = OUT_HTML.read_text()
    tag_open = '<script id="retain-gains-data" type="application/json">'
    if tag_open not in html:
        raise SystemExit("Could not find retain-gains-data script tag in index.html")
    new_html = re.sub(
        r'(<script id="retain-gains-data" type="application/json">).*?(</script>)',
        lambda m: m.group(1) + json.dumps(data, indent=2) + m.group(2),
        html,
        flags=re.DOTALL,
    )
    OUT_HTML.write_text(new_html)
    print(f"Wrote {len(all_outlets)} outlets, {len(gap_list)} open gaps "
          f"({sum(1 for g in gap_list if g['gapType']=='lostNewGain')} lost new gains, "
          f"{sum(1 for g in gap_list if g['gapType']=='unretained')} generic), "
          f"{len(opportunity_list)} open opportunities, {len(rep_summary)} reps.")


if __name__ == "__main__":
    main()
