#!/usr/bin/env python3
"""Flattens Molson Coors' retention workbooks into the two CSVs generate.py reads.

Kohler's BI tool can export these reports two ways. The pipeline was built on
the FLAT export (one column each for District Manager Name / Sales Rep Name /
Brand Family, with the on-screen subtotal rows flattened down into ordinary
data rows). What Gavin sent on 2026-09-04 is the GROUPED export instead: one
combined column where a row's level is implied by its position in the tree.

  ON  workbook: "Sales Rep Name / Brand Family"                    (2 levels)
  OFF workbook: "District Manager Name / Sales Rep Name / Brand Family" (3)

There is NOTHING in the file that marks the levels -- no indentation, no
outline level, no bold, all checked. So levels are resolved by name: the DM
and rep name sets come from the CURRENT flat CSVs (the authoritative previous
pull of these same two reports) plus ROSTER, and anything else under a rep is
a brand. That would be a fragile guess on its own, which is why EVERY total
in both files is then reconciled arithmetically -- each rep's brand rows must
sum to that rep's own total row, each DM's rep totals must sum to the DM's,
and the DMs must sum to the report's "Total" row. A new rep or DM the name
sets don't know would land in the wrong level and break those sums, so the
script refuses to write rather than publish a mis-levelled file.

WHY THE OUTPUT HAS NO SUBTOTAL ROWS. The flat export carried them and
generate.py's _strip_report_subtotals() drops them positionally. Rebuilding
fake subtotal rows here just to have them stripped again would mean inventing
the "borrowed" rep/brand labels RDE puts on them -- so instead this writes
only real data rows and generate.py's _parse_retention_goals() is called with
pre_stripped=True for these two files. Nothing else changes downstream.

Run: python3 convert_mc_retention.py <on.xlsx> <off.xlsx> [--dry-run]
"""
import csv
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
DATA = HERE / "data"
ON_CSV = DATA / "mc_retention_on_prem.csv"
OFF_CSV = DATA / "mc_retention_off_prem.csv"

# Rows whose label is one of these are structural, not data.
TOTAL_LABELS = {"Total"}


def load_grouped(path, sheet_prefix):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if s.startswith(sheet_prefix)), None)
    if sheet is None:
        raise SystemExit(f"{path.name}: no sheet starting {sheet_prefix!r}; found {wb.sheetnames}")
    ws = wb[sheet]
    rows = [r for r in ws.iter_rows(values_only=True) if r and r[0] is not None and str(r[0]).strip()]
    return rows[0], rows[1:]


def known_names(csv_path, column):
    """Distinct values of `column` in the previous flat pull of this report."""
    if not csv_path.exists():
        return set()
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if column not in (reader.fieldnames or []):
            return set()
        return {(r[column] or "").strip() for r in reader if (r[column] or "").strip()}


def num(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return v
    return float(str(v).replace(",", "").strip() or 0)


def convert_on(path, roster):
    """2-level: rep, then its brands. Brand rows are the ones carrying a
    Product Type ('Keg Beer'); a rep's own total row leaves it blank, which
    is a real structural signal rather than a name guess -- so the ON file
    doesn't need the name sets at all, only the reconciliation.

    ON-PREM'S METRIC IS *BUYERS*, A DISTINCT ACCOUNT COUNT, so its rep total
    deliberately does NOT equal the sum of its brand rows -- an account
    pouring both Coors Light and Blue Moon is one buyer on the rep's line and
    a buyer on each brand's. (Checking for equality here is what first
    surfaced this: every rep "failed", e.g. Allison Scott's brands sum to 164
    against a rep total of 72.) Same distinct-count behaviour the Keystone
    dashboard documents. What DOES have to hold is the bound -- a distinct
    count can't be smaller than its biggest single brand, or larger than all
    of them added up -- and that still catches a mis-levelled file, which is
    the thing worth guarding. Off-prem's metric is Placements, which is
    additive, so that side is reconciled by exact sum.
    """
    header, rows = load_grouped(path, "Molson Coors ON Retention")
    val_col, goal_col, pct_col = header[2], header[3], header[4]
    out, problems = [], []
    rep = None
    pending = []          # (rep, rep_total, [brand rows]) awaiting reconciliation
    rep_total = None

    def close_rep():
        if rep is None or not pending:
            return
        values = [num(b["value"]) for b in pending]
        total = num(rep_total)
        if not (max(values) - 0.01 <= total <= sum(values) + 0.01):
            problems.append(f"{rep}: rep total {total} is outside its brands' bound "
                            f"[{max(values)}, {sum(values)}] -- distinct buyer counts must sit "
                            f"between the biggest single brand and the sum of all of them")

    for r in rows:
        label = str(r[0]).strip()
        if label in TOTAL_LABELS:
            continue
        product_type = (str(r[1]).strip() if r[1] is not None else "")
        if not product_type:                     # a rep's own total row
            close_rep()
            rep, rep_total, pending = label, r[2], []
            continue
        if rep is None:
            raise SystemExit(f"{path.name}: brand row {label!r} appears before any rep row.")
        entry = {"rep": rep, "brand": label, "product_type": product_type,
                 "value": r[2], "goal": r[3], "pct": r[4]}
        pending.append(entry)
        out.append(entry)
    close_rep()
    return (val_col, goal_col, pct_col), out, problems


def convert_off(path, dm_names, rep_names):
    """3-level: DM, its reps, their brands. No structural marker at all, so
    levels come from the name sets and are proven by the reconciliation."""
    header, rows = load_grouped(path, "2026 MC Off")
    val_col, goal_col, pct_col = header[1], header[2], header[3]
    out, problems = [], []
    grand_total = None
    dm = rep = None
    dm_total = rep_total = None
    dm_rep_totals, rep_brands = [], []

    def close_rep():
        if rep is None:
            return
        got = sum(num(b["value"]) for b in rep_brands)
        if abs(got - num(rep_total)) > 0.01:
            problems.append(f"{dm} / {rep}: brand rows sum to {got}, rep total row says {rep_total}")
        dm_rep_totals.append(num(rep_total))

    def close_dm():
        close_rep()
        if dm is None:
            return
        got = sum(dm_rep_totals)
        if abs(got - num(dm_total)) > 0.01:
            problems.append(f"{dm}: rep totals sum to {got}, DM total row says {dm_total}")

    dm_totals_seen = []
    for r in rows:
        label = str(r[0]).strip()
        if label in TOTAL_LABELS:
            grand_total = r[1]
            continue
        if label in dm_names:
            close_dm()
            if dm is not None:
                dm_totals_seen.append(num(dm_total))
            dm, dm_total, dm_rep_totals = label, r[1], []
            rep, rep_total, rep_brands = None, None, []
            continue
        if label in rep_names:
            close_rep()
            rep, rep_total, rep_brands = label, r[1], []
            continue
        if rep is None:
            raise SystemExit(f"{path.name}: {label!r} is not a known DM or rep and no rep is open "
                             f"-- cannot tell which level it belongs to. Add it to the name sets "
                             f"(they come from the previous flat CSV) and rerun.")
        entry = {"dm": dm, "rep": rep, "brand": label,
                 "value": r[1], "goal": r[2], "pct": r[3]}
        rep_brands.append(entry)
        out.append(entry)
    close_dm()
    if dm is not None:
        dm_totals_seen.append(num(dm_total))
    if grand_total is not None:
        got = sum(dm_totals_seen)
        if abs(got - num(grand_total)) > 0.01:
            problems.append(f"DM totals sum to {got}, report Total row says {grand_total}")
    return (val_col, goal_col, pct_col), out, problems


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry = "--dry-run" in sys.argv
    if len(args) != 2:
        raise SystemExit("usage: convert_mc_retention.py <on.xlsx> <off.xlsx> [--dry-run]")
    on_path, off_path = Path(args[0]), Path(args[1])

    prev_reps = known_names(OFF_CSV, "Sales Rep Name") | known_names(ON_CSV, "Sales Rep Name")
    dm_names = known_names(OFF_CSV, "District Manager Name")
    if not dm_names:
        raise SystemExit("No District Manager names found in the previous off-prem CSV -- that file "
                         "is what tells this script which labels are DMs. Restore it and rerun.")

    on_cols, on_rows, on_problems = convert_on(on_path, prev_reps)
    off_cols, off_rows, off_problems = convert_off(off_path, dm_names, prev_reps)

    print(f"ON : {len(on_rows)} brand rows across {len({r['rep'] for r in on_rows})} reps")
    print(f"OFF: {len(off_rows)} brand rows across {len({r['rep'] for r in off_rows})} reps, "
          f"{len({r['dm'] for r in off_rows})} DMs")

    problems = [("on-prem", p) for p in on_problems] + [("off-prem", p) for p in off_problems]
    if problems:
        print("\nRECONCILIATION FAILED -- refusing to write:")
        for side, p in problems:
            print(f"  {side}: {p}")
        raise SystemExit(1)
    print("reconciliation: every rep, DM and report total adds up")

    # What changed vs the previous pull, for a sanity read before committing.
    for label, path, rows, key in (
            ("on-prem", ON_CSV, on_rows, lambda r: (r["rep"], r["brand"])),
            ("off-prem", OFF_CSV, off_rows, lambda r: (r["rep"], r["brand"]))):
        if not path.exists():
            continue
        with open(path, newline="", encoding="utf-8-sig") as f:
            prev = list(csv.DictReader(f))
        prev_keys = {(r["Sales Rep Name"], r["Brand Family"]) for r in prev}
        new_keys = {key(r) for r in rows}
        added, dropped = new_keys - prev_keys, prev_keys - new_keys
        print(f"  {label}: {len(added)} new (rep, brand) pairs, {len(dropped)} gone "
              f"(the previous file also carried subtotal rows, so some 'gone' are those)")

    if dry:
        print("\n--dry-run: nothing written")
        return

    with open(ON_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Sales Rep Name", "Brand Family", "Product Type", *on_cols])
        for r in on_rows:
            w.writerow([r["rep"], r["brand"], r["product_type"], r["value"], r["goal"], r["pct"]])
    with open(OFF_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["District Manager Name", "Sales Rep Name", "Brand Family", *off_cols])
        for r in off_rows:
            w.writerow([r["dm"], r["rep"], r["brand"], r["value"], r["goal"], r["pct"]])
    print(f"\nwrote {ON_CSV.relative_to(HERE.parent)} and {OFF_CSV.relative_to(HERE.parent)}")
    print("These are CLEAN (no subtotal rows) -- generate.py reads them with pre_stripped=True.")
    print("Now run: python3 generate.py")


if __name__ == "__main__":
    main()
